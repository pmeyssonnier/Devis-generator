"""
╔══════════════════════════════════════════════════════╗
║  BAG BATTER SRL — interface de chiffrage                                  ║
║  Déposer un métré · apparier les postes · télécharger l'offre              ║
╚══════════════════════════════════════════════════════╝

Interface Streamlit au-dessus de `chiffrage/`. Aucune logique de prix ici :
tout passe par le moteur, pour qu'il n'y ait jamais deux vérités sur un prix.

Ce que cette interface apporte et que la ligne de commande ne pouvait pas :
l'appariement des postes à l'écran. `MAPPING` est à refaire à chaque marché
— les codes appartiennent au pouvoir adjudicateur — et c'était jusqu'ici la
seule étape qui obligeait à éditer du Python.

Lancement local :
    streamlit run streamlit_app.py
"""

import copy
import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory

import streamlit as st

from chiffrage.bibliotheque import (
    ENTREPRISE,
    LOTS,
    MAPPING,
    OUVRAGES,
    OUVRAGES_A_VALIDER,
    PARAMS,
)
from chiffrage.devis_json import lire as lire_devis
from chiffrage.devis_json import serialiser as serialiser_devis
from chiffrage.devis_xlsx import exporter_devis
from chiffrage.controle_prix import analyser
from chiffrage.detection_colonnes import CHAMPS, CHAMPS_REQUIS
from chiffrage.depot_github import (
    ErreurDepot,
    commiter_lexique,
    commiter_parametres,
    commiter_table,
)
from chiffrage.justification_xlsx import exporter_justification
from chiffrage.lexique import (
    DEMOLITION,
    EXPRESSIONS,
    LOCAL,
    SURCOUCHE,
    SYNONYMES,
    adopter_local,
    ajouter_expression,
    ajouter_synonyme,
    est_demolition,
    sans_accents,
    surcouche_en_python,
    vider_surcouche,
)
from chiffrage.gen_metre import generer_metre
from openpyxl.utils import get_column_letter
from chiffrage.metre_io import (
    feuilles_avec_postes,
    lire_metre_complet,
    normaliser_unite,
    remplir_metre,
)
from chiffrage.moteur import (
    calcul_bordereau,
    calibration,
    coefficient_k,
    controle_coherence,
    devis,
    fiche_prix,
    tables_courantes,
)
from chiffrage.parametres import serialiser
from chiffrage.suggestion import (
    normaliser,
    proposer_mapping,
    score,
    suggerer,
)

# menu_items : « Get help » et « Report a bug » pointent par défaut vers
# Streamlit, pas vers nous — un client final qui a un souci de chiffrage n'a
# rien à faire sur un forum Streamlit. Les mettre à None les retire du menu.
# Le texte « About » remplace la mention par défaut de Streamlit.
st.set_page_config(page_title="Chiffrage BAG BATTER", page_icon="🧱",
                    layout="wide",
                    menu_items={"Get help": None, "Report a bug": None,
                                "About": "Outil de chiffrage BAG BATTER SRL — "
                                         "prix unitaires sur métré de marché "
                                         "public."})

# Un score sous ce seuil n'est pas montré comme une suggestion : afficher du
# bruit ferait perdre plus de temps qu'il n'en fait gagner.
SEUIL_SUGGESTION = 0.35
# Au-dessus de ce score, la suggestion est présentée comme solide — ce qui ne
# veut PAS dire juste. Voir chiffrage/suggestion.py.
SEUIL_CONFIANCE = 0.60

SANS_OUVRAGE = "— ne pas chiffrer —"


def _bordereau(params=None):
    """Le bordereau AUX PARAMÈTRES COURANTS.

    Surtout pas de `@st.cache_data` sans clé ici. Le bordereau
    dépendait « du code seul » tant que les coefficients étaient
    figés ; depuis que la barre latérale les règle, un cache sans
    paramètre sert un prix calculé avec d'AUTRES coefficients que
    ceux affichés.

    Le symptôme mesuré, marge portée à 25 % : 185 308 € affichés à
    l'écran, 210 581 € dans le fichier réellement produit. Deux
    vérités sur le même écran, 13 % d'écart, sur un document qui
    part chez un pouvoir adjudicateur.

    Le calcul prend 0,11 ms pour 49 ouvrages. Le cache économisait
    ça, au prix d'un faux prix affiché.
    """
    return calcul_bordereau(params)


def _libelle_ouvrage(code_ouv, bordereau):
    ligne = bordereau[code_ouv]
    return (f"{code_ouv} · {ligne['libelle_ouv']} "
            f"· {ligne['pu_vente']:.2f} €/{ligne['unite_ouv']}")


def _code_du_libelle(libelle):
    """« 40.20 · Étanchéité… · 45,00 €/m2 » -> « 40.20 »."""
    return libelle.split(" · ", 1)[0]


def _lignes_lisibles(lignes, bordereau):
    """Rend les postes d'un devis prêts pour le tableau, et la liste des
    ouvrages perdus en route.

    Ce qui est STOCKÉ est un code — court, stable, c'est lui qui part
    dans le `.json` et qui survit à un changement de libellé. Ce qui est
    MONTRÉ est le libellé complet : personne ne retient que « 40.20 »
    est l'étanchéité bicouche, et un devis chiffré sur un code confondu
    ne se voit qu'au chantier.

    Le libellé est reconstruit à CHAQUE exécution plutôt que conservé :
    il porte le prix de vente, qui bouge dès qu'on touche la marge ou un
    taux horaire. Un libellé figé afficherait un prix périmé et, pire,
    ne correspondrait plus à aucune option de la liste déroulante.
    """
    lisibles, perdus = [], []
    for ligne in lignes:
        code = ligne.get("code_ouv") or _code_du_libelle(
            str(ligne.get("ouvrage") or ""))
        try:
            qte = float(ligne.get("qte") or 0)
        except (TypeError, ValueError):
            qte = 0.0
        if not code:
            # Ligne à peine ajoutée, rien encore choisi : elle reste. La
            # jeter ferait disparaître sous les doigts la ligne qu'on
            # vient de créer.
            lisibles.append({"ouvrage": None, "qte": qte})
        elif code in bordereau:
            lisibles.append({"ouvrage": _libelle_ouvrage(code, bordereau),
                              "qte": qte})
        else:
            perdus.append(code)
    return lisibles, perdus


def _euro(montant):
    """1234.5 -> '1.234,50 €' (convention belge)."""
    return f"{montant:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".") + " €"


def _avertissement_calibration():
    """Rappel affiché sur chaque page : les prix ne sont pas calibrés.

    Une interface soignée donne l'impression d'un outil fini. Le dire une
    fois dans un README ne suffit pas : c'est ici que les documents sont
    produits, donc c'est ici que ça doit être écrit.
    """
    st.warning(
        "**Prix non calibrés.** Les taux horaires et surtout les rendements "
        "(h/unité) sont des ordres de grandeur du marché belge, pas les "
        "chiffres de l'entreprise. Tant qu'ils n'ont pas été relus, les "
        "documents produits ici montrent que la mécanique tourne — "
        "ils ne sont pas prêts à partir chez un client ou une commune.",
        icon="⚠️",
    )


# ══════════════════════════════════════════════════
#  Barre latérale — paramètres de prix
# ══════════════════════════════════════════════════

with st.sidebar:
    st.title("🧱 BAG BATTER")
    st.caption(f"{ENTREPRISE['adresse']} · {ENTREPRISE['cp_ville']}\n\n"
               f"TVA {ENTREPRISE['tva']}")

    st.subheader("Coefficient de vente")
    fg = st.number_input("Frais généraux (%)", 0.0, 100.0,
                         PARAMS["fg"] * 100, 0.5)
    fc = st.number_input("Frais de chantier (%)", 0.0, 100.0,
                         PARAMS["fc"] * 100, 0.5)
    aleas = st.number_input("Aléas (%)", 0.0, 100.0,
                            PARAMS["aleas"] * 100, 0.5)
    marge = st.number_input("Marge (%)", 0.0, 100.0,
                            PARAMS["marge"] * 100, 0.5)

    params = dict(PARAMS, fg=fg / 100, fc=fc / 100,
                   aleas=aleas / 100, marge=marge / 100)
    k = coefficient_k(params)
    st.metric("Coefficient K", f"{k:.4f}", f"{(k - 1) * 100:+.1f} %",
              help="pu_vente = déboursé sec × K")

    if abs(k - coefficient_k()) > 1e-9:
        st.info("Coefficient modifié pour cette session seulement. "
                 "Pour changer la référence, voir l'onglet "
                 "« ⚙️ Paramètres ».", icon="💡")

    anomalies = {c: v for c, v in controle_coherence().items() if v}
    if anomalies:
        st.error(f"Bibliothèque incohérente : {anomalies}", icon="🛑")

    st.divider()
    st.caption(
        f"{len(OUVRAGES)} ouvrages · {len(LOTS)} lots · "
        f"{len(OUVRAGES_A_VALIDER)} rendements jamais validés"
    )



# ══════════════════════════════════════════════════
#  1. Répondre à un métré imposé
# ══════════════════════════════════════════════════

(onglet_metre, onglet_devis, onglet_biblio,
 onglet_lexique, onglet_calib, onglet_params) = st.tabs(
    ["📥 Répondre à un métré", "🧾 Devis client", "📚 Bibliothèque",
     "🔤 Lexique", "🎯 Calibration", "⚙️ Paramètres"]
)


# ────────────────────────────────────────────
# L'onglet « métré » est une FONCTION, pas un bloc `with` : il a besoin
# de sorties anticipées (pas de fichier déposé, fichier illisible).
# `st.stop()` arrêterait le script ENTIER — les trois autres onglets
# ne s'afficheraient plus. `return` ne quitte que cet onglet.
# ────────────────────────────────────────────
def _repondre_a_un_metre(params):
    st.header("Répondre à un métré imposé")
    st.markdown(
        "Dépose le fichier Excel reçu du pouvoir adjudicateur. "
        "L'outil lit les postes, propose une correspondance vers les "
        "ouvrages de la bibliothèque, et te rend **son** fichier avec la "
        "seule colonne des prix unitaires remplie."
    )
    _avertissement_calibration()

    fichier = st.file_uploader("Métré imposé (.xlsx)", type=["xlsx", "xlsm"])

    if fichier is None:
        st.info(
            "Pas encore de métré ? Le bouton ci-dessous en fabrique un "
            "fictif (49 postes, 10 lots) pour essayer l'outil.",
            icon="💡",
        )
        if st.button("Générer un métré d'entraînement"):
            with TemporaryDirectory() as tmp:
                chemin = Path(tmp) / "METRE_entrainement.xlsx"
                generer_metre(str(chemin))
                st.download_button(
                    "⬇️ Télécharger le métré d'entraînement",
                    chemin.read_bytes(),
                    file_name="METRE_CSC_2026-TP-0147_Schaerbeek.xlsx",
                    mime="application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet",
                )
        return

    b = _bordereau(params)

    with TemporaryDirectory() as tmp:
        chemin_metre = Path(tmp) / fichier.name
        chemin_metre.write_bytes(fichier.getvalue())

        # ── Choix des feuilles ────────────────────────
        # Un métré réel se répartit souvent en « Lot 01 », « Lot 02 »…
        # plus un « Récapitulatif » qui REPREND les mêmes codes :
        # le traiter ferait compter les postes deux fois.
        try:
            inventaire = [f for f in feuilles_avec_postes(str(chemin_metre))
                           if f["nb_postes"]]
        except Exception as err:
            st.error(f"Lecture impossible : {err}", icon="🛑")
            return

        if not inventaire:
            st.error(
                "Aucune feuille ne contient de postes reconnaissables. "
                "L'outil cherche une colonne de codes (B) et une colonne "
                "de quantités (F).",
                icon="🛑")
            return

        if len(inventaire) > 1:
            st.markdown("**Feuilles à traiter**")
            retenues = []
            for feuille in inventaire:
                marque = " · récapitulatif présumé" if feuille["recapitulatif"] else ""
                if st.checkbox(
                    f"{feuille['nom']} — {feuille['nb_postes']} postes{marque}",
                    value=not feuille["recapitulatif"],
                    key=f"feuille_{feuille['nom']}",
                ):
                    retenues.append(feuille["nom"])
            if not retenues:
                st.warning("Aucune feuille retenue.", icon="⚠️")
                return
            st.caption(
                "Un récapitulatif reprend les codes des lots : le cocher "
                "ferait compter chaque poste deux fois. Les doublons sont "
                "signalés, jamais additionnés."
            )
        else:
            retenues = [inventaire[0]["nom"]]

        # ── Correspondance des colonnes ────────────────
        # Les colonnes étaient codées en dur : chez un autre pouvoir
        # adjudicateur, l'outil ne lisait rien — ou écrivait le prix
        # dans la mauvaise colonne, ce qui est pire. La détection
        # PROPOSE, l'humain valide avant tout chiffrage.
        detectees = next(
            (f["colonnes"] for f in inventaire if f["nom"] == retenues[0]),
            None)
        colonnes = dict(detectees or {})

        incertaines = [c for c in CHAMPS_REQUIS if c not in colonnes]
        libelle_champ = {
            "code": "Code du poste", "designation": "Désignation",
            "nature": "Nature", "unite": "Unité", "quantite": "Quantité",
            "pu": "Prix unitaire (colonne à remplir)", "montant": "Montant",
        }
        resume = " · ".join(
            f"{libelle_champ[c]} = {get_column_letter(colonnes[c])}"
            for c in CHAMPS_REQUIS if c in colonnes)

        with st.expander(
            ("⚠️ Colonnes à confirmer" if incertaines
             else f"🔎 Colonnes détectées — {resume}"),
            expanded=bool(incertaines),
        ):
            st.caption(
                "Corriger ici si la détection s'est trompée. Un prix "
                "écrit dans la mauvaise colonne rendrait l'offre "
                "silencieusement fausse."
            )
            lettres = [get_column_letter(i) for i in range(1, 31)]
            gauche, droite = st.columns(2)
            for i, champ in enumerate(CHAMPS):
                zone = gauche if i % 2 == 0 else droite
                actuelle = colonnes.get(champ)
                options = ["— absente —"] + lettres
                index = (lettres.index(get_column_letter(actuelle)) + 1
                          if actuelle and actuelle <= 30 else 0)
                choix = zone.selectbox(
                    libelle_champ[champ] + (" *" if champ in CHAMPS_REQUIS
                                             else ""),
                    options, index=index, key=f"col_{champ}")
                if choix == "— absente —":
                    colonnes.pop(champ, None)
                else:
                    colonnes[champ] = lettres.index(choix) + 1

            manquants = [libelle_champ[c] for c in CHAMPS_REQUIS
                          if c not in colonnes]
            if manquants:
                st.error(
                    "Champs indispensables non renseignés : "
                    + ", ".join(manquants)
                    + ". Sans eux, impossible de savoir quoi chiffrer, où "
                      "écrire, ni de vérifier les unités.",
                    icon="🛑")

        if any(c not in colonnes for c in CHAMPS_REQUIS):
            return

        try:
            lecture = lire_metre_complet(str(chemin_metre), retenues, colonnes)
        except Exception as err:
            st.error(f"Lecture impossible : {err}", icon="🛑")
            return
        postes = lecture["postes"]

        # Une ligne non lue est un poste ABSENT de l'offre — plus
        # discret qu'un poste sans prix, et tout aussi disqualifiant.
        if lecture["anomalies"]:
            bloquantes = [a for a in lecture["anomalies"]
                           if a["genre"] not in ("quantite_formule",
                                                  "quantite_nulle")]
            texte = "\n".join(
                f"- ligne **{a['ligne']}** · `{a['code']}` — {a['detail']}"
                for a in lecture["anomalies"])
            if bloquantes:
                st.error(
                    f"**{len(bloquantes)} ligne(s) du métré n'ont pas pu "
                    f"être lues — ces postes ne seront PAS chiffrés.**\n\n"
                    + texte, icon="🛑")
            else:
                st.warning(
                    f"**{len(lecture['anomalies'])} ligne(s) à "
                    f"vérifier.**\n\n" + texte, icon="⚠️")

        if not postes:
            st.error(
                "Aucun poste lu. Le fichier doit porter, par feuille, une "
                "colonne de codes au format NN.NN et une colonne de "
                "quantités numériques.",
                icon="🛑",
            )
            return

        # ── Appariement : recalculé si le fichier change ──────────────
        # La version du lexique entre dans la signature : ajouter un
        # synonyme dans l'onglet Lexique doit refaire l'appariement,
        # sinon l'essai n'a aucun effet visible ici.
        # Le contenu, pas le nom ni la taille : deux métrés différents
        # peuvent porter le même nom et peser pareil.
        signature = (
            hashlib.sha256(fichier.getvalue()).hexdigest(),
            tuple(retenues),
            tuple(sorted(colonnes.items())),
            st.session_state.get("lexique_version", 0),
        )
        # L'appariement est refait quand la signature change — mais
        # aussi quand l'un des deux états manque. Ils étaient lus en
        # accès direct : une session dont la signature avait survécu
        # sans son appariement (interruption au milieu du calcul,
        # session restaurée à moitié après un rafraîchissement) levait
        # un KeyError et affichait un bandeau rouge, là où il suffisait
        # de recalculer.
        etat_incomplet = ("proposition" not in st.session_state
                           or "mapping" not in st.session_state)
        if st.session_state.get("signature") != signature or etat_incomplet:
            proposition = proposer_mapping(
                postes, b, mapping_connu=MAPPING, seuil=SEUIL_SUGGESTION)
            st.session_state.signature = signature
            st.session_state.proposition = proposition
            st.session_state.mapping = {
                code: infos["code_ouv"]
                for code, infos in proposition.items()
            }

        proposition = st.session_state.proposition
        mapping = st.session_state.mapping

        # ── Tableau de contrôle ────────────────────────────────
        a_revoir = [
            p for p in postes
            if mapping.get(p["code"]) is None
            or proposition[p["code"]]["origine"] != "connu"
            and proposition[p["code"]]["score"] < SEUIL_CONFIANCE
        ]
        chiffres = [p for p in postes if mapping.get(p["code"])]

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Postes lus", len(postes))
        c2.metric("Appariés", len(chiffres))
        c3.metric("À revoir", len(a_revoir),
                    delta=None if not a_revoir else f"{len(a_revoir)} à confirmer",
                    delta_color="off" if not a_revoir else "inverse")
        total = sum(
            b[mapping[p["code"]]]["pu_vente"] * p["quantite"]
            for p in chiffres
        )
        c4.metric("Total estimé", _euro(total))

        # ── Postes à revoir ────────────────────────────────────
        st.subheader("Postes à revoir")
        if a_revoir:
            st.caption(
                "Suggestions issues du seul libellé. Un score élevé veut dire "
                "« regarde ici d'abord », jamais « c'est bon » : "
                "« dépose » et « pose » ne diffèrent que d'une lettre pour un "
                "algorithme, et sont opposés sur un chantier."
            )
        else:
            st.success("Chaque poste porte une correspondance.", icon="✅")

        tout_revoir = st.checkbox(
            "Revoir aussi les postes déjà appariés",
            help="Par défaut, seules postes incertains sont proposésés.",
        )

        a_afficher = postes if tout_revoir else [
            p for p in postes if p["code"] in {q["code"] for q in a_revoir}
        ]

        for poste in a_afficher:
            code = poste["code"]
            infos = proposition[code]
            unite_poste = normaliser_unite(poste["unite"])

            # Unité éliminatoire : un ouvrage dans une autre unité
            # ne peut pas chiffrer ce poste, quelle que soit la ressemblance
            # des libellés. On ne les propose donc pas.
            compatibles = [
                code_ouv for code_ouv, ligne in b.items()
                if normaliser_unite(ligne["unite_ouv"]) == unite_poste
            ]
            options = [SANS_OUVRAGE] + sorted(compatibles)

            actuel = mapping.get(code)
            index = options.index(actuel) if actuel in options else 0

            candidats = ", ".join(
                f"{c} ({s:.2f})" for c, s in infos["candidats"][:3]
            ) or "aucun candidat dans cette unité"

            marque = {"connu": "🔒", "suggere": "🟡", "aucun": "⚪"}[
                infos["origine"]
            ]

            col_gauche, col_droite = st.columns([3, 2])
            with col_gauche:
                st.markdown(
                    f"**{marque} {code} · {poste['designation']}**  \n"
                    f"`{poste['unite']}` × {poste['quantite']:g}"
                )
                st.caption(f"Candidats : {candidats}")
            with col_droite:
                choix = st.selectbox(
                    f"Ouvrage pour {code}",
                    options,
                    index=index,
                    format_func=lambda c: (
                        SANS_OUVRAGE if c == SANS_OUVRAGE
                        else _libelle_ouvrage(c, b)
                    ),
                    key=f"map_{code}",
                    label_visibility="collapsed",
                )
                mapping[code] = None if choix == SANS_OUVRAGE else choix

            if not compatibles:
                st.caption(
                    f"⚠️ Aucun ouvrage en « {poste['unite']} » : il faut le "
                    f"créer dans la bibliothèque, ce poste ne peut pas être "
                    f"chiffré."
                )
            st.divider()

        st.session_state.mapping = mapping

        # ── Production de l'offre ──────────────────────────────
        st.subheader("Produire l'offre")

        mapping_effectif = {poste: ouv for poste, ouv in mapping.items() if ouv}
        vides = [p["code"] for p in postes if not mapping.get(p["code"])]

        if vides:
            st.error(
                f"**{len(vides)} postes sans prix : offre irrégulière.** "
                f"Un seul poste vide suffit à faire rejeter l'offre "
                f"(art. 76 AR 18/04/2017). Postes concernés : "
                + ", ".join(vides),
                icon="🛑",
            )

        col_a, col_b = st.columns(2)
        with col_a:
            tva = st.radio("TVA", [0.21, 0.06],
                            format_func=lambda t: f"{t * 100:.0f} %",
                            horizontal=True,
                            help="21 % en marché public. 6 % : logement privé "
                                  "de plus de 10 ans, usage privé, consommateur final.")
        with col_b:
            st.write("")
            chiffrer = st.button("⚙️ Chiffrer et produire l'offre",
                                  type="primary", width="stretch",
                                  disabled=not mapping_effectif)

        if chiffrer:
            with TemporaryDirectory() as tmp:
                entree = Path(tmp) / fichier.name
                entree.write_bytes(fichier.getvalue())
                sortie = Path(tmp) / f"OFFRE_{Path(fichier.name).stem}.xlsx"
                rapport = remplir_metre(
                    str(entree), str(sortie), mapping=mapping,
                    params=params, tva=tva, feuilles=retenues,
                    colonnes=colonnes,
                )
                octets = sortie.read_bytes()

            st.session_state.offre = {
                "octets": octets,
                "nom": f"OFFRE_{Path(fichier.name).stem}"
                        f"_{datetime.now():%Y%m%d_%H%M}.xlsx",
                "rapport": rapport,
            }

        if st.session_state.get("offre"):
            offre = st.session_state.offre
            rapport = offre["rapport"]

            st.success(
                f"{len(rapport['chiffres'])} postes chiffrés · "
                f"{_euro(rapport['total_ht'])} HTVA · "
                f"{rapport['heures_mo']:.0f} h de main-d'œuvre",
                icon="✅",
            )

            if rapport["ecarts_unite"]:
                st.error(
                    "Écarts d'unité — ces postes ne sont **pas** chiffrés :\n"
                    + "\n".join(
                        f"- `{e['code']}` métré en « {e['unite_metre']} » mais "
                        f"`{e['code_ouv']}` est en « {e['unite_ouvrage']} »"
                        for e in rapport["ecarts_unite"]
                    ),
                    icon="🛑",
                )

            st.download_button(
                "⬇️ Télécharger l'offre à renvoyer",
                offre["octets"],
                file_name=offre["nom"],
                mime="application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet",
                type="primary",
            )
            st.caption(
                "C'est le fichier du pouvoir adjudicateur, avec la seule "
                "colonne des prix unitaires remplie : ses quantités et ses "
                "formules sont intactes."
            )

            # ── Contrôle des prix avant dépôt ──────────────
            st.divider()
            st.subheader("Contrôle des prix")
            st.caption(
                "Deux risques opposés : trop bas, l'offre est écartée pour "
                "prix anormalement bas (art. 36 AR 18/04/2017) ou le "
                "chantier s'exécute à perte ; trop haut, le marché est "
                "perdu. Ces contrôles regardent l'offre depuis l'intérieur "
                "de l'entreprise — les prix des concurrents, personne ne "
                "les a au moment de déposer."
            )

            lignes_controle = [
                {"code_ouv": c["code_ouv"], "qte": c["quantite"],
                 "pu_vente": c["pu"]}
                for c in rapport["chiffres"]
            ]

            rabais = st.slider(
                "Rabais commercial simulé", 0.0, 30.0, 0.0, 0.5,
                format="%.1f %%",
                help="Le seuil au-delà duquel l'offre cesse de couvrir "
                      "ses coûts est calculé juste en dessous.",
            ) / 100.0

            controle = analyser(lignes_controle, b, params=params,
                                 rabais=rabais)
            ind = controle["indicateurs"]

            m1, m2, m3 = st.columns(3)
            m1.metric(
                "Par heure travaillée", f"{ind['par_heure']:.2f} €",
                f"{ind['par_heure'] - ind['plancher']:+.2f} € / plancher",
                delta_color="normal" if ind["couvre"] else "inverse",
                help="Matériaux et matériel payés, ce que l'offre laisse "
                      "par heure de main-d'œuvre. Sous le plancher, chaque "
                      "heure travaillée coûte de l'argent.",
            )
            m2.metric("Rabais maximal",
                       f"{ind['rabais_maximal'] * 100:.1f} %",
                       help="Au-delà, l'offre ne couvre plus ses coûts.")
            m3.metric("Top 3 des postes",
                       f"{ind['concentration_top3'] * 100:.0f} %",
                       help="Part du total portée par trois postes.")

            for alerte in controle["alertes"]:
                texte = f"**{alerte['titre']}**  \n{alerte['detail']}"
                if alerte["niveau"] == "critique":
                    st.error(texte, icon="🛑")
                elif alerte["niveau"] == "attention":
                    st.warning(texte, icon="⚠️")
                else:
                    st.info(texte, icon="💡")
            if not controle["alertes"]:
                st.success("Aucun point d'attention.", icon="✅")

            # ── Dossier de justification ───────────────
            with st.expander("📄 Dossier de justification de prix"):
                st.markdown(
                    "Si le pouvoir adjudicateur conteste un prix, il doit "
                    "demander une justification écrite avant d'écarter "
                    "l'offre — et le délai de réponse est court. Ce "
                    "dossier contient, pour chaque poste visé, la "
                    "décomposition qui a **servi** à établir l'offre : "
                    "ressources, quantités, prix d'achat, déboursés, "
                    "coefficient. C'est ce qui la rend crédible."
                )
                defaut = [p["code_ouv"] for p in controle["postes"][:3]]
                a_justifier = st.multiselect(
                    "Postes à justifier",
                    sorted({p["code_ouv"] for p in controle["postes"]}),
                    default=defaut,
                    format_func=lambda c: _libelle_ouvrage(c, b),
                )
                if a_justifier:
                    with TemporaryDirectory() as tmp:
                        cible = Path(tmp) / "justification.xlsx"
                        exporter_justification(
                            a_justifier, str(cible),
                            marche={"reference": Path(fichier.name).stem},
                            params=params,
                        )
                        octets_just = cible.read_bytes()
                    st.download_button(
                        "⬇️ Télécharger le dossier",
                        octets_just,
                        file_name=f"JUSTIFICATION_{datetime.now():%Y%m%d_%H%M}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument."
                              "spreadsheetml.sheet",
                    )
                    st.caption(
                        "La lettre d'accompagnement est le premier onglet : "
                        "à relire et à signer avant envoi."
                    )

        # ── Réutiliser la correspondance ─────────────────────────────
        with st.expander("♻️ Réutiliser cette correspondance"):
            st.markdown(
                "Une commune réutilise ses propres codes d'un marché à "
                "l'autre. Garde ce fichier : au prochain métré de la même "
                "commune, la correspondance sera déjà faite."
            )
            st.download_button(
                "⬇️ Télécharger la correspondance (.json)",
                json.dumps(mapping_effectif, indent=2, ensure_ascii=False),
                file_name=f"MAPPING_{Path(fichier.name).stem}.json",
                mime="application/json",
            )
            repris = st.file_uploader("Reprendre une correspondance (.json)",
                                        type=["json"], key="up_map")
            for anomalie in st.session_state.get("mapping_anomalies", []):
                st.warning(anomalie, icon="⚠️")
            # BOUCLE DE RERUN À NE PAS RÉINTRODUIRE.
            # `repris` reste non-None à CHAQUE réexécution du script tant
            # que le fichier est déposé : un st.rerun() inconditionnel
            # ici relance le script sans fin. L'app ne répond plus, et
            # recharger la page n'y change rien puisque le fichier est
            # toujours là — il faut redémarrer le serveur.
            # On ne traite donc le dépôt QU'UNE FOIS, repéré par
            # l'empreinte de son contenu.
            depot_deja_lu = st.session_state.get("mapping_importe")
            empreinte = (hashlib.sha256(repris.getvalue()).hexdigest()
                          if repris is not None else None)
            if repris is not None and empreinte != depot_deja_lu:
                try:
                    charge = json.loads(repris.getvalue().decode("utf-8"))
                    inconnus = set(charge.values()) - set(b) - {None}
                    # Rangé dans l'état plutôt qu'affiché : le st.rerun()
                    # plus bas effaçait ce message avant qu'il soit lu, et
                    # la correspondance repartait amputée sans que rien ne
                    # le dise.
                    st.session_state.mapping_anomalies = (
                        ["Codes d'ouvrage inconnus, ignorés : "
                          + ", ".join(sorted(inconnus))] if inconnus else [])
                    st.session_state.mapping = {
                        poste: ouv for poste, ouv in charge.items()
                        if ouv in b or ouv is None
                    }
                    st.session_state.mapping_importe = empreinte
                    st.rerun()
                except Exception as err:
                    # Marqué lu malgré l'échec : sans ça, un fichier
                    # illisible reposerait la question à chaque rerun.
                    st.session_state.mapping_importe = empreinte
                    st.error(f"Fichier illisible : {err}", icon="🛑")


with onglet_metre:
    _repondre_a_un_metre(params)


# ══════════════════════════════════════════════════
#  2. Devis client
# ══════════════════════════════════════════════════

# Valeurs d'exemple, pas valeurs par défaut utiles : elles montrent la forme
# attendue de chaque champ et sont faites pour être écrasées. Elles passent par
# session_state — et non par l'argument `value=` des widgets — parce que
# reprendre un devis doit pouvoir les remplacer, ce qu'un `value=` empêcherait.
_DEVIS_EXEMPLE = {
    "devis_objet": "Rénovation de la façade arrière",
    "devis_reference": f"{date.today():%Y}-042",
    "devis_chantier": "Avenue Ernest Renan 62, 1030 Schaerbeek",
    "devis_client": "M. et Mme Dupont\nRue de l'Église 12\n1030 Schaerbeek",
    "tva_devis": 0.06,
    "lignes_devis": [{"code_ouv": "40.20", "qte": 22.0},
                      {"code_ouv": "40.30", "qte": 22.0}],
    # Les postes sont rangés par code et affichés par libellé : voir
    # _lignes_lisibles(). L'exemple est donc écrit en codes.
}


def _reprendre_un_devis(codes_connus):
    """Recharge un devis enregistré dans les champs de l'onglet.

    À appeler AVANT que les widgets concernés soient créés : Streamlit
    refuse qu'on écrive dans l'état d'un widget déjà instancié dans la
    même exécution.
    """
    depose = st.file_uploader("Reprendre un devis enregistré (.json)",
                               type=["json"], key="up_devis")
    # Même piège que pour la correspondance de métré : `depose` reste
    # non-None à chaque réexécution tant que le fichier est là. Un
    # st.rerun() inconditionnel bouclerait sans fin. On ne traite le
    # dépôt qu'une fois, repéré par l'empreinte de son contenu.
    deja_lu = st.session_state.get("devis_importe")
    empreinte = (hashlib.sha256(depose.getvalue()).hexdigest()
                  if depose is not None else None)
    if depose is None or empreinte == deja_lu:
        return
    st.session_state.devis_importe = empreinte
    try:
        charge, anomalies = lire_devis(depose.getvalue(), codes_connus)
    except Exception as err:
        # Marqué lu malgré l'échec : sans ça, un fichier illisible
        # reposerait la question à chaque rerun. Pas de rerun sur ce
        # chemin, donc le message peut s'afficher directement.
        st.session_state.pop("devis_anomalies", None)
        st.error(f"Fichier illisible : {err}", icon="🛑")
        return
    # Rangé dans l'état, PAS affiché ici : le st.rerun() qui suit
    # rejouerait le script depuis le début et effacerait l'avertissement
    # avant que quiconque l'ait lu. C'est l'appelant qui l'affiche.
    st.session_state.devis_anomalies = anomalies
    for champ in ("objet", "reference", "chantier", "client"):
        st.session_state[f"devis_{champ}"] = charge[champ]
    st.session_state.tva_devis = charge["tva"]
    st.session_state.lignes_devis = charge["lignes"]
    # L'éditeur garde ses modifications sous sa propre clé et les
    # REAPPLIQUE aux nouvelles données au rerun : sans cet oubli, les
    # lignes ajoutées à la main au devis précédent reviendraient se
    # coller au devis qu'on vient de charger.
    st.session_state.pop("editeur_devis", None)
    st.rerun()


with onglet_devis:
    st.header("Devis client")
    _avertissement_calibration()

    b = _bordereau(params)
    for cle, valeur in _DEVIS_EXEMPLE.items():
        # deepcopy : la liste de postes de l'exemple est partagée par le
        # module. La poser telle quelle dans l'état ferait éditer l'exemple
        # lui-même, et le devis suivant démarrerait sur les postes du
        # précédent.
        st.session_state.setdefault(cle, copy.deepcopy(valeur))

    with st.expander("📂 Reprendre un devis"):
        _reprendre_un_devis(b)
        for anomalie in st.session_state.get("devis_anomalies", []):
            st.warning(anomalie, icon="⚠️")
        st.caption(
            "Le fichier `.json` téléchargé en bas de cet onglet se redépose "
            "ici : en-tête, TVA et postes reviennent tels quels. Les **prix "
            "sont recalculés** aux valeurs actuelles de la bibliothèque — un "
            "devis rouvert est un devis à réémettre, pas une archive."
        )

    col_g, col_d = st.columns([2, 1])

    with col_g:
        objet = st.text_input("Objet", key="devis_objet")
        reference = st.text_input("Référence", key="devis_reference")
        chantier = st.text_input("Chantier", key="devis_chantier")
        client = st.text_area("Client", height=90, key="devis_client")
    with col_d:
        tva_devis = st.radio("TVA", [0.06, 0.21],
                              format_func=lambda t: f"{t * 100:.0f} %",
                              key="tva_devis")
        st.caption(
            "6 % : logement privé de plus de dix ans, usage principalement "
            "privé, facturation au consommateur final. Dans le doute, 21 %."
        )

    st.subheader("Postes")
    lisibles, perdus = _lignes_lisibles(st.session_state.lignes_devis, b)
    for code in perdus:
        st.warning(f"L'ouvrage « {code} » n'existe plus dans la "
                    "bibliothèque : poste retiré du devis.", icon="⚠️")

    # `options` en toutes lettres, pas en codes : la liste déroulante d'une
    # colonne n'accepte pas de `format_func`, contrairement aux selectbox du
    # reste de l'app. Montrer « 40.20 » demanderait de connaître par cœur une
    # cinquantaine de codes — et une erreur de choix ne se verrait qu'au
    # chantier. Le libellé porte aussi le prix unitaire : il sert de contrôle
    # de vraisemblance au moment même où on choisit.
    edite = st.data_editor(
        lisibles,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "ouvrage": st.column_config.SelectboxColumn(
                "Ouvrage", options=[_libelle_ouvrage(c, b) for c in sorted(b)],
                required=True, width="large"),
            "qte": st.column_config.NumberColumn(
                "Quantité", min_value=0.0, step=0.5, format="%.2f"),
        },
        key="editeur_devis",
    )
    st.caption("Onglet **📚 Bibliothèque** : la liste complète des ouvrages, "
                "leur code, leur unité et leur prix.")

    # Rangé en codes : un libellé porte le prix du jour, un code non. On
    # garde TOUTES les lignes, y compris celles en cours de saisie — filtrer
    # ici effacerait la ligne dont la quantité n'est pas encore tapée à la
    # première réexécution venue.
    st.session_state.lignes_devis = [
        {"code_ouv": (_code_du_libelle(ligne["ouvrage"])
                       if ligne.get("ouvrage") else None),
          "qte": ligne.get("qte")}
        for ligne in edite
    ]

    lignes = [
        (_code_du_libelle(ligne["ouvrage"]), float(ligne["qte"]))
        for ligne in edite
        if ligne.get("ouvrage") and ligne.get("qte")
    ]

    if lignes:
        d = devis(objet or "Devis", lignes, tva=tva_devis, params=params)

        c1, c2, c3 = st.columns(3)
        c1.metric("Total HTVA", _euro(d["total_ht"]))
        c2.metric(f"TVA {tva_devis * 100:.0f} %", _euro(d["montant_tva"]))
        c3.metric("Total TVAC", _euro(d["total_ttc"]))
        st.caption(f"Main-d'œuvre : {d['heures_mo']:.1f} h "
                    f"({d['jours_homme']:.1f} jours-homme)")

        tableau = [
            {
                "Code": poste["code_ouv"],
                "Désignation": poste["libelle_ouv"],
                "Un.": poste["unite_ouv"],
                "Qté": poste["qte"],
                "PU HTVA": poste["pu_vente"],
                "Montant HTVA": poste["montant"],
            }
            for poste in d["lignes"]
        ]
        st.dataframe(tableau, width="stretch", hide_index=True)

        with TemporaryDirectory() as tmp:
            cible = Path(tmp) / f"DEVIS_{reference}.xlsx"
            exporter_devis(d, str(cible), client=client, chantier=chantier,
                            reference=reference)
            octets = cible.read_bytes()

        col_xlsx, col_json = st.columns(2)
        col_xlsx.download_button(
            "⬇️ Télécharger le devis", octets,
            file_name=f"DEVIS_{reference}_{datetime.now():%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet",
            type="primary", width="stretch",
        )
        # Le .xlsx part au client, le .json revient ici. Sans lui, corriger
        # un devis de la semaine passée voulait dire tout ré-encoder — et un
        # simple rafraîchissement de la page suffisait à tout perdre.
        col_json.download_button(
            "💾 Enregistrer pour modifier plus tard (.json)",
            serialiser_devis(objet, reference, chantier, client, tva_devis,
                              lignes),
            file_name=f"DEVIS_{reference}.json",
            mime="application/json", width="stretch",
        )
    else:
        st.info("Ajoute au moins un poste.", icon="💡")

# ══════════════════════════════════════════════════
#  3. Bibliothèque
# ══════════════════════════════════════════════════

with onglet_biblio:
    st.header("Bibliothèque de prix")
    _avertissement_calibration()

    b = _bordereau(params)
    a_valider = set(OUVRAGES_A_VALIDER)

    # Faire défiler 49 lignes au doigt pour retrouver un code n'est pas une
    # recherche. Le filtre porte sur le code, le lot, la désignation et
    # l'unité à la fois : on tape ce dont on se souvient, quoi que ce soit.
    requete = st.text_input(
        "Chercher un ouvrage",
        placeholder="façade · enduit · 40.20 · toiture…",
        key="biblio_recherche")
    mots = sans_accents(requete).lower().split()

    def _texte(code, ligne):
        return sans_accents(
            f"{code} {ligne['lot']} {LOTS.get(ligne['lot'], '')} "
            f"{ligne['libelle_ouv']} {ligne['unite_ouv']}").lower()

    retenus = [(code, ligne) for code, ligne in sorted(b.items())
                if all(mot in _texte(code, ligne) for mot in mots)]

    lignes = [
        {
            "Code": code,
            "Lot": f"{ligne['lot']} — {LOTS.get(ligne['lot'], '')}",
            "Désignation": ligne["libelle_ouv"],
            "Un.": ligne["unite_ouv"],
            "h/unité": ligne["heures_mo"],
            "Déboursé": ligne["debourse_sec"],
            "PU vente": ligne["pu_vente"],
            "À valider": "⚠️" if code in a_valider else "",
        }
        for code, ligne in retenus
    ]
    if lignes:
        st.dataframe(lignes, width="stretch", hide_index=True, height=420)
        if mots:
            st.caption(f"{len(lignes)} ouvrage(s) sur {len(b)}.")
    else:
        # Le filtre littéral ne connaît que les mots de la bibliothèque.
        # « carrelage mural » n'y est pas : l'ouvrage s'appelle « faïence ».
        # On repasse donc par le moteur d'appariement, lexique compris —
        # le même qui apparie les métrés, pas une seconde vérité.
        st.warning("Aucun libellé ne contient tous ces mots.", icon="⚠️")
        proches = sorted(((score(requete, ligne["libelle_ouv"]), code)
                           for code, ligne in b.items()), reverse=True)[:5]
        st.caption(
            "Les plus proches selon le lexique — « carrelage mural » mène "
            "à « faïence ». Un score n'est pas une probabilité : il dit "
            "« regarde ici d'abord », jamais « c'est celui-là »."
        )
        st.dataframe(
            [{"Code": code, "Désignation": b[code]["libelle_ouv"],
               "Un.": b[code]["unite_ouv"], "PU vente": b[code]["pu_vente"],
               "Score": s}
              for s, code in proches],
            width="stretch", hide_index=True,
            column_config={"Score": st.column_config.NumberColumn(
                format="%.2f")})

    st.subheader("Fiche de justification de prix")
    st.caption(
        "La pièce à produire si un pouvoir adjudicateur conteste un prix "
        "jugé anormal (art. 36 AR 18/04/2017)."
    )
    code_fiche = st.selectbox("Ouvrage", sorted(b),
                                 format_func=lambda c: _libelle_ouvrage(c, b))
    st.code(fiche_prix(code_fiche), language="text")

    # ══════════════════════════════════════
    #  Atelier de correction — la séance de calibration
    # ══════════════════════════════════════
    st.divider()
    st.subheader("✏️ Corriger les prix et les rendements")
    st.markdown(
        "Les deux tables que le chef d'entreprise est seul à pouvoir "
        "corriger : ce que coûte une heure, et combien d'heures prend un "
        "mètre carré. **L'effet sur la calibration s'affiche en direct**, "
        "avant tout enregistrement."
    )

    if "tables_editees" not in st.session_state:
        st.session_state.tables_editees = copy.deepcopy(tables_courantes())
    tables = st.session_state.tables_editees
    origine = tables_courantes()

    # ── Corriger une valeur ────────────────────
    # Un tableur de 49 lignes ne se remplit pas au doigt : taper dans
    # une cellule, valider la saisie, en sortir — chaque geste rate une
    # fois sur deux sur un téléphone, et une correction qui ne « prend »
    # pas ne se voit pas. On corrige donc UNE valeur à la fois, en trois
    # gestes sûrs : choisir, saisir, appliquer.
    quoi = st.radio(
        "Que corriger ?",
        ["Un taux horaire ou un prix d'achat", "Un rendement (h/unité)"],
        horizontal=True, label_visibility="collapsed")

    origine_res = origine["ressources_par_code"]
    origine_ouv = origine["ouvrages_par_code"]

    if quoi.startswith("Un taux"):
        st.caption(
            "Les taux MO doivent être le **coût entreprise complet** — "
            "salaire, ONSS patronale, pécule, jours fériés, assurance loi, "
            "déplacements, EPI. Surtout pas le brut."
        )
        choix = st.selectbox(
            "Ressource",
            [r["code_res"] for r in tables["ressources"]],
            format_func=lambda c: (
                f"{c} · {origine_res[c]['libelle_res']} — "
                f"{tables['ressources_par_code'][c]['pu_res']:.2f} €"
                f"/{origine_res[c]['unite_res']}"),
            key="corr_res")
        actuelle = tables["ressources_par_code"][choix]
        if actuelle.get("note"):
            st.caption(f"🛈 {actuelle['note']}")

        col_val, col_btn = st.columns([2, 1])
        valeur = col_val.number_input(
            f"Nouveau prix ({origine_res[choix]['unite_res']})",
            min_value=0.0, value=float(actuelle["pu_res"]), step=0.5,
            format="%.2f", key="corr_res_valeur")
        col_btn.write("")
        if col_btn.button("Appliquer", width="stretch", key="corr_res_ok"):
            actuelle["pu_res"] = float(valeur)
            tables["ressources_par_code"] = {r["code_res"]: r
                                              for r in tables["ressources"]}
            st.rerun()
        if abs(actuelle["pu_res"] - origine_res[choix]["pu_res"]) > 1e-9:
            st.caption(f"Corrigé : {origine_res[choix]['pu_res']:.2f} € → "
                        f"**{actuelle['pu_res']:.2f} €**")
    else:
        st.caption(
            "Le rendement est la **seule donnée non achetable** : elle "
            "vient de l'expérience du chantier, et pèse la majeure partie "
            "du déboursé. C'est ici que la calibration se joue."
        )
        # La copie de travail, pas la constante du module : le ⚠️ doit
        # disparaître à l'écran dès qu'on lève le doute, sans attendre le
        # redéploiement.
        a_valider = set(tables["ouvrages_a_valider"])
        indices_mo = [
            i for i, c in enumerate(tables["composition"])
            if origine_res[c["code_res"]]["type_res"] == "MO"]
        choix = st.selectbox(
            "Ouvrage et main-d'œuvre",
            indices_mo,
            format_func=lambda i: (
                f"{'⚠️ ' if tables['composition'][i]['code_ouv'] in a_valider else ''}"
                f"{tables['composition'][i]['code_ouv']} · "
                f"{origine_ouv[tables['composition'][i]['code_ouv']]['libelle_ouv'][:44]}"
                f" — {tables['composition'][i]['qte_res']:.3f} h/"
                f"{origine_ouv[tables['composition'][i]['code_ouv']]['unite_ouv']}"),
            key="corr_rend")
        ligne = tables["composition"][choix]
        if ligne.get("note"):
            st.caption(f"🛈 {ligne['note']}")
        if ligne["code_ouv"] in a_valider:
            st.caption("⚠️ Rendement jamais confronté à un chantier réel.")

        col_val, col_btn = st.columns([2, 1])
        valeur = col_val.number_input(
            f"Heures par {origine_ouv[ligne['code_ouv']]['unite_ouv']}",
            min_value=0.0, value=float(ligne["qte_res"]), step=0.05,
            format="%.3f", key="corr_rend_valeur")
        col_btn.write("")
        if col_btn.button("Appliquer", width="stretch", key="corr_rend_ok"):
            ligne["qte_res"] = float(valeur)
            st.rerun()
        ancienne = origine["composition"][choix]["qte_res"]
        if abs(ligne["qte_res"] - ancienne) > 1e-9:
            st.caption(f"Corrigé : {ancienne:.3f} → **{ligne['qte_res']:.3f}** h")

        # ── Lever le doute ─────────────────────
        # Le ⚠️ ne dit pas « prix faux », il dit « rendement jamais
        # confronté au réel ». Le lever n'est donc pas un réglage
        # d'affichage : ça enregistre qu'un chantier a eu lieu et que les
        # heures relevées collent. Le sens inverse existe aussi — un
        # ouvrage sans ⚠️ n'a pas pour autant été vérifié, et poser le
        # doute sur celui qui déraille est le début de la calibration.
        code_courant = ligne["code_ouv"]
        if code_courant in a_valider:
            if st.button("✅ Rendement confirmé par un chantier réalisé",
                          key="corr_rend_valider", width="stretch"):
                tables["ouvrages_a_valider"] = [
                    c for c in tables["ouvrages_a_valider"]
                    if c != code_courant]
                st.rerun()
            st.caption(
                "À cocher **après** avoir relevé les heures réelles sur un "
                "chantier de cet ouvrage, pas avant : le ⚠️ retiré, plus "
                "rien ne signalera que ce rendement a été inventé."
            )
        elif st.button("⚠️ Remettre ce rendement en doute",
                        key="corr_rend_douter", width="stretch"):
            tables["ouvrages_a_valider"] = sorted(
                set(tables["ouvrages_a_valider"]) | {code_courant})
            st.rerun()

    # ── Créer un ouvrage ───────────────────────
    with st.expander("➕ Créer un ouvrage absent de la bibliothèque"):
        st.markdown(
            "Un ouvrage ne se résume pas à un prix : il lui faut un code, "
            "une **unité** — sans elle le contrôle d'unité devient "
            "inopérant — et surtout une **composition**. Sans composition, "
            "il se vendrait à 0 €, et le contrôle au chargement le refuse."
        )

        brouillon = st.session_state.setdefault(
            "nouvel_ouvrage", {"lignes": []})

        col_lot, col_num = st.columns(2)
        lot = col_lot.selectbox(
            "Lot", sorted(LOTS),
            format_func=lambda code: f"{code} — {LOTS[code]}",
            key="neuf_lot")
        # Numéro libre suivant, par pas de dix comme le reste de la
        # bibliothèque : on laisse de la place pour intercaler.
        pris = {int(o["code_ouv"].split(".")[1])
                 for o in tables["ouvrages"] if o["lot"] == lot}
        suivant = min(max(pris, default=0) + 10, 99)
        # Un widget à clé garde SA valeur d'un rerun à l'autre : sans
        # cette remise à jour, changer de lot laissait le numéro du lot
        # précédent, qui tombait sur un code déjà pris. La valeur doit
        # être posée AVANT d'instancier le widget.
        if st.session_state.get("neuf_dernier_lot") != lot:
            st.session_state.neuf_dernier_lot = lot
            st.session_state.neuf_num = suivant
        st.session_state.setdefault("neuf_num", suivant)
        # Pas de `value=` ici : le widget a une clé et sa valeur vient
        # de session_state. Passer les deux fait râler Streamlit, à
        # raison — laquelle des deux ferait foi ?
        numero = col_num.number_input(
            "Numéro", min_value=1, max_value=99, step=1, key="neuf_num")
        code_neuf = f"{lot}.{int(numero):02d}"

        libelle = st.text_input("Désignation", key="neuf_libelle",
                                 placeholder="Ce que le poste recouvre, "
                                              "tel qu'il sera lu par le client")
        unite = st.selectbox(
            "Unité",
            sorted({o["unite_ouv"] for o in tables["ouvrages"]}),
            key="neuf_unite",
            help="L'unité est éliminatoire à l'appariement : un poste "
                  "imposé au mètre courant ne sera jamais chiffré par un "
                  "ouvrage au m2.")

        deja_pris = code_neuf in tables["ouvrages_par_code"]
        if deja_pris:
            st.error(
                f"{code_neuf} existe déjà : "
                f"« {tables['ouvrages_par_code'][code_neuf]['libelle_ouv']} ». "
                f"Choisir un autre numéro.", icon="🛑")

        # ── Sa composition, une ligne à la fois ────────
        st.markdown(f"**Composition de {code_neuf}** — par "
                     f"{unite or 'unité'}")
        col_res, col_qte, col_add = st.columns([3, 2, 1])
        res_choisie = col_res.selectbox(
            "Ressource",
            [r["code_res"] for r in tables["ressources"]],
            format_func=lambda c: (
                f"{c} · {tables['ressources_par_code'][c]['libelle_res'][:34]}"
                f" — {tables['ressources_par_code'][c]['pu_res']:.2f} €"
                f"/{tables['ressources_par_code'][c]['unite_res']}"),
            key="neuf_res", label_visibility="collapsed")
        type_res = tables["ressources_par_code"][res_choisie]["type_res"]
        qte = col_qte.number_input(
            "Quantité", 0.0, value=1.0, step=0.05,
            format="%.3f", key="neuf_qte", label_visibility="collapsed",
            help="Sur une ressource MO, c'est le RENDEMENT en heures "
                  "par unité d'ouvrage.")
        col_add.write("")
        if col_add.button("➕", width="stretch", key="neuf_add",
                           disabled=qte <= 0):
            brouillon["lignes"] = [
                ligne for ligne in brouillon["lignes"]
                if ligne["code_res"] != res_choisie
            ] + [{"code_res": res_choisie, "qte_res": float(qte)}]
            st.rerun()
        if type_res == "MO":
            st.caption("🛈 Ressource de main-d'œuvre : cette quantité est un "
                        "**rendement**, en heures par unité d'ouvrage.")

        if brouillon["lignes"]:
            debourse = sum(
                ligne["qte_res"]
                * tables["ressources_par_code"][ligne["code_res"]]["pu_res"]
                for ligne in brouillon["lignes"])
            heures = sum(
                ligne["qte_res"] for ligne in brouillon["lignes"]
                if tables["ressources_par_code"][
                    ligne["code_res"]]["type_res"] == "MO")
            st.dataframe(
                [{"Ressource": ligne["code_res"],
                   "Désignation": tables["ressources_par_code"][
                       ligne["code_res"]]["libelle_res"],
                   "Type": tables["ressources_par_code"][
                       ligne["code_res"]]["type_res"],
                   "Quantité": ligne["qte_res"],
                   "Montant": round(
                       ligne["qte_res"] * tables["ressources_par_code"][
                           ligne["code_res"]]["pu_res"], 2)}
                  for ligne in brouillon["lignes"]],
                hide_index=True, width="stretch",
                column_config={
                    "Quantité": st.column_config.NumberColumn(format="%.3f"),
                    "Montant": st.column_config.NumberColumn(format="%.2f €"),
                })
            c1, c2, c3 = st.columns(3)
            c1.metric("Déboursé sec", f"{debourse:.2f} €")
            c2.metric("Prix de vente",
                       f"{debourse * coefficient_k(params):.2f} €")
            c3.metric("Main-d'œuvre", f"{heures:.3f} h")

            if st.button("↩️ Vider la composition", key="neuf_vider"):
                brouillon["lignes"] = []
                st.rerun()
        else:
            st.caption("Aucune ressource pour l'instant — un ouvrage sans "
                        "composition ne peut pas être créé.")

        pret = bool(libelle.strip()) and bool(unite) \
            and bool(brouillon["lignes"]) and not deja_pris
        if st.button(f"Créer {code_neuf}", type="primary", disabled=not pret,
                      key="neuf_creer", width="stretch"):
            tables["ouvrages"].append({
                "code_ouv": code_neuf, "lot": lot,
                "libelle_ouv": libelle.strip(), "unite_ouv": unite,
                "code_ref": ""})
            tables["ouvrages_par_code"][code_neuf] = tables["ouvrages"][-1]
            tables["composition"].extend(
                {"code_ouv": code_neuf, **ligne}
                for ligne in brouillon["lignes"])
            st.session_state.nouvel_ouvrage = {"lignes": []}
            st.rerun()
        if not pret and not deja_pris:
            st.caption("Il manque une désignation ou une composition.")

    # ── Effet en direct ────────────────────────
    modifs_res = [r for r in tables["ressources"]
                   if r["pu_res"] != origine["ressources_par_code"][
                       r["code_res"]]["pu_res"]]
    # Comparaison PAR CLÉ et non par position : un zip s'arrête à la
    # plus courte des deux listes, si bien que les lignes ajoutées en
    # créant un ouvrage passaient inaperçues — et le bouton
    # d'enregistrement n'apparaissait jamais pour elles.
    cle = lambda ligne: (ligne["code_ouv"], ligne["code_res"])  # noqa: E731
    compo_origine = {cle(c): c["qte_res"] for c in origine["composition"]}
    modifs_compo = [c for c in tables["composition"]
                     if compo_origine.get(cle(c)) != c["qte_res"]]
    ouvrages_neufs = [o for o in tables["ouvrages"]
                       if o["code_ouv"] not in origine["ouvrages_par_code"]]
    # Une validation ne change aucun prix : elle ne se verrait dans aucune
    # des comparaisons ci-dessus. Sans cette ligne, lever un doute ne
    # ferait apparaître aucun bouton d'enregistrement, et le travail serait
    # perdu au rafraîchissement suivant.
    modifs_valid = (sorted(tables["ouvrages_a_valider"])
                     != sorted(origine["ouvrages_a_valider"]))

    avant = calibration(params)
    apres = calibration(params, tables=tables)

    m1, m2, m3 = st.columns(3)
    reperes = ([f"+{len(ouvrages_neufs)} ouvrage(s)"] if ouvrages_neufs else [])
    if modifs_valid:
        avant_val, apres_val = (len(origine["ouvrages_a_valider"]),
                                 len(tables["ouvrages_a_valider"]))
        reperes.append(f"{avant_val} → {apres_val} à valider")
    m1.metric("Valeurs corrigées", len(modifs_res) + len(modifs_compo),
               " · ".join(reperes) or None, delta_color="off")
    m2.metric("Écart moyen absolu",
               f"{apres['ecart_moyen_absolu'] * 100:.1f} %",
               f"{(apres['ecart_moyen_absolu'] - avant['ecart_moyen_absolu']) * 100:+.1f} pt"
               if modifs_res or modifs_compo else None,
               delta_color="inverse")
    m3.metric("Devis hors cible",
               sum(1 for r in apres["lignes"] if abs(r["ecart"]) > 0.15),
               help="Cible : moins de 15 % d'écart sur CHAQUE ligne.")

    if modifs_res or modifs_compo or ouvrages_neufs or modifs_valid:
        st.dataframe(
            [{"Devis": r["devis"], "Objet": r["objet"],
               "Forfait vendu": r["forfait"], "Calculé": r["calcule"],
               "Écart": r["ecart"] * 100,
               "Écart avant": a["ecart"] * 100}
              for r, a in zip(apres["lignes"], avant["lignes"])],
            hide_index=True, width="stretch",
            column_config={
                "Forfait vendu": st.column_config.NumberColumn(format="%.0f €"),
                "Calculé": st.column_config.NumberColumn(format="%.0f €"),
                "Écart": st.column_config.NumberColumn(format="%+.1f %%"),
                "Écart avant": st.column_config.NumberColumn(format="%+.1f %%"),
            })

        # ── Enregistrer ────────────────────────
        a_ecrire = {}
        if modifs_res:
            a_ecrire["ressources"] = tables["ressources"]
        if modifs_compo or ouvrages_neufs:
            a_ecrire["composition"] = tables["composition"]
        if modifs_valid:
            a_ecrire["ouvrages_a_valider"] = sorted(
                tables["ouvrages_a_valider"])
        if ouvrages_neufs:
            # Le lot se déduit du code : on ne le réécrit pas dans le
            # fichier, sinon les deux pourraient diverger.
            a_ecrire["ouvrages"] = [
                {k: v for k, v in o.items() if k != "lot"}
                for o in tables["ouvrages"]]

        try:
            github = dict(st.secrets.get("github", {}))
        except Exception:
            github = {}
        depot, jeton = github.get("depot"), github.get("token")

        col_ok, col_raz = st.columns(2)
        with col_ok:
            if depot and jeton:
                if st.button(
                    f"📤 Enregistrer {len(a_ecrire)} table(s) dans le dépôt",
                    type="primary", width="stretch",
                ):
                    with st.spinner("Écriture dans le dépôt…"):
                        try:
                            for nom, contenu in a_ecrire.items():
                                commiter_table(
                                    nom,
                                    json.dumps(contenu, ensure_ascii=False,
                                                indent=2) + "\n",
                                    depot, jeton,
                                    branche=github.get("branche", "main"))
                        except ErreurDepot as err:
                            st.error(str(err), icon="🛑")
                        else:
                            st.success(
                                "Enregistré. L'app se redéploie dans une à "
                                "deux minutes ; ces valeurs deviendront "
                                "celles de la bibliothèque.", icon="✅")
            else:
                for nom, contenu in a_ecrire.items():
                    st.download_button(
                        f"⬇️ {nom}.json",
                        json.dumps(contenu, ensure_ascii=False, indent=2) + "\n",
                        file_name=f"{nom}.json", mime="application/json",
                        width="stretch", key=f"dl_{nom}")
        with col_raz:
            if st.button("↩️ Repartir des valeurs enregistrées",
                          width="stretch"):
                del st.session_state.tables_editees
                st.rerun()

        st.caption(
            "Tant que ce n'est pas enregistré, ces corrections ne valent "
            "que pour cet aperçu : les devis et les offres continuent "
            "d'utiliser les valeurs de la bibliothèque."
        )
    else:
        st.caption("Aucune valeur corrigée pour l'instant.")


# ══════════════════════════════════════════════════
#  4. Lexique métier
# ══════════════════════════════════════════════════

with onglet_lexique:
    st.header("Lexique métier")
    st.markdown(
        "Un pouvoir adjudicateur n'écrit pas « faïence » : il écrit "
        "« carrelage mural ». Ce lexique traduit son vocabulaire vers "
        "celui de la bibliothèque, **avant** toute comparaison de "
        "libellés. C'est lui qui fait la différence entre un poste "
        "apparié et un poste laissé sans prix."
    )

    b = _bordereau(params)

    # ── Banc d'essai ──────────────────────────
    st.subheader("Banc d'essai")
    st.caption(
        "Colle ici un libellé qui n'a pas été apparié, et regarde ce que "
        "l'outil en retient. C'est le cycle qui permet de régler le "
        "lexique sans écrire de Python."
    )

    col_lib, col_unite = st.columns([4, 1])
    with col_lib:
        essai = st.text_input(
            "Libellé du poste",
            "Sablage des maçonneries de façade",
            label_visibility="collapsed",
            placeholder="Libellé tel qu'il figure dans le métré",
        )
    with col_unite:
        # Le m2 par défaut, et non le premier par ordre alphabétique :
        # 29 des 49 ouvrages y sont, et surtout le libellé prérempli
        # est un travail de façade. Avec « FF » en tête, le banc
        # d'essai s'ouvrait sur cinq forfaits sans rapport avec
        # l'exemple — de quoi croire l'outil cassé au premier regard.
        unites = sorted({ligne["unite_ouv"] for ligne in b.values()})
        unite_essai = st.selectbox(
            "Unité",
            unites,
            index=unites.index("m2") if "m2" in unites else 0,
            label_visibility="collapsed",
        )

    if essai.strip():
        mots = normaliser(essai)
        avec_operation = normaliser(essai, garder_operation=True)

        st.markdown(
            "**Ce que l'outil retient :** "
            + (" · ".join(f"`{m}`" for m in mots) or "_rien_")
            + ("  \n**Opération :** dépose"
               if est_demolition(avec_operation)
               else "  \n**Opération :** mise en œuvre")
        )
        if not mots:
            st.warning(
                "Aucun mot significatif : tous ont été écartés comme mots "
                "vides de métré. Aucun appariement n'est possible.",
                icon="⚠️",
            )

        candidats = suggerer(
            {"designation": essai, "unite": unite_essai}, b, limite=5
        )
        if not candidats:
            st.error(
                f"Aucun ouvrage en « {unite_essai} » : ce n'est pas un "
                f"problème de vocabulaire, il manque l'ouvrage dans la "
                f"bibliothèque.",
                icon="🛑",
            )
        else:
            st.dataframe(
                [
                    {
                        "Score": s,
                        "": "✅" if s >= SEUIL_CONFIANCE
                             else ("🟡" if s >= SEUIL_SUGGESTION else "⚪"),
                        "Code": c,
                        "Ouvrage": b[c]["libelle_ouv"],
                        "PU": b[c]["pu_vente"],
                    }
                    for c, s in candidats
                ],
                hide_index=True,
                width="stretch",
                column_config={
                    "Score": st.column_config.ProgressColumn(
                        "Score", min_value=0.0, max_value=1.0, format="%.2f"),
                    "PU": st.column_config.NumberColumn(format="%.2f €"),
                },
            )
            st.caption(
                f"🟡 au-dessus de {SEUIL_SUGGESTION:.2f} : proposé comme "
                f"suggestion · ✅ au-dessus de {SEUIL_CONFIANCE:.2f} : "
                f"présenté comme solide — ce qui ne veut pas dire juste. "
                f"⚪ en dessous : l'outil se tait."
            )

    # ── Enrichir ───────────────────────────────
    st.subheader("Ajouter un terme")
    st.caption(
        "L'ajout prend effet immédiatement — pour **tous** les "
        "utilisateurs de l'app, pas seulement toi. Relance le banc "
        "d'essai ci-dessus pour en voir l'effet ; l'appariement de "
        "l'onglet « métré » est refait lui aussi."
    )

    col_var, col_canon, col_bouton = st.columns([2, 2, 1])
    with col_var:
        variante = st.text_input(
            "Terme du cahier des charges",
            placeholder="sablage  ·  carrelage mural",
        )
    with col_canon:
        canonique = st.text_input(
            "Terme de la bibliothèque",
            placeholder="nettoyage  ·  faience",
        )
    with col_bouton:
        st.write("")
        ajouter = st.button(
            "➕ Ajouter", width="stretch",
            disabled=not (variante.strip() and canonique.strip()),
        )

    if ajouter:
        # Une expression multi-mots se traduit AVANT la découpe en mots,
        # un mot seul APRÈS : ce ne sont pas les mêmes tables.
        try:
            if " " in variante.strip():
                ajouter_expression(variante, canonique)
            else:
                ajouter_synonyme(variante, canonique)
        except ValueError as err:
            st.error(str(err), icon="🛑")
        else:
            st.session_state.lexique_version = (
                st.session_state.get("lexique_version", 0) + 1
            )
            st.rerun()

    message_commit = st.session_state.pop("lexique_commit_texte", None)
    if message_commit:
        st.success(message_commit, icon="✅")

    # ── Ce qui a été ajouté à chaud ────────────────
    ajouts = dict(SURCOUCHE["expressions"], **SURCOUCHE["synonymes"])
    if ajouts:
        st.warning(
            f"**{len(ajouts)} terme(s) ajouté(s) à chaud.** Ils valent "
            "pour **l'app entière et tous ses utilisateurs**, pas pour "
            "toi seul — et **ils ne survivront pas au redémarrage**, "
            "que Streamlit Cloud déclenche tout seul après quelques "
            "heures d'inactivité. Pour les garder, colle le bloc "
            "ci-dessous dans `chiffrage/lexique.py` et commite.",
            icon="⚠️",
        )
        st.code(surcouche_en_python(), language="python")
        # ── Rendre permanent, si un jeton est configuré ─────────
        # `st.secrets` LÈVE quand aucun fichier de secrets n'existe —
        # ce n'est pas un dict vide. C'est le cas normal en local, et
        # un plantage y serait absurde : sans jeton, on n'affiche
        # simplement pas le bouton.
        try:
            github = dict(st.secrets.get("github", {}))
        except Exception:
            github = {}
        depot, jeton = github.get("depot"), github.get("token")

        if depot and jeton:
            st.markdown(
                f"**Commiter dans `{depot}`** — les termes sont écrits "
                f"dans `chiffrage/lexique_local.json`, **du JSON et non "
                f"du code**. L'app se redéploie seule ensuite, et ils "
                f"deviennent définitifs."
            )
            if st.button("📤 Commiter ces termes", type="primary",
                          width="stretch"):
                with st.spinner("Écriture dans le dépôt…"):
                    try:
                        fusion, url = commiter_lexique(
                            None, depot, jeton,
                            branche=github.get("branche", "main"),
                        )
                    except ErreurDepot as err:
                        st.error(str(err), icon="🛑")
                    else:
                        # Ce qui vient d'être commité devient la couche
                        # locale : les ajouts à chaud n'ont plus lieu
                        # d'être, ils sont dans le dépôt.
                        adopter_local(fusion)
                        st.session_state.lexique_version = (
                            st.session_state.get("lexique_version", 0) + 1
                        )
                        # Le message doit survivre au rerun qui suit,
                        # sinon personne ne le voit jamais.
                        st.session_state.lexique_commit_texte = (
                            f"Commité. [Voir le commit]({url}) — l'app se "
                            f"redéploie dans une à deux minutes.")
                        st.rerun()
        else:
            st.caption(
                "Aucun jeton GitHub configuré : les termes ne peuvent "
                "pas être commités depuis ici. Voir README, section "
                "« Rendre les termes permanents »."
            )

        col_dl, col_raz = st.columns(2)
        with col_dl:
            st.download_button(
                "⬇️ Télécharger le bloc",
                surcouche_en_python(),
                file_name="lexique_ajouts.py",
                mime="text/x-python",
                width="stretch",
            )
        with col_raz:
            if st.button("🗑️ Oublier ces ajouts", width="stretch",
                            help="Pour tout le monde, comme l'ajout."):
                vider_surcouche()
                st.session_state.lexique_version = (
                    st.session_state.get("lexique_version", 0) + 1
                )
                st.rerun()

    # ── Le lexique du dépôt ────────────────────
    st.subheader("Le lexique du dépôt")
    filtre = st.text_input("Filtrer", placeholder="faience, enduit…",
                           label_visibility="collapsed")

    entrees = (
        [("expression", k, v) for k, v in EXPRESSIONS.items()]
        + [("mot", k, v) for k, v in SYNONYMES.items() if v]
        + [("appris", k, v) for k, v in LOCAL["expressions"].items()]
        + [("appris", k, v) for k, v in LOCAL["synonymes"].items()]
    )
    if filtre.strip():
        motif = filtre.strip().lower()
        entrees = [e for e in entrees if motif in f"{e[1]} {e[2]}"]

    st.dataframe(
        [{"Type": t, "Terme du CSC": k, "→ Bibliothèque": v}
         for t, k, v in sorted(entrees, key=lambda e: (e[0], e[1]))],
        hide_index=True, width="stretch", height=260,
    )

    with st.expander("Marqueurs de démolition"):
        st.markdown(
            "Ces mots ne sont **pas** comparés comme les autres : ils sont "
            "retirés du libellé et traités comme une dimension à part, "
            "comme l'unité. Sans ça, « dépose du carrelage mural » "
            "s'appariait à « dépose de plafond » — deux fois le mot "
            "« dépose », et rien d'autre en commun. Une dépose comparée à "
            "une pose est fortement pénalisée, mais **pas éliminée** : "
            "l'unité est déclarée dans le métré, l'opération n'est "
            "qu'inférée de mots."
        )
        st.write(" · ".join(f"`{m}`" for m in sorted(DEMOLITION)))


# ══════════════════════════════════════════════════
#  5. Calibration
# ══════════════════════════════════════════════════

with onglet_calib:
    st.header("Calibration sur les devis historiques")
    st.markdown(
        "Les six devis forfaitaires réellement vendus, re-chiffrés avec la "
        "bibliothèque. **C'est ici que se juge la qualité des prix.**"
    )

    cal = calibration(params)

    lignes = [
        {
            "Devis": r["devis"],
            "Objet": r["objet"],
            "Forfait vendu": r["forfait"],
            "Calculé": r["calcule"],
            # ×100 EXPLICITE : contrairement au format « % » d'Excel,
            # st.column_config.NumberColumn ne multiplie pas. Passer la
            # fraction brute affichait « -0,1 % » pour un écart de
            # -11,5 % — trois devis hors cible avaient l'air parfaits.
            "Écart": r["ecart"] * 100,
            "h MO": r["heures_mo"],
            "€/h vendu": r["prix_horaire_implicite"],
        }
        for r in cal["lignes"]
    ]
    st.dataframe(
        lignes,
        width="stretch",
        hide_index=True,
        column_config={
            "Forfait vendu": st.column_config.NumberColumn(format="%.0f €"),
            "Calculé": st.column_config.NumberColumn(format="%.0f €"),
            "Écart": st.column_config.NumberColumn(format="%+.1f %%"),
            "h MO": st.column_config.NumberColumn(format="%.1f"),
            "€/h vendu": st.column_config.NumberColumn(format="%.0f €"),
        },
    )

    moyen = cal["ecart_moyen_absolu"]
    hors_cible = [r for r in cal["lignes"] if abs(r["ecart"]) > 0.15]
    st.metric("Écart moyen absolu", f"{moyen * 100:.1f} %",
              help="Cible : moins de 15 % sur CHAQUE ligne, pas en moyenne")

    if hors_cible:
        st.warning(
            f"**{len(hors_cible)} devis au-delà de 15 % d'écart** : "
            + ", ".join(r["devis"] for r in hors_cible)
            + ".\n\nDeux hypothèses, non tranchées faute de relevés :\n"
            "1. les quantités estimées sont trop élevées — elles viennent "
            "des descriptifs des devis PDF, **pas de relevés** ;\n"
            "2. ces chantiers ont été vendus sous leur coût analytique.",
            icon="⚠️",
        )

    st.subheader("Les 13 rendements jamais validés")
    st.markdown(
        "Ces ouvrages ont été créés pour couvrir des postes qui restaient "
        "sans prix — ce qui rendait toute offre irrégulière. Aucun n'a "
        "jamais été confronté à un chantier réel : **ce sont les premiers "
        "à valider.**"
    )
    st.dataframe(
        [
            {
                "Code": c,
                "Désignation": b[c]["libelle_ouv"],
                "Un.": b[c]["unite_ouv"],
                "h/unité": b[c]["heures_mo"],
                "PU vente": b[c]["pu_vente"],
            }
            for c in OUVRAGES_A_VALIDER
        ],
        width="stretch",
        hide_index=True,
    )


# ══════════════════════════════════════════════════
#  6. Paramètres
# ══════════════════════════════════════════════════

with onglet_params:
    st.header("Paramètres")
    st.markdown(
        "L'identité de l'entreprise et les coefficients de vente. Ce sont "
        "des valeurs **d'entreprise**, pas des constantes techniques : "
        "elles se règlent ici, sans toucher au code."
    )

    st.subheader("Identité")
    st.caption(
        "Figure en en-tête de chaque devis et de chaque courrier de "
        "justification de prix."
    )
    col_g, col_d = st.columns(2)
    saisie_entreprise = {}
    with col_g:
        saisie_entreprise["nom"] = st.text_input(
            "Raison sociale", ENTREPRISE["nom"])
        saisie_entreprise["adresse"] = st.text_input(
            "Adresse", ENTREPRISE["adresse"])
        saisie_entreprise["cp_ville"] = st.text_input(
            "Code postal et localité", ENTREPRISE["cp_ville"])
    with col_d:
        saisie_entreprise["pays"] = st.text_input("Pays", ENTREPRISE["pays"])
        saisie_entreprise["tva"] = st.text_input(
            "Numéro de TVA", ENTREPRISE["tva"])
    saisie_entreprise["activite"] = st.text_area(
        "Activité", ENTREPRISE["activite"], height=80)

    st.subheader("Coefficients de vente")
    st.caption(
        "`pu_vente = déboursé sec × K`, avec "
        "`K = (1+FG)(1+FC)(1+aléas)(1+marge)`. Les valeurs enregistrées "
        "ici deviennent celles de départ ; la barre latérale sert à "
        "simuler autre chose le temps d'une session."
    )
    c1, c2, c3, c4 = st.columns(4)
    saisie_params = dict(PARAMS)
    saisie_params["fg"] = c1.number_input(
        "Frais généraux (%)", 0.0, 100.0, PARAMS["fg"] * 100, 0.5,
        key="p_fg") / 100
    saisie_params["fc"] = c2.number_input(
        "Frais de chantier (%)", 0.0, 100.0, PARAMS["fc"] * 100, 0.5,
        key="p_fc") / 100
    saisie_params["aleas"] = c3.number_input(
        "Aléas (%)", 0.0, 100.0, PARAMS["aleas"] * 100, 0.5,
        key="p_aleas") / 100
    saisie_params["marge"] = c4.number_input(
        "Marge (%)", 0.0, 100.0, PARAMS["marge"] * 100, 0.5,
        key="p_marge") / 100

    t1, t2, t3 = st.columns(3)
    saisie_params["tva"] = t1.number_input(
        "TVA logement privé (%)", 0.0, 100.0, PARAMS["tva"] * 100, 1.0,
        key="p_tva", help="6 % : logement de plus de dix ans, usage "
                           "principalement privé, consommateur final.") / 100
    saisie_params["tva_marche_public"] = t2.number_input(
        "TVA marché public (%)", 0.0, 100.0,
        PARAMS["tva_marche_public"] * 100, 1.0, key="p_tva_mp") / 100
    t3.metric("Coefficient K", f"{coefficient_k(saisie_params):.4f}",
               f"{(coefficient_k(saisie_params) - coefficient_k(PARAMS)):+.4f}"
               if abs(coefficient_k(saisie_params)
                       - coefficient_k(PARAMS)) > 1e-9 else None)

    # ── Enregistrer ────────────────────────────
    try:
        contenu = serialiser(saisie_entreprise, saisie_params)
    except ValueError as err:
        st.error(str(err), icon="🛑")
        contenu = None

    modifie = contenu is not None and contenu != serialiser(ENTREPRISE, PARAMS)

    if not modifie:
        st.caption("Aucune modification par rapport aux valeurs enregistrées.")
    else:
        st.info(
            "Modifications non enregistrées. Elles ne s'appliqueront "
            "qu'une fois commitées — comme pour le lexique, rien ne "
            "survit au redémarrage de l'app.",
            icon="💡",
        )

    try:
        github = dict(st.secrets.get("github", {}))
    except Exception:
        github = {}
    depot, jeton = github.get("depot"), github.get("token")

    if modifie and depot and jeton:
        if st.button("📤 Enregistrer dans le dépôt", type="primary",
                      width="stretch"):
            with st.spinner("Écriture dans le dépôt…"):
                try:
                    url = commiter_parametres(
                        contenu, depot, jeton,
                        branche=github.get("branche", "main"))
                except ErreurDepot as err:
                    st.error(str(err), icon="🛑")
                else:
                    st.success(
                        f"Enregistré. [Voir le commit]({url}) — l'app se "
                        f"redéploie dans une à deux minutes, et ces "
                        f"valeurs deviendront celles de départ.",
                        icon="✅")
    elif modifie:
        st.caption(
            "Aucun jeton GitHub configuré : impossible d'enregistrer "
            "depuis ici. Le fichier à créer est "
            "`chiffrage/parametres_local.json` — son contenu exact est "
            "ci-dessous."
        )

    if contenu:
        with st.expander("Contenu de `chiffrage/parametres_local.json`"):
            st.code(contenu, language="json")
            st.download_button(
                "⬇️ Télécharger", contenu,
                file_name="parametres_local.json",
                mime="application/json", width="stretch")
