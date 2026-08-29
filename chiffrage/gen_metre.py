"""
╔══════════════════════════════════════════════════════════════════════════╗
║  GÉNÉRATEUR DE MÉTRÉ DE MARCHÉ PUBLIC (fictif, pour entraînement)        ║
║  CSC 2026-TP-0147 — Commune de Schaerbeek                                ║
╚══════════════════════════════════════════════════════════════════════════╝

Produit `METRE_CSC_2026-TP-0147_Schaerbeek.xlsx` : 49 postes, 10 lots, au
format qu'un pouvoir adjudicateur impose réellement — colonne « Quantité »
remplie, colonnes « PU » et « Montant » vides, formules de sous-total et de
total déjà en place.

C'est le fichier d'entrée de `metre_io.remplir_metre()` : il sert à
s'entraîner sans attendre un vrai cahier spécial des charges.

──────────────────────────────────────────────────────────────────────────
PIÈGE OPENPYXL CORRIGÉ ICI — À NE PAS RÉINTRODUIRE
──────────────────────────────────────────────────────────────────────────
`ws.insert_rows()` NE DÉCALE PAS les références des formules déjà écrites.
Une formule `=IF($G20="";"";$F20*$G20)` reste littéralement attachée à la
ligne 20 même après que sa cellule soit passée en ligne 21 : le fichier
s'ouvre sans erreur et calcule FAUX.

Conséquence : les lignes de sous-total sont écrites AU FIL DE LA BOUCLE,
jamais insérées après coup. Aucun appel à insert_rows/delete_rows dans ce
module — et il ne faut pas en ajouter.

Numérotation : le lot 09 démarre à 09.03. C'est volontaire — les postes
09.01/09.02 du référentiel ne sont pas repris dans ce marché. Un métré réel
présente régulièrement ce genre de trou : ne jamais « renuméroter pour faire
propre », les codes appartiennent au pouvoir adjudicateur.

Usage :
    python -m chiffrage.gen_metre                       # -> ./METRE_...xlsx
    python -m chiffrage.gen_metre /chemin/sortie.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# Identité du marché fictif
# ═══════════════════════════════════════════════════════════════════════════

MARCHE = {
    "reference": "CSC 2026-TP-0147",
    "pouvoir_adjudicateur": "Commune de Schaerbeek — Service des Travaux publics",
    "objet": "Rénovation d'un immeuble communal — enveloppe, parachèvements et techniques",
    "procedure": "Procédure négociée directe avec publication préalable",
    "depot": "Offres à déposer pour le 30/09/2026 à 11h00 au plus tard",
}

NOM_FICHIER_DEFAUT = "METRE_CSC_2026-TP-0147_Schaerbeek.xlsx"

LOTS_METRE = {
    "00": "Installations de chantier et généralités",
    "01": "Démolitions et déposes",
    "02": "Maçonnerie et structure",
    "03": "Façades",
    "04": "Étanchéité",
    "05": "Isolation",
    "06": "Plafonnage, chapes et revêtements",
    "07": "Peintures et finitions",
    "08": "Menuiseries et sanitaire",
    "09": "Électricité et conformité",
}

# Nature des postes :
#   QF = quantité forfaitaire garantie (payée telle quelle)
#   QP = quantité présumée (payée au métré réellement exécuté)
#   FF = forfait global, quantité 1
#
# (code, designation, nature, unite, quantite)
POSTES = [
    # ── 00 · Installations de chantier et généralités ─────────────────────
    ("00.01", "Installation de chantier, amenée et repli", "FF", "FF", 1),
    ("00.02", "Échafaudage de façade, location comprise", "QP", "m2", 180),
    ("00.03", "Signalisation, clôture et sécurisation des accès", "FF", "FF", 1),
    ("00.04", "Protection des ouvrages conservés", "QP", "m2", 120),
    ("00.05", "Évacuation des déchets, tri et taxes comprises", "QP", "m3", 45),
    ("00.06", "Dossier as-built, PV de réception, garanties", "FF", "FF", 1),

    # ── 01 · Démolitions et déposes ───────────────────────────────────────
    ("01.01", "Piquage d'enduit dégradé", "QP", "m2", 160),
    ("01.02", "Démolition de cloison légère", "QP", "m2", 38),
    ("01.03", "Dépose de plafond existant", "QP", "m2", 145),
    ("01.04", "Dépose de menuiserie extérieure", "QF", "pce", 12),
    ("01.05", "Dépose de revêtements de sol et chape existante", "QP", "m2", 95),
    ("01.06", "Dépose d'appareils sanitaires et tuyauterie", "QF", "pce", 8),

    # ── 02 · Maçonnerie et structure ──────────────────────────────────────
    ("02.01", "Maçonnerie de rebouchage en briques", "QP", "m2", 24),
    ("02.02", "Rejointoiement de maçonnerie, mortier de chaux", "QP", "m2", 85),
    ("02.03", "Réparation de béton dégradé, passivation des aciers", "QP", "m2", 18),
    ("02.04", "Pose de linteau préfabriqué, étançonnement compris", "QF", "m", 14),
    ("02.05", "Seuil en pierre bleue, pose comprise", "QF", "m", 9),

    # ── 03 · Façades ──────────────────────────────────────────────────────
    ("03.01", "Nettoyage haute pression de façade", "QP", "m2", 180),
    ("03.02", "Enduit de façade minéral armé", "QP", "m2", 165),
    ("03.03", "Peinture de façade siloxane", "QP", "m2", 165),
    ("03.04", "Solin, relevé et couvre-mur", "QF", "m", 42),
    ("03.05", "Cimentage hydrofuge de soubassement", "QP", "m2", 36),

    # ── 04 · Étanchéité ───────────────────────────────────────────────────
    ("04.01", "Étanchéité bitumineuse bicouche", "QP", "m2", 78),
    ("04.02", "Étanchéité EPDM collée", "QP", "m2", 45),
    ("04.03", "Étanchéité liquide sous carrelage", "QP", "m2", 32),
    ("04.04", "Évacuation PVC Ø 110", "QF", "m", 26),

    # ── 05 · Isolation ────────────────────────────────────────────────────
    ("05.01", "Isolation PIR 80 mm sous plafond", "QP", "m2", 145),
    ("05.02", "Isolation laine minérale 100 mm", "QP", "m2", 88),
    ("05.03", "Pare-vapeur et étanchéité à l'air", "QP", "m2", 145),

    # ── 06 · Plafonnage, chapes et revêtements ────────────────────────────
    ("06.01", "Chape de ravoirage armée 6 cm", "QP", "m2", 95),
    ("06.02", "Plafonnage sur maçonnerie, deux couches", "QP", "m2", 210),
    ("06.03", "Faux plafond BA13 sur ossature", "QP", "m2", 145),
    ("06.04", "Faïence murale, profils de finition compris", "QP", "m2", 48),
    ("06.05", "Carrelage de sol grès cérame", "QP", "m2", 95),
    ("06.06", "Rebouchage et enduit de rattrapage sur linteaux", "QP", "m", 32),
    ("06.07", "Cornières et profilés de finition", "QP", "m", 120),

    # ── 07 · Peintures et finitions ───────────────────────────────────────
    ("07.01", "Peinture murs intérieurs, deux couches", "QP", "m2", 320),
    ("07.02", "Peinture plafonds intérieurs, deux couches", "QP", "m2", 210),
    ("07.03", "Peinture sur menuiseries bois, ponçage compris", "QP", "m2", 42),
    ("07.04", "Enduit de lissage avant peinture", "QP", "m2", 180),

    # ── 08 · Menuiseries et sanitaire ─────────────────────────────────────
    ("08.01", "Châssis PVC double vitrage", "QF", "m2", 26),
    ("08.02", "Porte d'entrée bois massif", "QF", "pce", 1),
    ("08.03", "Garde-corps acier thermolaqué", "QF", "m", 18),
    ("08.04", "WC suspendu complet, bâti-support inclus", "QF", "pce", 3),
    ("08.05", "Essais d'étanchéité, rinçage, mise en service", "FF", "FF", 1),
    ("08.06", "Alimentation multicouche Ø 16", "QP", "m", 65),

    # ── 09 · Électricité et conformité ────────────────────────────────────
    ("09.03", "Prise de courant 2P+T encastrée", "QF", "pce", 24),
    ("09.04", "Mise à la terre et liaisons équipotentielles RGIE", "FF", "FF", 1),
    ("09.05", "Contrôle de conformité par organisme agréé", "FF", "FF", 1),
]

# ═══════════════════════════════════════════════════════════════════════════
# Mise en forme
# ═══════════════════════════════════════════════════════════════════════════

COLONNES = [
    ("A", "N°", 6),
    ("B", "Poste", 10),
    ("C", "Désignation des ouvrages", 58),
    ("D", "Nature", 8),
    ("E", "Unité", 8),
    ("F", "Quantité", 11),
    ("G", "PU HTVA", 13),
    ("H", "Montant HTVA", 15),
]

LIGNE_ENTETE = 9          # ligne des titres de colonnes
FORMAT_EUR = '#,##0.00 "€"'
FORMAT_QTE = "#,##0.00"

_BORDURE = Border(*(Side(style="thin", color="B0B0B0"),) * 4)
_FILL_TITRE = PatternFill("solid", fgColor="1F3864")
_FILL_LOT = PatternFill("solid", fgColor="D9E2F3")
_FILL_SAISIE = PatternFill("solid", fgColor="FFF2CC")   # colonne PU à remplir
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")


def _cell(ws, row, col, valeur, *, gras=False, fill=None, fmt=None, align=None):
    c = ws.cell(row=row, column=col, value=valeur)
    if gras:
        c.font = Font(bold=True)
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _BORDURE
    return c


def generer_metre(chemin=NOM_FICHIER_DEFAUT):
    """
    Écrit le classeur du métré et retourne (chemin, nb_postes, nb_lots).

    Une seule feuille « MÉTRÉ », structurée comme un vrai document de marché :
    cartouche du marché · en-têtes · un bloc par lot terminé par son
    sous-total · récapitulatif des lots · total général.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MÉTRÉ"

    for lettre, _, largeur in COLONNES:
        ws.column_dimensions[lettre].width = largeur

    # ── Cartouche ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    titre = ws["A1"]
    titre.value = f"{MARCHE['pouvoir_adjudicateur']}"
    titre.font = Font(bold=True, size=13, color="FFFFFF")
    titre.fill = _FILL_TITRE
    titre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for i, (label, valeur) in enumerate(
        [
            ("Référence", MARCHE["reference"]),
            ("Objet", MARCHE["objet"]),
            ("Procédure", MARCHE["procedure"]),
            ("Dépôt", MARCHE["depot"]),
            ("Document", "INVENTAIRE / MÉTRÉ RÉCAPITULATIF — à compléter par le soumissionnaire"),
        ],
        start=3,
    ):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        ws.cell(row=i, column=2, value=valeur)

    ws.cell(
        row=8,
        column=1,
        value="Nature : QF = quantité forfaitaire · QP = quantité présumée "
        "· FF = forfait global. Tout poste laissé sans prix rend l'offre "
        "irrégulière (art. 76 AR 18/04/2017).",
    ).font = Font(italic=True, size=9)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=8)

    # ── En-têtes de colonnes ──────────────────────────────────────────────
    for idx, (_, libelle, _) in enumerate(COLONNES, start=1):
        c = _cell(ws, LIGNE_ENTETE, idx, libelle, gras=True, align="center")
        c.fill = _FILL_LOT
    ws.freeze_panes = ws.cell(row=LIGNE_ENTETE + 1, column=1)

    # ── Corps : un bloc par lot, sous-total écrit AU FIL DE LA BOUCLE ─────
    #    (surtout pas d'insert_rows après coup — voir le cartouche du module)
    ligne = LIGNE_ENTETE + 1
    numero = 0
    lignes_sous_totaux = []          # (code_lot, libelle, n° de ligne du sous-total)

    for code_lot, libelle_lot in LOTS_METRE.items():
        postes_du_lot = [p for p in POSTES if p[0].startswith(code_lot + ".")]
        if not postes_du_lot:
            continue

        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=8)
        c = ws.cell(row=ligne, column=1, value=f"LOT {code_lot} — {libelle_lot.upper()}")
        c.font = Font(bold=True, size=11)
        c.fill = _FILL_LOT
        for col in range(1, 9):
            ws.cell(row=ligne, column=col).border = _BORDURE
        ligne += 1

        premiere = ligne
        for code, designation, nature, unite, qte in postes_du_lot:
            numero += 1
            _cell(ws, ligne, 1, numero, align="center")
            _cell(ws, ligne, 2, code, align="center")
            _cell(ws, ligne, 3, designation)
            _cell(ws, ligne, 4, nature, align="center")
            _cell(ws, ligne, 5, unite, align="center")
            _cell(ws, ligne, 6, float(qte), fmt=FORMAT_QTE)
            # Colonne G : à remplir par le soumissionnaire (laissée vide).
            _cell(ws, ligne, 7, None, fill=_FILL_SAISIE, fmt=FORMAT_EUR)
            # Colonne H : formule écrite pour CETTE ligne, à sa position finale.
            _cell(
                ws,
                ligne,
                8,
                f'=IF($G{ligne}="","",$F{ligne}*$G{ligne})',
                fmt=FORMAT_EUR,
            )
            ligne += 1
        derniere = ligne - 1

        _cell(ws, ligne, 3, f"Sous-total lot {code_lot}", gras=True, fill=_FILL_TOTAL)
        for col in (1, 2, 4, 5, 6, 7):
            _cell(ws, ligne, col, None, fill=_FILL_TOTAL)
        _cell(
            ws,
            ligne,
            8,
            f"=SUM(H{premiere}:H{derniere})",
            gras=True,
            fill=_FILL_TOTAL,
            fmt=FORMAT_EUR,
        )
        lignes_sous_totaux.append((code_lot, libelle_lot, ligne))
        ligne += 2

    # ── Récapitulatif ─────────────────────────────────────────────────────
    ligne += 1
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=8)
    c = ws.cell(row=ligne, column=1, value="RÉCAPITULATIF PAR LOT")
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center")
    ligne += 1

    premiere_recap = ligne
    for code_lot, libelle_lot, ligne_st in lignes_sous_totaux:
        _cell(ws, ligne, 2, code_lot, align="center")
        _cell(ws, ligne, 3, libelle_lot)
        _cell(ws, ligne, 8, f"=H{ligne_st}", fmt=FORMAT_EUR)
        ligne += 1
    derniere_recap = ligne - 1

    _cell(ws, ligne, 3, "TOTAL DE L'OFFRE — HTVA", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 8, f"=SUM(H{premiere_recap}:H{derniere_recap})",
          gras=True, fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_ht = ligne
    ligne += 1

    _cell(ws, ligne, 3, "TVA 21 %", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 8, f"=H{ligne_ht}*0.21", gras=True, fill=_FILL_TOTAL,
          fmt=FORMAT_EUR)
    ligne_tva = ligne
    ligne += 1

    _cell(ws, ligne, 3, "TOTAL DE L'OFFRE — TVAC", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 8, f"=H{ligne_ht}+H{ligne_tva}", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)

    ligne += 3
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=8)
    ws.cell(
        row=ligne,
        column=1,
        value="Le soumissionnaire complète exclusivement la colonne « PU HTVA ». "
        "Les colonnes Quantité et Montant ne peuvent être modifiées.",
    ).font = Font(italic=True, size=9)

    ws.print_title_rows = f"{LIGNE_ENTETE}:{LIGNE_ENTETE}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    wb.save(chemin)
    return chemin, len(POSTES), len(lignes_sous_totaux)


def _colonne_de(nom):
    """Utilitaire : 'Quantité' -> 'F'. Utilisé par metre_io pour rester d'accord
    avec la structure produite ici."""
    for lettre, libelle, _ in COLONNES:
        if libelle == nom:
            return lettre
    raise KeyError(nom)


if __name__ == "__main__":
    cible = sys.argv[1] if len(sys.argv) > 1 else NOM_FICHIER_DEFAUT
    chemin, nb_postes, nb_lots = generer_metre(cible)
    print(f"Métré généré : {chemin}")
    print(f"  {nb_postes} postes · {nb_lots} lots")
    print(f"  colonne à remplir : {_colonne_de('PU HTVA')} "
          f"(quantités en {_colonne_de('Quantité')}, "
          f"montants calculés en {_colonne_de('Montant HTVA')})")
    print(f"  dernière colonne utilisée : {get_column_letter(len(COLONNES))}")
