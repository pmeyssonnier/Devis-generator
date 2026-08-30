"""
╔══════════════════════════════════════════════════════════════════════════╗
║  IMPORT D'UN MÉTRÉ IMPOSÉ ET REMPLISSAGE DE L'OFFRE                      ║
╚══════════════════════════════════════════════════════════════════════════╝

Trois fonctions :
    lire_metre(chemin)                  -> liste des postes imposés
    lire_metre_complet(chemin)          -> postes ET anomalies
    remplir_metre(chemin, sortie, ...)  -> écrit les PU et rend un rapport

──────────────────────────────────────────────────────────────────────────
DEUX RÈGLES NON NÉGOCIABLES
──────────────────────────────────────────────────────────────────────────

1. OUVRIR SANS `data_only=True`.
   Avec `data_only=True`, openpyxl ne voit que la dernière valeur mise en
   cache par Excel ; à la sauvegarde, TOUTES les formules du pouvoir
   adjudicateur sont remplacées par ces valeurs figées. Le fichier renvoyé
   ne recalcule plus rien — et personne ne s'en aperçoit avant l'ouverture
   des offres. Ce module ouvre donc toujours en mode formules, et lit les
   quantités qui, elles, sont des valeurs saisies.

2. NE JAMAIS DÉSACTIVER LE CONTRÔLE D'UNITÉ.
   Chiffrer au m² un poste imposé au mètre courant ne se voit qu'au moment
   de facturer. Quand l'unité du métré ne correspond pas à celle de
   l'ouvrage, ce module N'ÉCRIT PAS de prix et remonte le poste dans
   `ecarts_unite` : c'est un arbitrage humain, pas une conversion
   automatique.

3. RIEN N'EST ÉCARTÉ EN SILENCE.
   Un poste qu'on ne sait pas lire doit APPARAÎTRE, pas disparaître. La
   règle a été apprise à la dure : une quantité écrite `=12.5*3` — ce
   qu'un pouvoir adjudicateur fait couramment — était ignorée, le poste
   sortait du décompte, et le rapport annonçait fièrement « tous les
   postes portent un prix » sur une offre amputée de trois lignes. Le
   garde-fou certifiait l'inverse de la vérité.

   Toute ligne écartée part donc dans `anomalies`, et le rapport ne peut
   plus se déclarer complet tant qu'il en reste.

Rappel : un poste laissé sans prix rend l'offre IRRÉGULIÈRE et entraîne son
rejet (art. 76 AR 18/04/2017). `remplir_metre()` liste explicitement les
postes restés vides — cette liste doit être à zéro avant l'envoi.
"""

import re
import shutil

from openpyxl import load_workbook

from .bibliotheque import MAPPING, PARAMS
from .moteur import calcul_bordereau

# ══════════════════════════════════════════
#  À quoi ressemble un code de poste
# ══════════════════════════════════════════
#
# Les codes appartiennent au pouvoir adjudicateur, pas à nous. `NN.NN`
# était la forme du métré d'entraînement, et elle seule était acceptée :
# un cahier des charges numéroté `01.02.03` ou `3.2` rendait ZÉRO poste,
# avec pour tout message « aucun poste lu ».
#
# Le motif accepte donc : 3.2 · 03.02 · 01.02.03 · 03.02.A · 1.01.10 ·
# A.1.2 · 03-02 · 03/02.
#
# Élargir un motif, c'est risquer de prendre pour un poste ce qui n'en
# est pas. Trois garde-fous :
#   · au moins un séparateur — « 2026 » n'est pas un code ;
#   · au moins un chiffre — « A.B » non plus ;
#   · et surtout, LE CODE NE SUFFIT PAS : une ligne n'est un poste que
#     si elle porte aussi une quantité lisible. C'est cette seconde
#     condition, pas la sévérité du motif, qui écarte le bruit.
#
# Modifiable si un marché impose une forme exotique — sans jamais
# supprimer le contrôle.
RE_CODE_POSTE = re.compile(
    r"^[A-Za-z0-9]{1,3}(?:[./-][A-Za-z0-9]{1,3}){1,4}$")


def est_code_poste(valeur):
    """La cellule ressemble-t-elle à un code de poste ?"""
    if not isinstance(valeur, str):
        return False
    code = valeur.strip()
    return bool(RE_CODE_POSTE.match(code)) and any(c.isdigit() for c in code)

COL_CODE = 2       # B
COL_DESIGNATION = 3  # C
COL_NATURE = 4     # D
COL_UNITE = 5      # E
COL_QUANTITE = 6   # F
COL_PU = 7         # G
COL_MONTANT = 8    # H

# Équivalences d'écriture admises pour la comparaison d'unités. Volontairement
# minimaliste : uniquement des variantes typographiques du MÊME unité, jamais
# une conversion (m -> m2 n'y figurera jamais).
_EQUIV_UNITES = {
    "m²": "m2",
    "m^2": "m2",
    "m3": "m3",
    "m³": "m3",
    "mct": "m",
    "ml": "m",
    "p": "pce",
    "pc": "pce",
    "u": "pce",
    "pièce": "pce",
    "piece": "pce",
    "ff": "ff",
    "forfait": "ff",
    "gp": "ff",
    "h": "h",
}


def normaliser_unite(unite):
    """'m²' -> 'm2', 'PCE' -> 'pce', 'FF' -> 'ff'. Aucune conversion physique."""
    if unite is None:
        return ""
    u = str(unite).strip().lower().replace(" ", "")
    return _EQUIV_UNITES.get(u, u)


def _anomalie(genre, code, ligne, detail, feuille=None):
    return {"genre": genre, "code": code, "ligne": ligne,
            "detail": detail, "feuille": feuille}


def feuilles_avec_postes(chemin):
    """
    Inventaire des feuilles du classeur et de ce qu'elles contiennent.

    Retourne [{"nom", "nb_postes", "recapitulatif"}], dans l'ordre du
    classeur. Sert à laisser l'utilisateur choisir : un métré réel se
    répartit souvent en « Lot 01 », « Lot 02 »… plus un
    « Récapitulatif » qui REPREND les mêmes codes. Traiter ce dernier
    ferait compter les postes deux fois.

    `recapitulatif` est une simple présomption tirée du nom de la
    feuille — un indice pour pré-décocher la case, jamais une décision.
    """
    wb = load_workbook(chemin)
    inventaire = []
    for nom in wb.sheetnames:
        ws = wb[nom]
        nb = sum(
            1 for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            if len(row) >= COL_QUANTITE
            and est_code_poste(row[COL_CODE - 1].value)
        )
        minuscule = nom.lower()
        inventaire.append({
            "nom": nom,
            "nb_postes": nb,
            "recapitulatif": any(mot in minuscule for mot in
                                  ("recap", "récap", "synth", "total",
                                   "resume", "résumé")),
        })
    wb.close()
    return inventaire


def _feuilles_a_lire(wb, feuilles):
    """Résout l'argument `feuilles` : None, un nom, ou une liste."""
    if feuilles is None:
        # Par défaut, TOUTES les feuilles : mieux vaut lire un
        # récapitulatif et signaler les doublons que perdre un lot
        # entier sans que personne ne s'en aperçoive.
        return list(wb.sheetnames)
    if isinstance(feuilles, str):
        return [feuilles]
    return list(feuilles)


def lire_metre_complet(chemin, feuilles=None):
    """
    Lit les postes d'un métré imposé, et rend AUSSI ce qu'il n'a pas su lire.

    Retourne {"postes": [...], "anomalies": [...], "feuilles": [...]}.
    Chaque poste porte sa `feuille` et sa `ligne` : c'est ce qui permet
    de réécrire le prix au bon endroit dans un classeur multi-feuilles.

    `feuilles` : None (toutes), un nom, ou une liste de noms.

    ── Pourquoi le classeur est ouvert DEUX FOIS ────────────────────
    Sans `data_only`, une cellule de quantité contenant une formule rend
    la formule, pas son résultat : `"=12.5*3"` et non `37.5`. Avec
    `data_only=True`, on obtient la dernière valeur calculée par Excel —
    mais on perdrait les formules du pouvoir adjudicateur à la sauvegarde.

    Les deux besoins s'opposent, donc on lit deux fois : le classeur des
    formules pour la structure, celui des valeurs pour les quantités
    calculées. L'écriture, elle, se fait toujours sur le premier.

    Un piège demeure : la valeur en cache n'existe que si Excel a ouvert
    et enregistré le fichier au moins une fois. Un classeur produit par
    un programme n'en a pas. Dans ce cas le poste n'est PAS deviné : il
    part en anomalie, à saisir à la main.
    """
    wb = load_workbook(chemin)                    # formules — cf. cartouche

    # Second classeur : les valeurs mises en cache par Excel. Il ne sert
    # QU'À LIRE, et n'est jamais sauvegardé.
    try:
        wb_valeurs = load_workbook(chemin, data_only=True)
    except Exception:
        wb_valeurs = None

    noms = [n for n in _feuilles_a_lire(wb, feuilles) if n in wb.sheetnames]
    postes, anomalies, vus = [], [], {}

    for nom in noms:
        ws = wb[nom]
        ws_valeurs = (wb_valeurs[nom]
                       if wb_valeurs is not None and nom in wb_valeurs.sheetnames
                       else None)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if len(row) < COL_QUANTITE:
                continue
            brut = row[COL_CODE - 1].value
            if not est_code_poste(brut):
                continue
            code = brut.strip()
            numero = row[0].row
            designation = (row[COL_DESIGNATION - 1].value or "")
            unite = (row[COL_UNITE - 1].value or "")
            qte = row[COL_QUANTITE - 1].value

            # Une ligne sans quantité NI désignation NI unité n'est
            # probablement pas un poste : un titre, une référence, un
            # numéro d'article. On ne l'invente pas en anomalie.
            ressemble = bool(str(designation).strip()) or bool(str(unite).strip())

            # ── Quantité calculée par formule ────────────────────
            if isinstance(qte, str) and qte.lstrip().startswith("="):
                cache = (ws_valeurs.cell(numero, COL_QUANTITE).value
                          if ws_valeurs is not None else None)
                if isinstance(cache, (int, float)):
                    anomalies.append(_anomalie(
                        "quantite_formule", code, numero,
                        f"Quantité calculée par la formule « {qte} ». Valeur "
                        f"en cache utilisée : {cache:g}. À vérifier.", nom))
                    qte = cache
                else:
                    anomalies.append(_anomalie(
                        "quantite_illisible", code, numero,
                        f"Quantité calculée par la formule « {qte} », sans "
                        f"valeur en cache — le classeur n'a jamais été ouvert "
                        f"dans Excel. Poste NON chiffré : ouvrir puis "
                        f"enregistrer le fichier, ou saisir la quantité.", nom))
                    continue

            if not isinstance(qte, (int, float)):
                if ressemble:
                    anomalies.append(_anomalie(
                        "quantite_absente", code, numero,
                        f"Aucune quantité lisible (cellule : {qte!r}). "
                        f"Poste NON chiffré.", nom))
                continue

            if qte < 0:
                anomalies.append(_anomalie(
                    "quantite_negative", code, numero,
                    f"Quantité négative ({qte:g}). Poste NON chiffré.", nom))
                continue

            if qte == 0:
                anomalies.append(_anomalie(
                    "quantite_nulle", code, numero,
                    "Quantité nulle — poste « pour mémoire » ou oubli du "
                    "pouvoir adjudicateur. Chiffré à 0 €, à confirmer.", nom))

            if code in vus:
                # Deux lignes pour un même code : on ne choisit PAS à la
                # place de l'utilisateur, mais on ne l'ignore plus en
                # silence. Le cas le plus fréquent est une feuille
                # « Récapitulatif » qui reprend les codes des lots :
                # la chiffrer aussi doublerait le montant de l'offre.
                ou = vus[code]
                anomalies.append(_anomalie(
                    "code_duplique", code, numero,
                    f"Ce code figure déjà en « {ou[0]} » ligne {ou[1]}. "
                    f"Seule la première occurrence est chiffrée ; la "
                    f"quantité de celle-ci ({qte:g}) est ignorée. Si cette "
                    f"feuille est un récapitulatif, la décocher.", nom))
                continue
            vus[code] = (nom, numero)

            postes.append({
                "code": code,
                "designation": str(designation).strip(),
                "nature": str(row[COL_NATURE - 1].value or "").strip(),
                "unite": str(unite).strip(),
                "quantite": float(qte),
                "ligne": numero,
                "feuille": nom,
            })

    wb.close()
    return {"postes": postes, "anomalies": anomalies, "feuilles": noms}


def lire_metre(chemin, feuilles=None):
    """Les seuls postes lisibles. Voir lire_metre_complet() pour ce qui
    a été écarté — et il vaut mieux le regarder."""
    return lire_metre_complet(chemin, feuilles)["postes"]


def remplir_metre(
    chemin_metre,
    chemin_sortie,
    mapping=None,
    params=None,
    tva=None,
    feuilles=None,
):
    """
    Recopie le métré imposé et y écrit les prix unitaires de la bibliothèque.

    Écrit UNIQUEMENT la colonne PU (G). Les quantités et les formules de
    montant/sous-total/total du pouvoir adjudicateur sont laissées
    intactes — c'est le fichier du PA, pas le nôtre.

    tva : 0,21 par défaut ici (marché public). Sert au rapport, pas au
          fichier : le taux du classeur est celui imposé par le PA.

    Retourne un rapport :
        anomalies[]     lignes qu'on n'a PAS su lire — quantité en
                        formule, absente, négative, code en double.
                        Chacune est un poste absent de l'offre.
        postes          nombre de postes lus
        chiffres[]      postes prix écrit  {code, ouvrage, pu, quantite, montant, heures}
        non_couverts[]  postes sans correspondance dans le mapping
        ecarts_unite[]  correspondance trouvée mais unité incompatible -> NON chiffrés
        vides[]         codes restés sans prix (non_couverts + ecarts_unite)
        total_ht, montant_tva, total_tvac, heures_mo, jours_homme
        fichier         chemin du fichier produit
    """
    mapping = MAPPING if mapping is None else mapping
    taux = PARAMS["tva_marche_public"] if tva is None else tva
    bordereau = calcul_bordereau(params)

    shutil.copyfile(chemin_metre, chemin_sortie)
    wb = load_workbook(chemin_sortie)   # PAS de data_only : cf. cartouche

    lecture = lire_metre_complet(chemin_sortie, feuilles)
    postes, anomalies = lecture["postes"], lecture["anomalies"]

    chiffres, non_couverts, ecarts_unite = [], [], []
    total_ht = heures = 0.0

    for poste in postes:
        code_ouv = mapping.get(poste["code"])
        if code_ouv is None or code_ouv not in bordereau:
            non_couverts.append(
                {
                    "code": poste["code"],
                    "designation": poste["designation"],
                    "unite": poste["unite"],
                    "quantite": poste["quantite"],
                    "motif": "aucun ouvrage en bibliothèque"
                    if code_ouv is None
                    else f"mapping vers un ouvrage inexistant ({code_ouv})",
                }
            )
            continue

        ref = bordereau[code_ouv]
        if normaliser_unite(poste["unite"]) != normaliser_unite(ref["unite_ouv"]):
            # Contrôle d'unité : on ne convertit pas, on ne chiffre pas.
            ecarts_unite.append(
                {
                    "code": poste["code"],
                    "designation": poste["designation"],
                    "unite_metre": poste["unite"],
                    "code_ouv": code_ouv,
                    "unite_ouvrage": ref["unite_ouv"],
                }
            )
            continue

        # Le prix retourne dans SA feuille : dans un classeur à un lot
        # par onglet, tout écrire sur la première les mettrait tous au
        # même endroit — et le pouvoir adjudicateur recevrait un
        # classeur incohérent sans qu'aucune erreur ne soit levée.
        wb[poste["feuille"]].cell(
            row=poste["ligne"], column=COL_PU, value=ref["pu_vente"])
        montant = round(ref["pu_vente"] * poste["quantite"], 2)
        h = round(ref["heures_mo"] * poste["quantite"], 2)
        total_ht += montant
        heures += h
        chiffres.append(
            {
                "code": poste["code"],
                "code_ouv": code_ouv,
                "feuille": poste["feuille"],
                "designation": poste["designation"],
                "unite": poste["unite"],
                "quantite": poste["quantite"],
                "pu": ref["pu_vente"],
                "montant": montant,
                "heures": h,
            }
        )

    wb.save(chemin_sortie)
    wb.close()

    total_ht = round(total_ht, 2)
    # Une ligne écartée à la lecture est un poste ABSENT de l'offre —
    # aussi grave qu'un poste sans prix, et bien plus discret.
    bloquantes = [a for a in anomalies
                   if a["genre"] not in ("quantite_formule", "quantite_nulle")]

    return {
        "fichier": chemin_sortie,
        "anomalies": anomalies,
        "anomalies_bloquantes": bloquantes,
        "postes": len(postes),
        "chiffres": chiffres,
        "non_couverts": non_couverts,
        "ecarts_unite": ecarts_unite,
        "feuilles": lecture["feuilles"],
        "vides": [p["code"] for p in non_couverts] + [e["code"] for e in ecarts_unite],
        "total_ht": total_ht,
        "tva_taux": taux,
        "montant_tva": round(total_ht * taux, 2),
        "total_tvac": round(total_ht * (1 + taux), 2),
        "heures_mo": round(heures, 2),
        "jours_homme": round(heures / 8.0, 2),
    }


def imprimer_rapport(rapport):
    """Rapport de remplissage lisible, à relire AVANT d'envoyer l'offre."""
    r = rapport
    out = [
        f"Fichier produit : {r['fichier']}",
        f"{r['postes']} postes lus · {len(r['chiffres'])} chiffrés "
        f"· {len(r['vides'])} sans prix",
        "",
        f"Total des postes chiffrés : {r['total_ht']:,.2f} € HTVA".replace(",", " "),
        f"TVA {r['tva_taux'] * 100:.0f} % : {r['montant_tva']:,.2f} € "
        f"-> {r['total_tvac']:,.2f} € TVAC".replace(",", " "),
        f"Main-d'œuvre : {r['heures_mo']:.0f} h "
        f"({r['jours_homme']:.0f} jours-homme)",
    ]
    if r["ecarts_unite"]:
        out += ["", "⚠️  ÉCARTS D'UNITÉ — non chiffrés, à arbitrer à la main :"]
        for e in r["ecarts_unite"]:
            out.append(
                f"   {e['code']}  métré en « {e['unite_metre']} » mais "
                f"{e['code_ouv']} est en « {e['unite_ouvrage']} » — "
                f"{e['designation']}"
            )
    if r["non_couverts"]:
        out += ["", f"⚠️  {len(r['non_couverts'])} POSTES NON COUVERTS "
                    "(il manque l'ouvrage en bibliothèque) :"]
        for p in r["non_couverts"]:
            out.append(f"   {p['code']}  {p['designation']} "
                       f"[{p['unite']} × {p['quantite']:g}]")
    if r.get("anomalies"):
        out += ["", f"⚠️  {len(r['anomalies'])} LIGNE(S) NON LUE(S) OU "
                     "DOUTEUSE(S) DANS LE MÉTRÉ :"]
        for a in r["anomalies"]:
            out.append(f"   ligne {a['ligne']} · {a['code']} — {a['detail']}")

    if r["vides"] or r.get("anomalies_bloquantes"):
        manquants = list(r["vides"]) + [a["code"] for a
                                          in r.get("anomalies_bloquantes", [])]
        out += [
            "",
            "🛑 OFFRE IRRÉGULIÈRE EN L'ÉTAT — art. 76 AR 18/04/2017.",
            f"   {len(manquants)} postes ne partiraient pas chiffrés : "
            + ", ".join(dict.fromkeys(manquants)),
            "   Les chiffrer à la main, créer les ouvrages manquants, ou "
            "corriger le métré avant envoi.",
        ]
    else:
        out += ["", "✅ Tous les postes du métré portent un prix."]
    return "\n".join(out)
