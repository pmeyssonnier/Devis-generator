"""
╔══════════════════════════════════════════════════════════════════════════╗
║  EXPORT DE LA BIBLIOTHÈQUE VERS EXCEL — 6 onglets                        ║
╚══════════════════════════════════════════════════════════════════════════╝

Produit `bibliotheque_prix_bagbatter.xlsx` :

    PARAMS       les 4 coefficients + K + taux de TVA
    RESSOURCES   49 lignes, prix d'achat et taux horaires
    OUVRAGES     49 lignes, 9 lots
    COMPOSITION  les liens ouvrage <-> ressource (rendements MO compris)
    BORDEREAU    le calculé : déboursés, PU de vente, heures par unité
    MAPPING      correspondance postes de métré imposé -> ouvrages,
                 avec la colonne « statut » qui isole les 13 ouvrages dont le
                 rendement n'a jamais été validé par le chef d'entreprise

L'onglet BORDEREAU porte de VRAIES FORMULES Excel (pas des valeurs figées) :
le chef d'entreprise peut modifier un taux horaire dans RESSOURCES ou un
coefficient dans PARAMS et voir tous les prix de vente bouger. C'est ce qui
rend le fichier utilisable par quelqu'un qui ne lit pas de Python.

Arborescence Drive attendue par le client (créée par la cellule 2 du
notebook d'origine) :

    MyDrive/BAG_BATTER/Chiffrage/
        01_bibliotheque/     bibliotheque_prix_bagbatter.xlsx
        02_metres_recus/     métrés reçus des pouvoirs adjudicateurs
        03_offres_remises/   métrés complétés, horodatés
        04_archives/         versions successives de la bibliothèque

Usage :
    python -m chiffrage.export_xlsx [chemin_de_sortie.xlsx]
"""

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .bibliotheque import (
    COMPOSITION,
    LOTS,
    MAPPING,
    OUVRAGES_A_VALIDER,
    OUVRAGES,
    OUVRAGES_PAR_CODE,
    PARAMS,
    RESSOURCES,
    RESSOURCES_PAR_CODE,
)
from .moteur import calcul_bordereau, coefficient_k

NOM_FICHIER_DEFAUT = "bibliotheque_prix_bagbatter.xlsx"

FORMAT_EUR = '#,##0.00 "€"'
FORMAT_PCT = "0.00 %"
_FILL_ENTETE = PatternFill("solid", fgColor="1F3864")
_FILL_CALC = PatternFill("solid", fgColor="E2EFDA")


def _entetes(ws, libelles, largeurs):
    for i, (libelle, largeur) in enumerate(zip(libelles, largeurs), start=1):
        c = ws.cell(row=1, column=i, value=libelle)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _FILL_ENTETE
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.freeze_panes = "A2"


def exporter_bibliotheque(chemin=NOM_FICHIER_DEFAUT, params=None):
    """Écrit le classeur à 6 onglets. Retourne (chemin, nb_onglets)."""
    p = params or PARAMS
    bordereau = calcul_bordereau(p)
    wb = Workbook()

    # ── 1. PARAMS ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "PARAMS"
    _entetes(ws, ["Paramètre", "Valeur", "Commentaire"], [26, 14, 72])
    lignes_params = [
        ("Frais généraux", p["fg"], "Siège, véhicules, assurances, administratif"),
        ("Frais de chantier", p["fc"], "Non imputables à un ouvrage précis"),
        ("Aléas", p["aleas"], "Provision pour imprévus"),
        ("Marge", p["marge"], "Bénéfice visé"),
    ]
    for i, (libelle, valeur, com) in enumerate(lignes_params, start=2):
        ws.cell(row=i, column=1, value=libelle)
        ws.cell(row=i, column=2, value=valeur).number_format = FORMAT_PCT
        ws.cell(row=i, column=3, value=com)
    ws.cell(row=6, column=1, value="Coefficient K").font = Font(bold=True)
    c = ws.cell(row=6, column=2, value="=(1+B2)*(1+B3)*(1+B4)*(1+B5)")
    c.font = Font(bold=True)
    c.fill = _FILL_CALC
    c.number_format = "0.0000"
    ws.cell(row=6, column=3, value="pu_vente = déboursé sec × K — recalculé par Excel")
    ws.cell(row=7, column=1, value="TVA logement privé")
    ws.cell(row=7, column=2, value=p["tva"]).number_format = FORMAT_PCT
    ws.cell(row=7, column=3,
            value="6 % : logement > 10 ans, usage privé, consommateur final")
    ws.cell(row=8, column=1, value="TVA marché public").font = Font(bold=True)
    ws.cell(row=8, column=2, value=p["tva_marche_public"]).number_format = FORMAT_PCT
    ws.cell(row=8, column=3,
            value="21 % — taux à utiliser pour toute offre en marché public")
    ws.cell(row=10, column=1, value=f"Export du {date.today():%d/%m/%Y}").font = (
        Font(italic=True)
    )

    # ── 2. RESSOURCES ─────────────────────────────────────────────────────
    ws = wb.create_sheet("RESSOURCES")
    _entetes(ws, ["code_res", "libelle_res", "type_res", "unite_res", "pu_res"],
             [12, 46, 10, 12, 14])
    for i, r in enumerate(RESSOURCES, start=2):
        ws.cell(row=i, column=1, value=r["code_res"])
        ws.cell(row=i, column=2, value=r["libelle_res"])
        ws.cell(row=i, column=3, value=r["type_res"])
        ws.cell(row=i, column=4, value=r["unite_res"])
        ws.cell(row=i, column=5, value=r["pu_res"]).number_format = FORMAT_EUR
    ligne_res = {r["code_res"]: i for i, r in enumerate(RESSOURCES, start=2)}

    # ── 3. OUVRAGES ───────────────────────────────────────────────────────
    ws = wb.create_sheet("OUVRAGES")
    _entetes(ws, ["code_ouv", "lot", "libelle_lot", "libelle_ouv", "unite_ouv",
                  "code_ref"], [12, 8, 34, 58, 12, 18])
    for i, o in enumerate(OUVRAGES, start=2):
        ws.cell(row=i, column=1, value=o["code_ouv"])
        ws.cell(row=i, column=2, value=o["lot"])
        ws.cell(row=i, column=3, value=LOTS.get(o["lot"], ""))
        ws.cell(row=i, column=4, value=o["libelle_ouv"])
        ws.cell(row=i, column=5, value=o["unite_ouv"])
        ws.cell(row=i, column=6, value=o["code_ref"])

    # ── 4. COMPOSITION ────────────────────────────────────────────────────
    #    Les lignes MO portent le RENDEMENT (h/unité) : c'est la colonne que
    #    le chef d'entreprise doit relire en priorité.
    ws = wb.create_sheet("COMPOSITION")
    _entetes(ws, ["code_ouv", "libelle_ouv", "code_res", "libelle_res",
                  "type_res", "unite_res", "qte_res", "pu_res", "montant"],
             [12, 50, 12, 42, 10, 12, 12, 12, 14])
    for i, c in enumerate(COMPOSITION, start=2):
        res = RESSOURCES_PAR_CODE[c["code_res"]]
        ouv = OUVRAGES_PAR_CODE[c["code_ouv"]]
        ws.cell(row=i, column=1, value=c["code_ouv"])
        ws.cell(row=i, column=2, value=ouv["libelle_ouv"])
        ws.cell(row=i, column=3, value=c["code_res"])
        ws.cell(row=i, column=4, value=res["libelle_res"])
        ws.cell(row=i, column=5, value=res["type_res"])
        ws.cell(row=i, column=6, value=res["unite_res"])
        ws.cell(row=i, column=7, value=c["qte_res"]).number_format = "0.000"
        ws.cell(row=i, column=8,
                value=f"=RESSOURCES!E{ligne_res[c['code_res']]}"
                ).number_format = FORMAT_EUR
        cell = ws.cell(row=i, column=9, value=f"=G{i}*H{i}")
        cell.number_format = FORMAT_EUR
        cell.fill = _FILL_CALC
    derniere_compo = len(COMPOSITION) + 1

    # ── 5. BORDEREAU (formules vivantes) ──────────────────────────────────
    ws = wb.create_sheet("BORDEREAU")
    _entetes(ws, ["code_ouv", "lot", "libelle_ouv", "unite_ouv", "deb_mo",
                  "deb_mat", "deb_eqp", "debourse_sec", "K", "pu_vente",
                  "heures_mo"],
             [12, 8, 56, 12, 13, 13, 13, 15, 10, 14, 12])
    for i, o in enumerate(OUVRAGES, start=2):
        code = o["code_ouv"]
        ref = bordereau[code]
        ws.cell(row=i, column=1, value=code)
        ws.cell(row=i, column=2, value=o["lot"])
        ws.cell(row=i, column=3, value=o["libelle_ouv"])
        ws.cell(row=i, column=4, value=o["unite_ouv"])
        plage = f"COMPOSITION!$A$2:$A${derniere_compo}"
        types = f"COMPOSITION!$E$2:$E${derniere_compo}"
        montants = f"COMPOSITION!$I$2:$I${derniere_compo}"
        qtes = f"COMPOSITION!$G$2:$G${derniere_compo}"
        for col, type_res in ((5, "MO"), (6, "MAT"), (7, "EQP")):
            ws.cell(
                row=i, column=col,
                value=f'=SUMIFS({montants},{plage},$A{i},{types},"{type_res}")',
            ).number_format = FORMAT_EUR
        ws.cell(row=i, column=8, value=f"=E{i}+F{i}+G{i}").number_format = FORMAT_EUR
        ws.cell(row=i, column=9, value="=PARAMS!$B$6").number_format = "0.0000"
        cell = ws.cell(row=i, column=10, value=f"=ROUND(H{i}*I{i},2)")
        cell.number_format = FORMAT_EUR
        cell.fill = _FILL_CALC
        cell.font = Font(bold=True)
        ws.cell(
            row=i, column=11,
            value=f'=SUMIFS({qtes},{plage},$A{i},{types},"MO")',
        ).number_format = "0.000"
        # Contrôle : la valeur calculée par Python doit correspondre à ce que
        # recalcule Excel. Écart -> une des deux implémentations a dérivé.
        ws.cell(row=i, column=13, value=ref["pu_vente"]).number_format = FORMAT_EUR
    ws.cell(row=1, column=13, value="pu_vente (Python)").font = Font(
        bold=True, color="FFFFFF"
    )
    ws.cell(row=1, column=13).fill = _FILL_ENTETE
    ws.column_dimensions["M"].width = 18

    # ── 6. MAPPING ────────────────────────────────────────────────────────
    ws = wb.create_sheet("MAPPING")
    _entetes(ws, ["poste_metre", "code_ouv", "libelle_ouv", "unite_ouv",
                  "statut"], [14, 12, 58, 12, 30])
    a_valider = set(OUVRAGES_A_VALIDER)
    for ligne, (poste, code_ouv) in enumerate(sorted(MAPPING.items()), start=2):
        ouv = OUVRAGES_PAR_CODE.get(code_ouv, {})
        ws.cell(row=ligne, column=1, value=poste)
        ws.cell(row=ligne, column=2, value=code_ouv)
        ws.cell(row=ligne, column=3, value=ouv.get("libelle_ouv", ""))
        ws.cell(row=ligne, column=4, value=ouv.get("unite_ouv", ""))
        if code_ouv in a_valider:
            ws.cell(row=ligne, column=5,
                    value="RENDEMENT À VALIDER").font = Font(bold=True)
        else:
            ws.cell(row=ligne, column=5, value="couvert")

    wb.save(chemin)
    return chemin, len(wb.sheetnames)


if __name__ == "__main__":
    cible = sys.argv[1] if len(sys.argv) > 1 else NOM_FICHIER_DEFAUT
    chemin, nb = exporter_bibliotheque(cible)
    print(f"Bibliothèque exportée : {chemin}")
    print(f"  {nb} onglets · {len(RESSOURCES)} ressources · {len(OUVRAGES)} ouvrages "
          f"· {len(COMPOSITION)} lignes de composition")
    print(f"  coefficient K = {coefficient_k():.4f}")
