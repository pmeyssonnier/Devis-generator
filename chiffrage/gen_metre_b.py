"""
╔══════════════════════════════════════════════════════════════════════════╗
║  DEUXIÈME MÉTRÉ D'ENTRAÎNEMENT (fictif)                                  ║
║  CSC 2026/043 — Régie foncière de la Ville de Bruxelles                  ║
╚══════════════════════════════════════════════════════════════════════════╝

Le premier métré d'entraînement (`gen_metre.py`) est le document dont
l'outil est né : ses colonnes sont celles qui étaient codées en dur, et
ses 49 codes figurent tous dans `MAPPING`. Il ne peut donc plus rien
apprendre — il passe par construction.

Celui-ci est écrit par un AUTRE pouvoir adjudicateur, et c'est tout son
intérêt : il range ses colonnes autrement, numérote ses postes à sa
façon, éclate le métré en une feuille par lot, et rédige ses
désignations sans savoir ce que contient notre bibliothèque. Il
s'entraîne donc sur ce qui casse en vrai :

    · DÉTECTION DES COLONNES — l'unité est en C, la quantité en D, le
      prix unitaire en E. Aucune colonne « Nature ». Les intitulés sont
      « U », « Qté », « P.U. (€) » : abrégés, comme dans la vraie vie.
    · EN-TÊTES À HAUTEUR VARIABLE — le cartouche n'a pas la même taille
      d'une feuille à l'autre. Une ligne d'en-tête fixe ne marche pas.
    · PLUSIEURS FEUILLES + UN RÉCAPITULATIF qui REPREND les mêmes codes.
      Le traiter comme un lot compterait chaque poste deux fois.
    · CODES D'UNE AUTRE FORME — `1.01`, et `3.04.a` pour les variantes.
    · QUANTITÉS EN FORMULE — `=32*7.5`, ce qu'un métreur écrit
      couramment. Les valeurs calculées sont injectées dans le fichier
      comme Excel le ferait ; UN poste en est privé volontairement (voir
      POSTE_SANS_CACHE), pour que la lecture ait aussi son cas d'échec.
      Une formule STOCKÉE s'écrit toujours en syntaxe américaine : la
      virgule y sépare des arguments, elle ne marque pas les décimales.
      `=32*7,5` se lit donc « 32 fois 7 » suivi d'un « 5 » égaré, et le
      classeur s'ouvre sur un message de récupération de contenu. Excel
      affiche la virgule tout seul, selon la langue du poste.
    · ÉCRITURES D'UNITÉ — `m²`, `ML`, `PC`, à normaliser sans convertir.
    · UN ÉCART D'UNITÉ RÉEL — le solin (2.08) est imposé au m² là où la
      bibliothèque le tient au mètre. Aucun ouvrage ne lui est donc
      proposé ; et si on le rapproche de 40.60 à la main, l'écriture du
      prix doit être REFUSÉE et le poste remonté à l'humain. Jamais
      converti.
    · DEUX QUASI-CORRESPONDANCES — « faux plafonds » (1.06) et
      « descentes d'eau pluviale » (3.06) désignent des ouvrages que la
      bibliothèque a, sous d'autres mots, et passent sous le seuil de
      suggestion. C'est le cas d'usage du lexique.
    · UN POSTE SANS ÉQUIVALENT — le sciage de trémie (1.09) n'existe pas
      dans la bibliothèque. Il faut créer l'ouvrage, ou renoncer au
      marché.
    · UNE LIGNE « POUR MÉMOIRE » sans quantité, qui ne doit pas
      disparaître en silence.

Aucun de ces pièges n'est gratuit : chacun a été rencontré, et la
plupart ont coûté un défaut. Un métré d'entraînement qui passe du
premier coup n'entraîne personne.

Le piège openpyxl d'`insert_rows` vaut ici comme dans `gen_metre.py` :
les sous-totaux sont écrits au fil de la boucle, jamais insérés après
coup. Aucun appel à insert_rows/delete_rows dans ce module.

Usage :
    python -m chiffrage metre2                     # -> ./METRE_CSC_2026-043...
    python -m chiffrage metre2 /chemin/sortie.xlsx
"""

import re
import shutil
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ═══════════════════════════════════════════════════════════════════════════
# Identité du marché fictif
# ═══════════════════════════════════════════════════════════════════════════

MARCHE = {
    "reference": "CSC 2026/043",
    "pouvoir_adjudicateur": "Ville de Bruxelles — Régie foncière",
    "objet": "École communale n° 12, annexe rue des Tanneurs — "
              "rénovation lourde",
    "procedure": "Procédure ouverte, marché à prix unitaires",
    "depot": "Ouverture des offres le 27/10/2026 à 10h00",
}

NOM_FICHIER_DEFAUT = "METRE_CSC_2026-043_Bruxelles.xlsx"

# Ce poste-ci porte une formule SANS valeur calculée : c'est le fichier
# tel qu'un tableur qui ne l'a jamais ouvert le laisse. La quantité est
# alors illisible, et le poste doit apparaître en anomalie plutôt que
# disparaître du décompte.
POSTE_SANS_CACHE = "2.04"

# (code, désignation, unité, quantité)
# Une quantité str commençant par « = » est une formule ; sa valeur
# calculée est injectée ensuite, comme Excel le ferait.
LOTS = [
    ("Lot 1 - Chantier et démolitions", [
        ("1.01", "Installation du chantier, amenée et repli du matériel",
         "FF", 1),
        ("1.02", "Échafaudage de façade y compris location et démontage",
         "m²", "=32*7.5"),
        ("1.03", "Bâchage et protection des ouvrages maintenus", "m²", 210),
        ("1.04", "Piquage des enduits de façade dégradés", "m²", 240),
        ("1.05", "Démolition de cloisons légères", "m²", 64),
        ("1.06", "Dépose des faux plafonds existants", "m²", 186),
        ("1.07", "Dépose des châssis extérieurs existants", "PC", 18),
        ("1.08", "Évacuation en conteneur, tri et taxes de mise en décharge",
         "m3", 62),
        ("1.09", "Sciage de béton armé et découpe de trémie d'escalier",
         "ML", 11),
    ]),
    ("Lot 2 - Maçonnerie et façades", [
        ("2.01", "Rebouchage de baies en maçonnerie de briques", "m²", 31),
        ("2.02", "Rejointoiement au mortier de chaux naturelle", "m²",
         "=140+65"),
        ("2.03", "Réparation des bétons et passivation des armatures",
         "m²", 26),
        ("2.04", "Nettoyage de la façade sous haute pression", "m²",
         "=32*7.5"),
        ("2.05", "Enduit de façade minéral sur treillis d'armature", "m²",
         205),
        ("2.06", "Mise en peinture de la façade, produit siloxane", "m²", 205),
        ("2.07", "Seuils et appuis en pierre bleue", "ML", 16),
        ("2.08", "Solin, relevé et couvre-mur en zinc", "m²", 38),
        ("2.09", "Cimentage hydrofuge du soubassement", "m²", 44),
    ]),
    ("Lot 3 - Étanchéité et isolation", [
        ("3.01", "Étanchéité bitumineuse bicouche sur toiture plate", "m²",
         96),
        ("3.02", "Étanchéité EPDM en membrane collée", "m²", 58),
        ("3.03", "Isolation en panneaux PIR 80 mm sous dalle", "m²", 174),
        ("3.04.a", "Isolation en laine minérale 100 mm, variante ossature "
                    "bois", "m²", 92),
        ("3.05", "Pare-vapeur et traitement de l'étanchéité à l'air", "m²",
         174),
        ("3.06", "Descentes d'eau pluviale en PVC Ø 110", "ML", 34),
    ]),
    ("Lot 4 - Plafonnage et finitions", [
        ("4.01", "Plafonnage des maçonneries, deux couches dressées", "m²",
         268),
        ("4.02", "Faux plafond en plaques BA13 sur ossature", "m²", 186),
        ("4.03", "Chape de ravoirage armée", "m²", 124),
        ("4.04", "Carrelage de sol en grès cérame", "m²", 124),
        ("4.05", "Faïence murale des sanitaires", "m²", 56),
        ("4.06", "Enduit de lissage avant mise en peinture", "m²", 268),
        ("4.07", "Peinture des murs intérieurs, deux couches", "m²", 268),
        ("4.08", "Peinture des plafonds intérieurs", "m²", 186),
        ("4.09", "Cornières et profilés de finition", "ML", 145),
    ]),
    ("Lot 5 - Châssis et sanitaire", [
        ("5.01", "Châssis en PVC à double vitrage, pose comprise", "m²", 38),
        ("5.02", "Porte d'entrée en bois massif, quincaillerie comprise",
         "PC", 2),
        ("5.03", "Garde-corps en acier thermolaqué", "ML", 24),
        ("5.04", "Étanchéité liquide sous carrelage des sanitaires", "m²",
         28),
        ("5.05", "WC suspendu complet sur bâti-support", "PC", 6),
        ("5.06", "Alimentation en tube multicouche Ø 16", "ML", 88),
        ("5.07", "Essais d'étanchéité, rinçage et mise en service", "FF", 1),
        ("5.08", "Mobilier de vestiaire — POUR MÉMOIRE, hors marché",
         "PC", None),
    ]),
    ("Lot 6 - Électricité", [
        ("6.01", "Prises de courant 2P+T encastrées", "PC", 42),
        ("6.02", "Mise à la terre et liaisons équipotentielles RGIE", "FF", 1),
        ("6.03", "Contrôle de conformité par organisme agréé", "FF", 1),
        ("6.04", "Dossier as-built, PV de réception et garanties", "FF", 1),
    ]),
]

NOM_RECAP = "Récapitulatif"

# ═══════════════════════════════════════════════════════════════════════════
# Mise en forme
# ═══════════════════════════════════════════════════════════════════════════
#
# L'ordre des colonnes n'est PAS celui du premier métré, et c'est le
# but : rien ici ne doit pouvoir être lu de mémoire.

COLONNES = [
    ("A", "N° poste", 11),
    ("B", "Désignation des travaux", 62),
    ("C", "U", 7),
    ("D", "Qté", 10),
    ("E", "P.U. (€)", 13),
    ("F", "Total (€)", 15),
]

FORMAT_EUR = '#,##0.00'
FORMAT_QTE = "#,##0.00"

_BORDURE = Border(*(Side(style="hair", color="808080"),) * 4)
_FILL_TITRE = PatternFill("solid", fgColor="7F0000")
_FILL_ENTETE = PatternFill("solid", fgColor="F2DCDB")
_FILL_SAISIE = PatternFill("solid", fgColor="FFFFCC")
_FILL_TOTAL = PatternFill("solid", fgColor="EDEDED")


def _cell(ws, row, col, valeur, *, gras=False, fill=None, fmt=None,
           align=None, italique=False):
    c = ws.cell(row=row, column=col, value=valeur)
    if gras or italique:
        c.font = Font(bold=gras, italic=italique)
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _BORDURE
    return c


def _cartouche(ws, titre_lot, hauteur):
    """Le bandeau du haut. Sa HAUTEUR change d'une feuille à l'autre —
    ce n'est pas une coquetterie : une ligne d'en-tête fixe est
    exactement ce qui ne survit pas au métré suivant."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1,
                 value=f"{MARCHE['pouvoir_adjudicateur']} — "
                        f"{MARCHE['reference']}")
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    lignes = [("Objet", MARCHE["objet"]), ("Lot", titre_lot)]
    if hauteur > 2:
        lignes.append(("Procédure", MARCHE["procedure"]))
    if hauteur > 3:
        lignes.append(("Dépôt", MARCHE["depot"]))

    for i, (label, valeur) in enumerate(lignes, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2, value=valeur)
    return 2 + len(lignes) + 1          # ligne des en-têtes de colonnes


def _feuille_de_lot(wb, titre_lot, postes, hauteur_cartouche):
    # Excel refuse un nom de feuille de plus de 31 caractères, et le
    # récapitulatif REFERENCE ces noms : `='Lot 3'!F14`. On vérifie
    # plutôt que de tronquer en silence, sinon la référence pointerait
    # une feuille qui n'existe pas. Pas de virgule non plus — une
    # référence à une feuille dont le nom en contient une reste valide,
    # mais c'est un détail de citation dont ce fichier n'a pas besoin.
    assert len(titre_lot) <= 31, f"nom de feuille trop long : {titre_lot}"
    assert "," not in titre_lot, f"virgule dans un nom de feuille : {titre_lot}"
    ws = wb.create_sheet(titre_lot)
    for lettre, _, largeur in COLONNES:
        ws.column_dimensions[lettre].width = largeur

    ligne_entete = _cartouche(ws, titre_lot, hauteur_cartouche)
    for idx, (_, libelle, _) in enumerate(COLONNES, start=1):
        _cell(ws, ligne_entete, idx, libelle, gras=True, align="center",
              fill=_FILL_ENTETE)
    ws.freeze_panes = ws.cell(row=ligne_entete + 1, column=1)

    formules = {}                       # ligne -> valeur calculée à injecter
    ligne = ligne_entete + 1
    premiere = ligne
    for code, designation, unite, quantite in postes:
        _cell(ws, ligne, 1, code, align="center")
        _cell(ws, ligne, 2, designation)
        _cell(ws, ligne, 3, unite, align="center")
        if isinstance(quantite, str) and quantite.startswith("="):
            _cell(ws, ligne, 4, quantite, fmt=FORMAT_QTE)
            if code != POSTE_SANS_CACHE:
                formules[ligne] = _evaluer(quantite)
        elif quantite is None:
            _cell(ws, ligne, 4, None, fmt=FORMAT_QTE)
        else:
            _cell(ws, ligne, 4, float(quantite), fmt=FORMAT_QTE)
        # Colonne E : la seule que le soumissionnaire remplit.
        _cell(ws, ligne, 5, None, fill=_FILL_SAISIE, fmt=FORMAT_EUR)
        _cell(ws, ligne, 6, f'=IF($E{ligne}="","",$D{ligne}*$E{ligne})',
              fmt=FORMAT_EUR)
        ligne += 1
    derniere = ligne - 1

    _cell(ws, ligne, 2, "Sous-total du lot (€ HTVA)", gras=True,
          fill=_FILL_TOTAL)
    for col in (1, 3, 4, 5):
        _cell(ws, ligne, col, None, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"=SUM(F{premiere}:F{derniere})", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_total = ligne

    ligne += 2
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
    ws.cell(row=ligne, column=1,
             value="Le soumissionnaire complète la colonne « P.U. (€) » "
                    "uniquement. Tout poste sans prix rend l'offre "
                    "irrégulière (art. 76 AR 18/04/2017).").font = Font(
        italic=True, size=9)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    return ws, ligne_total, formules


def _feuille_recap(wb, totaux):
    """Le récapitulatif REPREND les codes de lot. Le lire comme un lot
    de plus compterait chaque poste deux fois — c'est à l'utilisateur de
    le décocher, et à `feuilles_avec_postes()` de le présumer."""
    ws = wb.create_sheet(NOM_RECAP)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["F"].width = 15

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"{MARCHE['reference']} — RÉCAPITULATIF DES LOTS"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center", vertical="center")

    ligne = 3
    premiere = ligne
    for titre_lot, feuille, ligne_total in totaux:
        _cell(ws, ligne, 1, titre_lot.split(" - ")[0], align="center")
        _cell(ws, ligne, 2, titre_lot.split(" - ", 1)[1])
        _cell(ws, ligne, 6, f"='{feuille}'!F{ligne_total}", fmt=FORMAT_EUR)
        ligne += 1
    derniere = ligne - 1

    _cell(ws, ligne, 2, "TOTAL DE L'OFFRE — HTVA", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"=SUM(F{premiere}:F{derniere})", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_ht = ligne
    ligne += 1
    _cell(ws, ligne, 2, "TVA 21 %", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"=F{ligne_ht}*0.21", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_tva = ligne
    ligne += 1
    _cell(ws, ligne, 2, "TOTAL DE L'OFFRE — TVAC", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"=F{ligne_ht}+F{ligne_tva}", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# Valeurs calculées des formules
# ═══════════════════════════════════════════════════════════════════════════
#
# openpyxl écrit une formule mais JAMAIS sa valeur : un classeur qu'aucun
# tableur n'a ouvert ne contient aucun cache. Or c'est ce cache que lit
# `metre_io` — et un métré réel, sorti d'Excel, en a toujours un.
#
# Sans cette injection, le fichier d'entraînement mentirait par excès de
# difficulté : toutes ses quantités en formule seraient illisibles, ce
# qui n'arrive pas dans la vraie vie. On l'écrit donc à la main dans le
# XML, exactement comme Excel le ferait — et on en prive UN poste, pour
# que le cas d'échec existe aussi.


def _evaluer(formule):
    """La valeur d'une formule arithmétique simple : `=32*7.5` -> 240.0.

    Volontairement bornée aux nombres et aux quatre opérations : ce
    module écrit un fichier d'entraînement, il n'implémente pas Excel.

    La virgule est REFUSÉE, et pas par purisme : dans le XML d'un
    classeur, une formule est toujours en syntaxe américaine, où la
    virgule sépare des arguments. `=32*7,5` y devient « 32 fois 7 »
    suivi d'un « 5 » égaré, et le fichier s'ouvre sur un message de
    récupération de contenu.
    """
    expression = formule.lstrip("=").replace(" ", "")
    if not re.fullmatch(r"[0-9.+\-*/()]+", expression):
        raise ValueError(
            f"formule non évaluable : {formule} — une formule stockée "
            f"s'écrit en syntaxe US, décimales au point")
    return float(eval(expression))      # noqa: S307 — expression validée


def _injecter_valeurs(chemin, valeurs_par_feuille):
    """Écrit `<v>…</v>` dans les cellules de formule visées.

    valeurs_par_feuille : {nom_de_feuille: {ligne: valeur}} — la colonne
    est toujours D, celle des quantités.
    """
    if not any(valeurs_par_feuille.values()):
        return
    source = Path(chemin)
    tampon = source.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(source) as lu:
        ordre = [i.filename for i in lu.infolist()]
        contenus = {nom: lu.read(nom) for nom in ordre}
        feuilles = _feuilles_du_classeur(contenus)

    for nom_feuille, valeurs in valeurs_par_feuille.items():
        cible = feuilles.get(nom_feuille)
        if not cible:
            continue
        xml = contenus[cible].decode("utf-8")
        for ligne, valeur in valeurs.items():
            # openpyxl écrit déjà un `<v/>` VIDE derrière la formule. Il
            # faut le REMPLACER : deux `<v>` dans une même cellule sont
            # invalides, et le classeur s'ouvre alors sur un message de
            # récupération de contenu. Le défaut ne se voit qu'à
            # l'ouverture dans un tableur — la lecture Python, elle,
            # rendait la bonne quantité.
            motif = re.compile(
                rf'(<c r="D{ligne}"[^>]*>)(<f>[^<]*</f>)'
                rf'(?:<v\s*/>|<v>[^<]*</v>)?')
            xml, remplacees = motif.subn(
                lambda m, v=valeur: f"{m.group(1)}{m.group(2)}<v>{v}</v>",
                xml, count=1)
            if not remplacees:
                raise RuntimeError(
                    f"formule introuvable en D{ligne} de « {nom_feuille} »")
        contenus[cible] = xml.encode("utf-8")

    with zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as ecrit:
        for nom in ordre:
            ecrit.writestr(nom, contenus[nom])
    shutil.move(tampon, source)


def _feuilles_du_classeur(contenus):
    """{nom de feuille: chemin XML} — l'ordre de workbook.xml fait foi."""
    workbook = contenus["xl/workbook.xml"].decode("utf-8")
    noms = re.findall(r'<sheet[^>]*name="([^"]+)"', workbook)
    return {nom: f"xl/worksheets/sheet{i}.xml"
            for i, nom in enumerate(noms, start=1)}


# ═══════════════════════════════════════════════════════════════════════════
# Point d'entrée
# ═══════════════════════════════════════════════════════════════════════════


def generer_metre(chemin=NOM_FICHIER_DEFAUT):
    """Écrit le classeur et retourne (chemin, nb_postes, nb_lots)."""
    wb = Workbook()
    wb.remove(wb.active)

    totaux, a_injecter = [], {}
    for i, (titre_lot, postes) in enumerate(LOTS):
        # 2, 3 ou 4 lignes de cartouche selon le lot : l'en-tête des
        # colonnes ne tombe donc pas au même endroit d'une feuille à
        # l'autre.
        ws, ligne_total, formules = _feuille_de_lot(
            wb, titre_lot, postes, 2 + i % 3)
        totaux.append((titre_lot, ws.title, ligne_total))
        if formules:
            a_injecter[ws.title] = formules

    _feuille_recap(wb, totaux)
    wb.save(chemin)
    _injecter_valeurs(chemin, a_injecter)

    return chemin, sum(len(p) for _, p in LOTS), len(LOTS)


if __name__ == "__main__":
    cible = sys.argv[1] if len(sys.argv) > 1 else NOM_FICHIER_DEFAUT
    chemin, nb_postes, nb_lots = generer_metre(cible)
    print(f"Métré généré : {chemin}")
    print(f"  {nb_postes} postes · {nb_lots} feuilles de lot + "
          f"« {NOM_RECAP} »")
