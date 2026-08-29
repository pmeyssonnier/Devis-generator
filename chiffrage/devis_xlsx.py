"""
╔══════════════════════════════════════════════════════════════════════════╗
║  DEVIS CLIENT AU FORMAT EXCEL                                            ║
╚══════════════════════════════════════════════════════════════════════════╝

Transforme un devis calculé par `moteur.devis()` en document envoyable :
en-tête de l'entreprise, coordonnées du client, postes détaillés et regroupés
par lot, sous-totaux, TVA, conditions, cadre de signature.

C'est l'inverse exact du problème de départ : au lieu d'un « FF 1,00 » pour
dix pages de prestations, chaque poste porte son unité, sa quantité et son
prix unitaire. Le même détail sert donc au privé ET au marché public.

Le classeur porte de VRAIES FORMULES (montant = quantité × PU, sous-totaux,
TVA) : si le client négocie une quantité au téléphone, la correction se fait
dans la cellule et tout se recalcule. Pas besoin de relancer Python.

Discipline openpyxl identique à gen_metre.py : les lignes de sous-total sont
écrites au fil de la boucle, jamais insérées après coup — `insert_rows` ne
décale pas les références des formules déjà écrites.

Usage :
    from chiffrage import devis
    from chiffrage.devis_xlsx import exporter_devis

    d = devis("Rénovation façade arrière", [("40.20", 26), ("40.30", 26)])
    exporter_devis(d, "devis_2026-042.xlsx",
                   client="M. et Mme Dupont\\nRue de l'Église 12\\n1030 Schaerbeek",
                   chantier="Avenue Ernest Renan 62, 1030 Schaerbeek",
                   reference="2026-042")

En ligne de commande :
    python -m chiffrage devis 40.20:26 40.30:26 --sortie=devis.xlsx \\
        --client="M. Dupont" --chantier="Av. Renan 62"
"""

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .bibliotheque import ENTREPRISE, LOTS

FORMAT_EUR = '#,##0.00 "€"'
FORMAT_QTE = "#,##0.00"

_BORDURE = Border(*(Side(style="thin", color="B0B0B0"),) * 4)
_FILL_TITRE = PatternFill("solid", fgColor="1F3864")
_FILL_LOT = PatternFill("solid", fgColor="D9E2F3")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

COLONNES = [
    ("A", "Code", 10),
    ("B", "Désignation des ouvrages", 56),
    ("C", "Unité", 9),
    ("D", "Quantité", 12),
    ("E", "PU HTVA", 13),
    ("F", "Montant HTVA", 16),
]

# Conditions par défaut. Volontairement factuelles et courtes : tout ce qui
# engage juridiquement (acompte, pénalités, garanties) doit être décidé par le
# chef d'entreprise, pas hérité d'un modèle.
CONDITIONS_DEFAUT = [
    "Offre établie sur la base des quantités reprises ci-dessus. "
    "Toute quantité réellement exécutée différente sera adaptée au métré.",
    "Les prix sont fermes pendant la durée de validité de l'offre.",
    "Ne sont pas compris : les travaux non explicitement décrits ci-dessus.",
]

# Mention TVA : le taux réduit est conditionnel, il ne s'applique pas au choix
# du client. Le dire sur le devis évite une régularisation après facturation.
MENTION_TVA_6 = (
    "TVA 6 % : taux réduit applicable au logement privé de plus de dix ans, "
    "à usage principalement privé, facturé au consommateur final "
    "(AR n° 20, tableau A, rubrique XXXVIII). "
    "À défaut, le taux de 21 % s'applique."
)
MENTION_TVA_21 = "TVA 21 % : taux normal."


def _cell(ws, row, col, valeur, *, gras=False, fill=None, fmt=None, align=None,
          italique=False, taille=None):
    c = ws.cell(row=row, column=col, value=valeur)
    if gras or italique or taille:
        c.font = Font(bold=gras, italic=italique, size=taille or 11)
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _BORDURE
    return c


def _lignes_par_lot(lignes):
    """Regroupe les lignes du devis par lot, lots dans l'ordre croissant.

    L'ordre de saisie est conservé À L'INTÉRIEUR d'un lot : c'est en général
    l'ordre d'exécution du chantier, et le client s'y retrouve mieux.
    """
    groupes = {}
    for ligne in lignes:
        groupes.setdefault(ligne["code_ouv"].split(".")[0], []).append(ligne)
    return sorted(groupes.items())


def exporter_devis(
    d,
    chemin,
    client=None,
    chantier=None,
    reference=None,
    date_devis=None,
    validite_jours=30,
    conditions=None,
):
    """
    Écrit le devis `d` (sortie de `moteur.devis()`) dans un classeur Excel.

    client   : bloc d'adresse libre, sauts de ligne avec « \\n »
    chantier : adresse du chantier, si différente de celle du client
    reference: numéro de devis ; à défaut, dérivé de la date

    Retourne (chemin, nb_lignes).

    Un devis dont certains codes sont inconnus (`d["inconnus"]`) est REFUSÉ :
    envoyer un document dont il manque des postes est pire que ne rien
    envoyer, et l'erreur ne se verrait qu'au moment de facturer.
    """
    if d["inconnus"]:
        raise ValueError(
            "Devis incomplet, export refusé — codes d'ouvrage inconnus : "
            + ", ".join(d["inconnus"])
        )
    if not d["lignes"]:
        raise ValueError("Devis vide, rien à exporter.")

    jour = date_devis or date.today()
    reference = reference or f"{jour:%Y}-{jour:%m%d}"

    wb = Workbook()
    ws = wb.active
    ws.title = "DEVIS"
    for lettre, _, largeur in COLONNES:
        ws.column_dimensions[lettre].width = largeur

    # ── En-tête entreprise ────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = ENTREPRISE["nom"]
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"{ENTREPRISE['adresse']} · {ENTREPRISE['cp_ville']} · "
        f"{ENTREPRISE['pays']} · TVA {ENTREPRISE['tva']}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:F3")
    ws["A3"] = ENTREPRISE["activite"]
    ws["A3"].font = Font(italic=True, size=9)
    ws["A3"].alignment = Alignment(horizontal="center")

    # ── Cartouche du devis ────────────────────────────────────────────────
    ligne = 5
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
    c = ws.cell(row=ligne, column=1, value=f"DEVIS N° {reference}")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    ligne += 2

    infos = [
        ("Objet", d["nom"]),
        ("Date", f"{jour:%d/%m/%Y}"),
        ("Validité", f"{validite_jours} jours — jusqu'au "
                     f"{jour + timedelta(days=validite_jours):%d/%m/%Y}"),
    ]
    if chantier:
        infos.insert(1, ("Chantier", chantier))
    for label, valeur in infos:
        ws.cell(row=ligne, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=ligne, start_column=2, end_row=ligne, end_column=6)
        ws.cell(row=ligne, column=2, value=valeur)
        ligne += 1

    if client:
        ligne += 1
        ws.cell(row=ligne, column=1, value="Client").font = Font(bold=True)
        nb_lignes_client = client.count("\n") + 1
        ws.merge_cells(
            start_row=ligne, start_column=2,
            end_row=ligne + max(0, nb_lignes_client - 1), end_column=6,
        )
        cc = ws.cell(row=ligne, column=2, value=client)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        ligne += nb_lignes_client
    ligne += 1

    # ── Tableau des postes ────────────────────────────────────────────────
    for idx, (_, libelle, _) in enumerate(COLONNES, start=1):
        c = _cell(ws, ligne, idx, libelle, gras=True, align="center")
        c.fill = _FILL_LOT
    ligne += 1

    lignes_sous_totaux = []
    nb_lignes = 0
    for code_lot, lignes_du_lot in _lignes_par_lot(d["lignes"]):
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
        c = ws.cell(
            row=ligne, column=1,
            value=f"LOT {code_lot} — {LOTS.get(code_lot, '').upper()}",
        )
        c.font = Font(bold=True, size=11)
        c.fill = _FILL_LOT
        for col in range(1, 7):
            ws.cell(row=ligne, column=col).border = _BORDURE
        ligne += 1

        premiere = ligne
        for poste in lignes_du_lot:
            _cell(ws, ligne, 1, poste["code_ouv"], align="center")
            _cell(ws, ligne, 2, poste["libelle_ouv"])
            _cell(ws, ligne, 3, poste["unite_ouv"], align="center")
            _cell(ws, ligne, 4, float(poste["qte"]), fmt=FORMAT_QTE)
            _cell(ws, ligne, 5, poste["pu_vente"], fmt=FORMAT_EUR)
            # Formule écrite pour CETTE ligne, à sa position finale.
            _cell(ws, ligne, 6, f"=ROUND(D{ligne}*E{ligne},2)", fmt=FORMAT_EUR)
            ligne += 1
            nb_lignes += 1
        derniere = ligne - 1

        _cell(ws, ligne, 2, f"Sous-total lot {code_lot}", gras=True,
              fill=_FILL_TOTAL)
        for col in (1, 3, 4, 5):
            _cell(ws, ligne, col, None, fill=_FILL_TOTAL)
        _cell(ws, ligne, 6, f"=SUM(F{premiere}:F{derniere})", gras=True,
              fill=_FILL_TOTAL, fmt=FORMAT_EUR)
        lignes_sous_totaux.append(ligne)
        ligne += 1

    # ── Totaux ────────────────────────────────────────────────────────────
    ligne += 1
    somme = "+".join(f"F{n}" for n in lignes_sous_totaux)
    _cell(ws, ligne, 2, "TOTAL HTVA", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"={somme}", gras=True, fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_ht = ligne
    ligne += 1

    taux = d["tva_taux"]
    _cell(ws, ligne, 2, f"TVA {taux * 100:.0f} %", gras=True, fill=_FILL_TOTAL)
    _cell(ws, ligne, 6, f"=ROUND(F{ligne_ht}*{taux},2)", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR)
    ligne_tva = ligne
    ligne += 1

    _cell(ws, ligne, 2, "TOTAL À PAYER — TVAC", gras=True, fill=_FILL_TOTAL,
          taille=12)
    _cell(ws, ligne, 6, f"=F{ligne_ht}+F{ligne_tva}", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR, taille=12)
    ligne += 2

    # ── Mention TVA et conditions ─────────────────────────────────────────
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
    ws.cell(row=ligne, column=1,
            value=MENTION_TVA_6 if abs(taux - 0.06) < 1e-9 else MENTION_TVA_21
            ).font = Font(italic=True, size=9)
    ws.row_dimensions[ligne].height = 28
    ws.cell(row=ligne, column=1).alignment = Alignment(wrap_text=True,
                                                       vertical="top")
    ligne += 2

    ws.cell(row=ligne, column=1, value="Conditions").font = Font(bold=True)
    ligne += 1
    for condition in (conditions if conditions is not None else CONDITIONS_DEFAUT):
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
        c = ws.cell(row=ligne, column=1, value=f"·  {condition}")
        c.font = Font(size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ligne].height = 24
        ligne += 1

    # ── Signature ─────────────────────────────────────────────────────────
    ligne += 2
    ws.cell(row=ligne, column=1, value="Pour accord, le client").font = Font(bold=True)
    ws.cell(row=ligne, column=5, value="Pour l'entreprise").font = Font(bold=True)
    ligne += 1
    ws.cell(row=ligne, column=1,
            value="Date et signature, précédées de « lu et approuvé »"
            ).font = Font(italic=True, size=9)
    ws.cell(row=ligne, column=5, value=ENTREPRISE["nom"]).font = Font(size=9)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1

    wb.save(chemin)
    return chemin, nb_lignes
