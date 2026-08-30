"""
╔══════════════════════════════════════════════╗
║  DOSSIER DE JUSTIFICATION DE PRIX                                    ║
╚══════════════════════════════════════════════╝

Quand un pouvoir adjudicateur juge un prix anormalement bas, il doit
demander une justification écrite avant d'écarter l'offre (art. 36
AR 18/04/2017). Le délai de réponse est court.

Ce module produit la pièce attendue : pour chaque poste mis en cause,
la décomposition complète du prix — ressource par ressource,
quantité, prix unitaire, déboursés par nature, coefficient, prix de
vente. Rien n'est rédigé pour l'occasion : ce sont les chiffres qui
ont SERVI à établir l'offre, sortis de la même bibliothèque. C'est
précisément ce qui rend la justification crédible.

Le classeur porte une lettre d'accompagnement à relire et à signer.
Elle est volontairement sobre et factuelle : une justification de
prix n'est pas un argumentaire commercial.

Usage :
    from chiffrage.justification_xlsx import exporter_justification
    exporter_justification(["40.20", "70.50"], "justification.xlsx",
                            marche={"reference": "CSC 2026-TP-0147", ...})
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .bibliotheque import (
    COMPOSITION,
    ENTREPRISE,
    LOTS,
    OUVRAGES_PAR_CODE,
    PARAMS,
    RESSOURCES_PAR_CODE,
)
from .moteur import calcul_bordereau, coefficient_k

FORMAT_EUR = '#,##0.00 "€"'
_BORDURE = Border(*(Side(style="thin", color="B0B0B0"),) * 4)
_FILL_TITRE = PatternFill("solid", fgColor="1F3864")
_FILL_ENTETE = PatternFill("solid", fgColor="D9E2F3")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

# Lettre type. Factuelle, sans plaidoyer : les chiffres des onglets
# suivants font le travail. À relire et signer avant envoi.
CORPS_LETTRE = [
    "Monsieur, Madame,",
    "",
    "Faisant suite à votre demande de justification des prix unitaires "
    "de notre offre citée en objet, vous trouverez ci-après la "
    "décomposition détaillée des postes concernés.",
    "",
    "Chaque prix est établi analytiquement à partir de notre "
    "bibliothèque de prix unitaires : main-d'œuvre valorisée au coût "
    "entreprise complet, matériaux aux prix d'achat rendus chantier, "
    "matériel au coût de location. Le déboursé sec ainsi obtenu est "
    "majoré d'un coefficient couvrant les frais généraux, les frais "
    "de chantier, les aléas et la marge.",
    "",
    "Les onglets suivants détaillent, pour chaque poste, l'ensemble "
    "des ressources mises en œuvre par unité d'ouvrage, leurs "
    "quantités et leurs prix unitaires.",
    "",
    "Nous restons à votre disposition pour tout complément.",
    "",
    "Veuillez agréer, Monsieur, Madame, l'expression de nos "
    "salutations distinguées.",
]


def _cell(ws, row, col, valeur, *, gras=False, fill=None, fmt=None,
           align=None, italique=False, taille=None, bordure=True):
    c = ws.cell(row=row, column=col, value=valeur)
    if gras or italique or taille:
        c.font = Font(bold=gras, italic=italique, size=taille or 11)
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center")
    if bordure:
        c.border = _BORDURE
    return c


def _feuille_courrier(wb, codes, marche, jour):
    ws = wb.active
    ws.title = "Courrier"
    for lettre, largeur in zip("ABCDEF", (14, 18, 18, 14, 14, 14)):
        ws.column_dimensions[lettre].width = largeur

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = ENTREPRISE["nom"]
    c.font = Font(bold=True, size=15, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (f"{ENTREPRISE['adresse']} · {ENTREPRISE['cp_ville']} · "
                 f"TVA {ENTREPRISE['tva']}")
    ws["A2"].alignment = Alignment(horizontal="center")

    ligne = 4
    for label, valeur in [
        ("Destinataire", (marche or {}).get("pouvoir_adjudicateur", "")),
        ("Marché", (marche or {}).get("reference", "")),
        ("Objet", "Justification des prix unitaires — "
                   "art. 36 AR du 18/04/2017"),
        ("Date", f"{jour:%d/%m/%Y}"),
        ("Postes concernés", ", ".join(codes)),
    ]:
        ws.cell(row=ligne, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=ligne, start_column=2, end_row=ligne,
                        end_column=6)
        cellule = ws.cell(row=ligne, column=2, value=valeur)
        cellule.alignment = Alignment(wrap_text=True, vertical="top")
        ligne += 1

    ligne += 1
    for paragraphe in CORPS_LETTRE:
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne,
                        end_column=6)
        cellule = ws.cell(row=ligne, column=1, value=paragraphe)
        cellule.alignment = Alignment(wrap_text=True, vertical="top")
        if paragraphe:
            ws.row_dimensions[ligne].height = 30
        ligne += 1

    ligne += 2
    ws.cell(row=ligne, column=4, value="Pour l'entreprise").font = Font(bold=True)
    ws.cell(row=ligne + 1, column=4, value=ENTREPRISE["nom"]).font = Font(size=9)
    ws.cell(row=ligne + 3, column=1,
            value="Lettre à relire et à signer avant envoi.").font = Font(
                italic=True, size=9)


def _feuille_poste(wb, code_ouv, bordereau, params):
    ouv = OUVRAGES_PAR_CODE[code_ouv]
    ref = bordereau[code_ouv]
    p = params or PARAMS

    # Le nom d'onglet Excel n'admet ni ':' ni '/' ni plus de 31 signes.
    ws = wb.create_sheet(f"Poste {code_ouv}"[:31])
    for lettre, largeur in zip("ABCDEFG", (12, 40, 8, 10, 11, 11, 13)):
        ws.column_dimensions[lettre].width = largeur

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{code_ouv} — {ouv['libelle_ouv']}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _FILL_TITRE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    ws["A2"] = (f"Lot {ouv['lot']} · {LOTS.get(ouv['lot'], '')} · "
                 f"unité : {ouv['unite_ouv']}")
    ws["A2"].alignment = Alignment(horizontal="center")

    ligne = 4
    for idx, entete in enumerate(
        ["Ressource", "Désignation", "Type", "Unité", "Quantité",
         "PU d'achat", "Montant"], start=1
    ):
        _cell(ws, ligne, idx, entete, gras=True, fill=_FILL_ENTETE,
              align="center")
    ligne += 1

    premiere = ligne
    for comp in COMPOSITION:
        if comp["code_ouv"] != code_ouv:
            continue
        res = RESSOURCES_PAR_CODE[comp["code_res"]]
        _cell(ws, ligne, 1, res["code_res"], align="center")
        _cell(ws, ligne, 2, res["libelle_res"])
        _cell(ws, ligne, 3, res["type_res"], align="center")
        _cell(ws, ligne, 4, res["unite_res"], align="center")
        _cell(ws, ligne, 5, comp["qte_res"], fmt="0.000")
        _cell(ws, ligne, 6, res["pu_res"], fmt=FORMAT_EUR)
        # Formule ancrée sur SA ligne : le destinataire doit pouvoir
        # refaire le calcul dans son tableur.
        _cell(ws, ligne, 7, f"=E{ligne}*F{ligne}", fmt=FORMAT_EUR)
        ligne += 1
    derniere = ligne - 1

    ligne += 1
    for libelle, valeur, gras in [
        ("Déboursé main-d'œuvre", ref["deb_mo"], False),
        ("Déboursé matériaux", ref["deb_mat"], False),
        ("Déboursé matériel", ref["deb_eqp"], False),
        ("DÉBOURSÉ SEC", None, True),
    ]:
        _cell(ws, ligne, 2, libelle, gras=gras,
              fill=_FILL_TOTAL if gras else None)
        if valeur is None:
            _cell(ws, ligne, 7, f"=SUM(G{premiere}:G{derniere})", gras=True,
                  fill=_FILL_TOTAL, fmt=FORMAT_EUR)
            ligne_debourse = ligne
        else:
            _cell(ws, ligne, 7, round(valeur, 2), fmt=FORMAT_EUR)
        ligne += 1

    ligne += 1
    _cell(ws, ligne, 2, f"Heures de main-d'œuvre par {ouv['unite_ouv']}")
    _cell(ws, ligne, 7, ref["heures_mo"], fmt="0.000")
    ligne += 2

    for libelle, taux in [
        ("Frais généraux", p["fg"]), ("Frais de chantier", p["fc"]),
        ("Aléas", p["aleas"]), ("Marge", p["marge"]),
    ]:
        _cell(ws, ligne, 2, libelle)
        _cell(ws, ligne, 7, taux, fmt="0.0 %")
        ligne += 1

    _cell(ws, ligne, 2, "Coefficient K", gras=True)
    _cell(ws, ligne, 7, round(coefficient_k(p), 4), gras=True, fmt="0.0000")
    ligne_k = ligne
    ligne += 2

    _cell(ws, ligne, 2, f"PRIX DE VENTE HTVA / {ouv['unite_ouv']}", gras=True,
          fill=_FILL_TOTAL, taille=12)
    _cell(ws, ligne, 7, f"=ROUND(G{ligne_debourse}*G{ligne_k},2)", gras=True,
          fill=_FILL_TOTAL, fmt=FORMAT_EUR, taille=12)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1


def exporter_justification(codes, chemin, marche=None, params=None,
                            date_reponse=None):
    """
    Écrit le dossier : une lettre, puis un onglet par poste.

    codes : les codes d'ouvrage mis en cause par le PA.
    Retourne (chemin, nb_postes).

    Un code inconnu est refusé plutôt qu'ignoré : un dossier de
    justification amputé d'un poste demandé se retourne contre son
    auteur.
    """
    codes = list(dict.fromkeys(codes))       # dédoublonne, garde l'ordre
    if not codes:
        raise ValueError("Aucun poste à justifier.")
    inconnus = [c for c in codes if c not in OUVRAGES_PAR_CODE]
    if inconnus:
        raise ValueError(
            "Postes inconnus, dossier non produit : " + ", ".join(inconnus))

    bordereau = calcul_bordereau(params)
    wb = Workbook()
    _feuille_courrier(wb, codes, marche, date_reponse or date.today())
    for code in codes:
        _feuille_poste(wb, code, bordereau, params)
    wb.save(chemin)
    return chemin, len(codes)
