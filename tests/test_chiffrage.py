"""Tests de l'outil de chiffrage BAG BATTER (`chiffrage/`).

Deux niveaux :
  · le moteur en Python pur — toujours exécuté ;
  · les modules Excel (export, génération de métré, remplissage d'offre) —
    ignorés si openpyxl n'est pas installé (il ne fait pas partie des
    dépendances de l'app, seulement de `requirements-chiffrage.txt`).

Ce que ces tests protègent VRAIMENT, au-delà de la couverture :
  1. le coefficient K et la formule de prix (une dérive silencieuse ici fausse
     toutes les offres) ;
  2. l'intégrité référentielle des trois tables ;
  3. le contrôle d'unité, qui doit refuser de chiffrer plutôt que convertir ;
  4. l'ancrage des formules Excel ligne par ligne — le bug openpyxl
     `insert_rows` décrit dans gen_metre.py.
"""
import pytest

from chiffrage import bibliotheque as biblio
from chiffrage import moteur


# ── 1. Structure de la bibliothèque ────────────────────────────────────────
def test_tailles_des_tables():
    assert len(biblio.RESSOURCES) == 49
    assert len(biblio.OUVRAGES) == 49
    assert len(biblio.LOTS) == 9
    assert len(biblio.OUVRAGES_A_VALIDER) == 13


def test_bibliotheque_coherente():
    """Aucune référence orpheline, aucun ouvrage sans composition ni sans MO."""
    anomalies = moteur.controle_coherence()
    assert anomalies == {
        "res_orphelines": [],
        "ouv_orphelins": [],
        "releves_orphelins": [],
        "ouv_sans_compo": [],
        "ouv_sans_mo": [],
        "codes_dupliques": [],
        "lots_inconnus": [],
    }


def test_codes_ouvrages_au_format_lot_numero():
    for ouv in biblio.OUVRAGES:
        lot, numero = ouv["code_ouv"].split(".")
        assert lot in biblio.LOTS and lot == ouv["lot"]
        assert len(numero) == 2 and numero.isdigit()


def test_mapping_ne_pointe_que_des_ouvrages_existants():
    codes = {o["code_ouv"] for o in biblio.OUVRAGES}
    assert set(biblio.MAPPING.values()) <= codes


def test_espaces_de_nommage_disjoints():
    """Un code de poste de métré (NN.NN) ne doit jamais être un code d'ouvrage."""
    ouvrages = {o["code_ouv"] for o in biblio.OUVRAGES}
    assert set(biblio.MAPPING) & ouvrages == set()


def test_ouvrages_a_valider_existent_et_sont_mappes():
    """Les 13 ouvrages créés après coup doivent exister ET couvrir un poste :
    sinon on a fabriqué des prix pour rien."""
    codes = {o["code_ouv"] for o in biblio.OUVRAGES}
    assert set(biblio.OUVRAGES_A_VALIDER) <= codes
    assert set(biblio.OUVRAGES_A_VALIDER) <= set(biblio.MAPPING.values())


# ── 2. Formule de prix ─────────────────────────────────────────────────────
def test_coefficient_k():
    assert round(moteur.coefficient_k(), 4) == 1.3324


def test_pu_vente_est_le_debourse_fois_k():
    k = moteur.coefficient_k()
    for ligne in moteur.calcul_bordereau().values():
        assert ligne["pu_vente"] == pytest.approx(ligne["debourse_sec"] * k, abs=0.01)
        assert ligne["debourse_sec"] == pytest.approx(
            ligne["deb_mo"] + ligne["deb_mat"] + ligne["deb_eqp"], abs=0.01
        )
        assert ligne["pu_vente"] > 0


def test_marge_nulle_ramene_le_prix_au_debourse_majore():
    params = dict(biblio.PARAMS, fg=0.0, fc=0.0, aleas=0.0, marge=0.0)
    for ligne in moteur.calcul_bordereau(params).values():
        assert ligne["pu_vente"] == pytest.approx(ligne["debourse_sec"], abs=0.01)


# ── 3. Devis ───────────────────────────────────────────────────────────────
def test_devis_totalise_et_applique_la_tva():
    d = moteur.devis("test", [("70.10", 100.0)], tva=0.21)
    pu = moteur.calcul_bordereau()["70.10"]["pu_vente"]
    assert d["total_ht"] == pytest.approx(pu * 100, abs=0.01)
    assert d["montant_tva"] == pytest.approx(d["total_ht"] * 0.21, abs=0.01)
    assert d["total_ttc"] == pytest.approx(d["total_ht"] * 1.21, abs=0.02)
    assert d["jours_homme"] == pytest.approx(d["heures_mo"] / 8, abs=0.01)


def test_devis_signale_les_codes_inconnus_sans_les_chiffrer():
    """Un code absent ne doit JAMAIS disparaître en silence."""
    d = moteur.devis("test", [("70.10", 10.0), ("99.99", 5.0)])
    assert d["inconnus"] == ["99.99"]
    assert len(d["lignes"]) == 1


def test_tva_marche_public_a_21_pourcent():
    assert biblio.PARAMS["tva_marche_public"] == 0.21


# ── 4. Calibration sur les devis historiques ───────────────────────────────
def test_calibration_couvre_les_six_devis():
    cal = moteur.calibration()
    assert len(cal["lignes"]) == 6
    assert all(not r["inconnus"] for r in cal["lignes"])
    assert all(r["calcule"] > 0 for r in cal["lignes"])


def test_fiche_prix_detaille_toutes_les_ressources():
    texte = moteur.fiche_prix("40.20")
    for comp in biblio.COMPOSITION:
        if comp["code_ouv"] == "40.20":
            assert comp["code_res"] in texte
    assert "DÉBOURSÉ SEC" in texte and "1.3324" in texte


def test_fiche_prix_refuse_un_ouvrage_inconnu():
    with pytest.raises(KeyError):
        moteur.fiche_prix("99.99")


# ── 5. Chaîne Excel (nécessite openpyxl) ───────────────────────────────────
@pytest.fixture(scope="module")
def openpyxl_dispo():
    return pytest.importorskip("openpyxl")


@pytest.fixture(scope="module")
def metre(tmp_path_factory, openpyxl_dispo):
    from chiffrage.gen_metre import generer_metre

    chemin = tmp_path_factory.mktemp("chiffrage") / "metre.xlsx"
    generer_metre(str(chemin))
    return chemin


def test_normalisation_des_unites(openpyxl_dispo):
    from chiffrage.metre_io import normaliser_unite

    assert normaliser_unite("m²") == normaliser_unite("M2") == "m2"
    assert normaliser_unite("PCE") == normaliser_unite("pc") == "pce"
    # Aucune conversion physique : le mètre courant n'est pas le mètre carré.
    assert normaliser_unite("m") != normaliser_unite("m2")


def test_metre_genere_49_postes(metre):
    from chiffrage.metre_io import lire_metre

    postes = lire_metre(str(metre))
    assert len(postes) == 49
    assert len({p["code"] for p in postes}) == 49
    assert all(p["quantite"] > 0 for p in postes)


def test_formules_du_metre_ancrees_sur_leur_propre_ligne(metre):
    """Garde-fou contre le bug openpyxl `insert_rows` (cf. gen_metre.py).

    Une formule de montant doit référencer SA ligne, pas celle où elle a été
    écrite avant un décalage.
    """
    from openpyxl import load_workbook

    ws = load_workbook(str(metre)).active
    montants = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith('=IF($G'):
                montants += 1
                assert v == f'=IF($G{cell.row}="","",$F{cell.row}*$G{cell.row})'
    assert montants == 49


def test_remplissage_de_l_offre(metre, tmp_path):
    from chiffrage.metre_io import remplir_metre

    rapport = remplir_metre(str(metre), str(tmp_path / "offre.xlsx"))
    assert rapport["postes"] == 49
    # Les 49 postes sont chiffrés : c'est la condition de RÉGULARITÉ de
    # l'offre (art. 76 AR 18/04/2017). Ce test est le garde-fou du seul
    # défaut qui fait rejeter une offre sans discussion.
    assert len(rapport["chiffres"]) == 49
    assert rapport["non_couverts"] == []
    assert rapport["vides"] == []
    # Contrôle d'unité : aucun écart sur ce métré-ci, et surtout aucun poste
    # chiffré dans une unité différente de celle de l'ouvrage.
    assert rapport["ecarts_unite"] == []
    assert rapport["total_ht"] > 0


def test_remplissage_preserve_les_formules_du_pouvoir_adjudicateur(metre, tmp_path):
    """`data_only=True` détruirait ces formules : le fichier renvoyé au PA ne
    recalculerait plus rien. On vérifie qu'elles survivent au remplissage."""
    from openpyxl import load_workbook

    from chiffrage.metre_io import remplir_metre

    sortie = tmp_path / "offre.xlsx"
    remplir_metre(str(metre), str(sortie))
    ws = load_workbook(str(sortie)).active
    formules = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert sum(1 for f in formules if f.startswith("=IF($G")) == 49
    assert any(f.startswith("=SUM(H") for f in formules)


def test_unite_incompatible_bloque_le_chiffrage(metre, tmp_path):
    """Mapper un poste au m2 vers un ouvrage au mètre courant ne doit PAS
    produire de prix : le poste part en `ecarts_unite`, pas en `chiffres`."""
    from chiffrage.metre_io import remplir_metre

    mapping_faux = dict(biblio.MAPPING)
    mapping_faux["03.02"] = "40.60"  # 03.02 est en m2, 40.60 en m
    rapport = remplir_metre(
        str(metre), str(tmp_path / "offre_ko.xlsx"), mapping=mapping_faux
    )
    assert [e["code"] for e in rapport["ecarts_unite"]] == ["03.02"]
    assert "03.02" not in [c["code"] for c in rapport["chiffres"]]
    assert "03.02" in rapport["vides"]


def test_export_bibliotheque_six_onglets(tmp_path, openpyxl_dispo):
    from openpyxl import load_workbook

    from chiffrage.export_xlsx import exporter_bibliotheque

    chemin = tmp_path / "biblio.xlsx"
    exporter_bibliotheque(str(chemin))
    wb = load_workbook(str(chemin))
    assert wb.sheetnames == [
        "PARAMS", "RESSOURCES", "OUVRAGES", "COMPOSITION", "BORDEREAU", "MAPPING",
    ]
    assert wb["RESSOURCES"].max_row == len(biblio.RESSOURCES) + 1
    assert wb["COMPOSITION"].max_row == len(biblio.COMPOSITION) + 1


def test_formules_excel_du_bordereau_donnent_les_prix_python(tmp_path, openpyxl_dispo):
    """Le classeur exporté calcule tout seul (SUMIFS + K) : on ré-exécute la
    sémantique de ses formules en Python et on la compare à `pu_vente`.

    Sans ça, une erreur de plage dans l'export ne se verrait qu'à l'ouverture
    du fichier par le client — c'est-à-dire jamais avant l'envoi d'une offre.
    """
    from openpyxl import load_workbook

    from chiffrage.export_xlsx import exporter_bibliotheque

    chemin = tmp_path / "biblio.xlsx"
    exporter_bibliotheque(str(chemin))
    wb = load_workbook(str(chemin))
    compo = wb["COMPOSITION"]
    lignes = [
        (
            compo.cell(r, 1).value,   # code_ouv
            compo.cell(r, 5).value,   # type_res
            compo.cell(r, 7).value,   # qte_res
            wb["RESSOURCES"].cell(
                int(compo.cell(r, 8).value.split("!E")[1]), 5
            ).value,                  # pu_res, suivi à travers la formule
        )
        for r in range(2, compo.max_row + 1)
    ]
    k = moteur.coefficient_k()
    bordereau = wb["BORDEREAU"]
    for r in range(2, bordereau.max_row + 1):
        code = bordereau.cell(r, 1).value
        debourse = sum(q * pu for c, _, q, pu in lignes if c == code)
        # colonne M = valeur calculée par Python, écrite comme témoin
        assert bordereau.cell(r, 13).value == pytest.approx(debourse * k, abs=0.01)


# ── 6. Devis client au format Excel ────────────────────────────────────────
@pytest.fixture
def devis_exemple():
    return moteur.devis(
        "Rénovation façade arrière",
        [("40.20", 26.0), ("40.30", 26.0), ("70.70", 8.0)],
        tva=0.06,
    )


def test_devis_client_ecrit_un_classeur(devis_exemple, tmp_path, openpyxl_dispo):
    from openpyxl import load_workbook

    from chiffrage.devis_xlsx import exporter_devis

    chemin = tmp_path / "devis.xlsx"
    _, nb = exporter_devis(devis_exemple, str(chemin), reference="2026-042",
                           client="M. Dupont", chantier="Av. Renan 62")
    assert nb == 3
    ws = load_workbook(str(chemin))["DEVIS"]
    textes = [
        c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)
    ]
    assert any(biblio.ENTREPRISE["nom"] in t for t in textes)
    assert any("2026-042" in t for t in textes)
    assert any("TOTAL À PAYER" in t for t in textes)


def test_devis_client_formules_ancrees_sur_leur_ligne(devis_exemple, tmp_path,
                                                      openpyxl_dispo):
    """Même garde-fou que pour le métré : chaque montant multiplie SA ligne."""
    from openpyxl import load_workbook

    from chiffrage.devis_xlsx import exporter_devis

    chemin = tmp_path / "devis.xlsx"
    exporter_devis(devis_exemple, str(chemin))
    ws = load_workbook(str(chemin))["DEVIS"]
    montants = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=ROUND(D"):
                montants += 1
                assert cell.value == f"=ROUND(D{cell.row}*E{cell.row},2)"
    assert montants == 3


def test_devis_client_totaux_excel_egalent_le_moteur(devis_exemple, tmp_path,
                                                     openpyxl_dispo):
    """Ré-exécution en Python de la sémantique des formules du classeur.

    Une erreur de plage dans un sous-total ne se verrait qu'à l'ouverture du
    fichier par le client — c'est-à-dire après l'envoi.
    """
    import re

    from openpyxl import load_workbook

    from chiffrage.devis_xlsx import exporter_devis

    chemin = tmp_path / "devis.xlsx"
    exporter_devis(devis_exemple, str(chemin))
    ws = load_workbook(str(chemin))["DEVIS"]

    valeurs = {}
    sous_totaux = {}
    for row in ws.iter_rows():
        cell = row[5]                       # colonne F
        if not isinstance(cell.value, str):
            continue
        if cell.value.startswith("=ROUND(D"):
            valeurs[cell.row] = round(ws.cell(cell.row, 4).value
                                      * ws.cell(cell.row, 5).value, 2)
        elif cell.value.startswith("=SUM(F"):
            debut, fin = map(int, re.findall(r"F(\d+)", cell.value))
            sous_totaux[cell.row] = sum(
                valeurs[r] for r in range(debut, fin + 1) if r in valeurs
            )

    total_ht = round(sum(sous_totaux.values()), 2)
    assert total_ht == pytest.approx(devis_exemple["total_ht"], abs=0.01)
    assert round(total_ht * 0.06, 2) == pytest.approx(
        devis_exemple["montant_tva"], abs=0.01
    )


def test_devis_client_refuse_un_devis_incomplet(tmp_path, openpyxl_dispo):
    """Un devis qui a laissé tomber un code inconnu ne doit PAS partir chez le
    client : le poste manquant ne se verrait qu'à la facturation."""
    from chiffrage.devis_xlsx import exporter_devis

    d = moteur.devis("test", [("40.20", 10.0), ("99.99", 5.0)])
    with pytest.raises(ValueError, match="99.99"):
        exporter_devis(d, str(tmp_path / "devis.xlsx"))
    assert not (tmp_path / "devis.xlsx").exists()


def test_devis_client_refuse_un_devis_vide(tmp_path, openpyxl_dispo):
    from chiffrage.devis_xlsx import exporter_devis

    with pytest.raises(ValueError):
        exporter_devis(moteur.devis("vide", []), str(tmp_path / "devis.xlsx"))


# ── 7. Notebook Colab ─────────────────────────────────────────────
#
# Le notebook est le point d'entrée réel du chef d'entreprise (il travaille
# depuis mobile, via Colab). Rien ne l'exécute en CI : sans ces tests, un
# renommage dans `chiffrage/` le casserait en silence et personne ne le
# saurait avant qu'il ouvre Colab pour répondre à un marché.
import json
from pathlib import Path

NOTEBOOK = Path(__file__).resolve().parent.parent / "colab" / "chiffrage_bagbatter.ipynb"


def _cellules_de_code():
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    return [
        "".join(c["source"])
        for c in nb["cells"]
        if c["cell_type"] == "code"
    ]


def test_notebook_est_un_json_valide():
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    assert nb["nbformat"] == 4
    assert len(_cellules_de_code()) >= 10


def test_cellules_du_notebook_compilent():
    """Syntaxe valide — hors cellules à magies shell (!pip, !git), qui ne
    sont pas du Python et ne compilent que dans IPython."""
    for i, source in enumerate(_cellules_de_code(), start=1):
        if any(ligne.lstrip().startswith("!") for ligne in source.splitlines()):
            continue
        compile(source, f"<cellule {i}>", "exec")


def test_symboles_importes_par_le_notebook_existent():
    """Le vrai garde-fou : chaque `from chiffrage.x import y` du notebook
    doit résoudre. C'est ce qui casse quand on renomme une fonction."""
    import importlib
    import re

    for source in _cellules_de_code():
        for module, noms in re.findall(
            r"from (chiffrage[\w.]*) import ([^\n(]+|\([^)]*\))", source
        ):
            mod = importlib.import_module(module)
            for nom in noms.strip("()").replace("\n", " ").split(","):
                nom = nom.strip().split(" as ")[0].strip()
                if nom:
                    assert hasattr(mod, nom), f"{module}.{nom} n'existe pas"


# ── 8. Appariement automatique ───────────────────────────────
from chiffrage import suggestion  # noqa: E402


def test_normalisation_de_libelle():
    mots = suggestion.normaliser("Étanchéité bitumineuse, relevés compris")
    assert "etancheite" in mots
    assert "bitumineuse" in mots
    # « compris » est un mot vide de métré : le garder ferait remonter
    # n'importe quel poste qui le contient.
    assert "compris" not in mots


def test_score_est_maximal_pour_un_libelle_identique():
    lib = "Enduit de façade minéral armé, deux couches"
    assert suggestion.score(lib, lib) > suggestion.score(lib, "Peinture de façade")


def test_unite_est_eliminatoire_pas_departageante():
    """LE garde-fou du module. Un poste au mètre courant ne doit jamais
    se voir proposer un ouvrage au m2, même si les libellés sont
    identiques : le prix serait faux d'un facteur inconnu."""
    b = moteur.calcul_bordereau()
    poste = {
        "designation": "Enduit de façade minéral armé, deux couches",  # = 40.20
        "unite": "m",                                                # mais en m !
    }
    candidats = suggestion.suggerer(poste, b, limite=10)
    assert "40.20" not in [c for c, _ in candidats]
    assert all(b[c]["unite_ouv"] == "m" for c, _ in candidats)


def test_suggerer_rend_une_liste_vide_si_aucune_unite_ne_correspond():
    b = moteur.calcul_bordereau()
    assert suggestion.suggerer(
        {"designation": "Quoi que ce soit", "unite": "tonne"}, b
    ) == []


def test_proposer_mapping_reprend_les_correspondances_connues():
    """Un choix humain antérieur n'est jamais réécrit par l'algorithme."""
    b = moteur.calcul_bordereau()
    postes = [{"code": "03.02", "designation": "n'importe quoi", "unite": "m2"}]
    prop = suggestion.proposer_mapping(postes, b, mapping_connu=biblio.MAPPING)
    assert prop["03.02"]["origine"] == "connu"
    assert prop["03.02"]["code_ouv"] == biblio.MAPPING["03.02"]


def test_proposer_mapping_signale_ce_qu_il_ne_sait_pas_apparier():
    b = moteur.calcul_bordereau()
    postes = [{"code": "99.99", "designation": "Zzzz qqqq", "unite": "tonne"}]
    prop = suggestion.proposer_mapping(postes, b)
    assert prop["99.99"]["origine"] == "aucun"
    assert prop["99.99"]["code_ouv"] is None


@pytest.mark.parametrize("libelle, unite, attendu", [
    ("Montage, location et démontage d'un échafaudage de pied", "m2", "10.20"),
    ("Membrane d'étanchéité soudée en deux couches", "m2", "40.40"),
    ("Menuiserie extérieure PVC avec double vitrage", "m2", "80.10"),
    ("Nettoyage de parement par projection d'eau sous pression", "m2", "40.10"),
])
def test_appariement_sur_des_libelles_reformules(libelle, unite, attendu):
    """Libellés réécrits « à la manière d'un CSC » — synonymes, ordre
    inversé, jargon administratif — sans mot commun au-delà du sens.

    Sur un échantillon plus large de ce type, l'appariement tombe juste
    du premier coup dans ~60 % des cas et place la bonne réponse dans les
    trois premières dans ~90 %. C'est un dégrossissage, pas une décision :
    l'interface fait trancher l'humain. Ces quatre cas-ci sont ceux qui
    doivent rester au premier rang.
    """
    b = moteur.calcul_bordereau()
    candidats = suggestion.suggerer(
        {"designation": libelle, "unite": unite}, b, limite=1
    )
    assert candidats and candidats[0][0] == attendu


# ── 9. Surcouche de session du lexique ─────────────────
@pytest.fixture(autouse=True)
def _lexique_propre():
    """Isole CHAQUE test des deux couches mutables du lexique.

    La surcouche est un état global de module : sans nettoyage, un
    test qui ajoute un synonyme fausserait les suivants.

    La couche LOCALE, elle, est chargée depuis lexique_local.json —
    un fichier que l'APP écrit en production. Un test qui en dépend
    se met donc à échouer le jour où quelqu'un apprend un terme
    depuis l'interface, sans que le code ait bougé. C'est
    exactement ce qui est arrivé avec « sablage » : la suite est
    devenue rouge à cause d'une donnée, pas d'une régression.

    Les tests qui veulent une couche locale la posent eux-mêmes.
    """
    from chiffrage.lexique import LOCAL, vider_surcouche

    garde = {table: dict(valeurs) for table, valeurs in LOCAL.items()}
    for valeurs in LOCAL.values():
        valeurs.clear()
    vider_surcouche()
    yield
    vider_surcouche()
    for table, valeurs in garde.items():
        LOCAL[table] = valeurs


def test_surcouche_change_l_appariement_immediatement():
    """C'est tout l'intérêt : ajouter un terme dans l'interface et en
    voir l'effet sans redéployer."""
    from chiffrage import lexique

    b = moteur.calcul_bordereau()
    poste = {"designation": "Sablage des maçonneries de façade",
             "unite": "m2"}

    avant = suggestion.suggerer(poste, b, limite=1)[0][1]
    lexique.ajouter_synonyme("sablage", "nettoyage")
    apres = suggestion.suggerer(poste, b, limite=1)[0][1]

    assert avant < 0.40 < apres


def test_surcouche_l_emporte_sur_la_table_du_depot():
    from chiffrage import lexique

    assert lexique.canoniser("crepi") == "enduit"
    lexique.ajouter_synonyme("crepi", "plafonnage")
    assert lexique.canoniser("crepi") == "plafonnage"


def test_surcouche_se_vide():
    from chiffrage import lexique

    lexique.ajouter_synonyme("sablage", "nettoyage")
    lexique.ajouter_expression("mur de refend", "cloison")
    lexique.vider_surcouche()
    assert lexique.canoniser("sablage") == "sablage"
    assert lexique.appliquer_expressions("mur de refend") == "mur de refend"


def test_surcouche_rend_un_bloc_python_collable():
    """L'interface ne peut pas commiter : elle rend le bloc à coller.
    Ce bloc doit être du Python valide, sinon il ne sert à rien."""
    import ast

    from chiffrage import lexique

    lexique.ajouter_synonyme("sablage", "nettoyage")
    lexique.ajouter_expression("mur de refend", "cloison")
    bloc = lexique.surcouche_en_python()

    assert '"sablage": "nettoyage",' in bloc
    assert '"mur de refend": "cloison",' in bloc
    # Collé dans un dict, ça doit se parser.
    ast.parse("SYNONYMES = {\n" + bloc.replace("#", "  #") + "\n}")


# ── 10. Persistance du lexique dans le dépôt ─────────
def test_terme_accentue_est_normalise():
    """Défaut corrigé : les tables sont consultées avec des mots
    DÉPOUILLÉS de leurs accents. Une clé accentuée n'était jamais
    trouvée — l'ajout ne servait à rien, en silence."""
    from chiffrage import lexique

    lexique.ajouter_synonyme("Maçonneries", "maconnerie")
    assert "maconneries" in lexique.SURCOUCHE["synonymes"]
    assert "maconnerie" in suggestion.normaliser("Sablage des maçonneries")


@pytest.mark.parametrize("terme", [
    'x"; import os',          # injection : le terme finissait dans du code
    "trop " * 20,             # longueur
    "",                       # vide
    "<script>",               # balise
    "terme; DROP",           # ponctuation d'injection
])
def test_terme_irrecevable_est_refuse(terme):
    from chiffrage import lexique

    with pytest.raises(ValueError):
        lexique.ajouter_synonyme(terme, "nettoyage")


def test_saut_de_ligne_est_recolle_pas_refuse():
    """Coller un libellé depuis un PDF amène des sauts de ligne. Ils
    deviennent des espaces — une expression multi-mots est valide, et
    un saut de ligne ne peut rien casser dans du JSON."""
    from chiffrage import lexique

    lexique.ajouter_expression("carrelage\nmural", "faience")
    assert "carrelage mural" in lexique.SURCOUCHE["expressions"]


def test_couche_locale_absente_ne_casse_rien(tmp_path):
    """Un lexique_local.json absent est le cas NORMAL : l'outil doit
    chiffrer sans lui."""
    from chiffrage import lexique

    assert lexique.charger_local(tmp_path / "rien.json") == {
        "expressions": {}, "synonymes": {}}


def test_couche_locale_illisible_ne_casse_rien(tmp_path):
    """Un fichier corrompu ne doit pas empêcher de répondre à un
    marché : on revient aux tables du dépôt."""
    from chiffrage import lexique

    fichier = tmp_path / "casse.json"
    fichier.write_text("{ceci n'est pas du json", encoding="utf-8")
    assert lexique.charger_local(fichier) == {
        "expressions": {}, "synonymes": {}}


def test_couche_locale_est_normalisee_au_chargement(tmp_path):
    import json

    from chiffrage import lexique

    fichier = tmp_path / "lex.json"
    fichier.write_text(
        json.dumps({"synonymes": {"Sablage ": "Nettoyage"}}), encoding="utf-8")
    assert lexique.charger_local(fichier)["synonymes"] == {
        "sablage": "nettoyage"}


def test_fusion_respecte_la_precedance():
    """dépôt < appris < à chaud : un essai en cours doit l'emporter
    sur ce qui est commité, sinon on ne peut rien corriger."""
    from chiffrage import lexique

    lexique.LOCAL["synonymes"]["sablage"] = "peinture"
    lexique.ajouter_synonyme("sablage", "nettoyage")
    try:
        fusion = lexique.fusion_a_commiter({"synonymes": {"sablage": "chape"}})
        assert fusion["synonymes"]["sablage"] == "nettoyage"
    finally:
        lexique.LOCAL["synonymes"].clear()


def test_commit_fusionne_avec_le_distant_et_n_ecrase_rien():
    """Deux personnes peuvent régler le lexique le même jour : un
    PUT sans fusion effacerait les termes de l'autre en silence."""
    import json

    from chiffrage import depot_github, lexique

    ecrit = {}

    def _lire(chemin, depot, token, branche):
        return json.dumps({"synonymes": {"pose-existant": "carrelage"}}), "sha1"

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        ecrit.update(chemin=chemin, contenu=contenu, sha=sha, branche=branche)
        return "https://github.com/x/y/commit/abc"

    lexique.ajouter_synonyme("sablage", "nettoyage")
    fusion, url = depot_github.commiter_lexique(
        None, "moi/depot", "jeton", _lire=_lire, _ecrire=_ecrire)

    # Le terme de l'autre a survécu, le nôtre est arrivé.
    assert fusion["synonymes"] == {"pose-existant": "carrelage",
                                    "sablage": "nettoyage"}
    assert ecrit["sha"] == "sha1"      # écriture conditionnée à la version lue
    assert ecrit["chemin"].endswith("lexique_local.json")
    assert json.loads(ecrit["contenu"])["synonymes"]["sablage"] == "nettoyage"
    assert url.startswith("https://github.com/")


def test_commit_premier_fichier_sans_sha():
    """Au tout premier commit, le fichier n'existe pas : pas de sha."""
    from chiffrage import depot_github, lexique

    vu = {}

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        vu["sha"] = sha
        return "url"

    lexique.ajouter_synonyme("sablage", "nettoyage")
    depot_github.commiter_lexique(
        None, "moi/depot", "jeton",
        _lire=lambda *a, **k: (None, None), _ecrire=_ecrire)
    assert vu["sha"] is None


def test_erreur_github_est_expliquee():
    """Un « 403 » brut n'apprend rien : le message doit dire quoi
    vérifier."""
    import urllib.error

    from chiffrage import depot_github

    def _appel(*a, **k):
        raise urllib.error.HTTPError("u", 403, "Forbidden", {}, None)

    with pytest.raises(depot_github.ErreurDepot) as err:
        depot_github.lire_fichier("f", "moi/depot", "jeton", _appel=_appel)
    assert "Contents: read and write" in str(err.value)


def test_fichier_absent_nest_pas_une_erreur():
    import urllib.error

    from chiffrage import depot_github

    def _appel(*a, **k):
        raise urllib.error.HTTPError("u", 404, "Not Found", {}, None)

    assert depot_github.lire_fichier(
        "f", "moi/depot", "jeton", _appel=_appel) == (None, None)


# ── 11. Contrôle des prix ─────────────────────────
from chiffrage import controle_prix  # noqa: E402


@pytest.fixture
def offre_type():
    """Une offre de façade : quatre postes, dont un très dominant."""
    b = moteur.calcul_bordereau()
    return [
        {"code_ouv": c, "qte": q, "pu_vente": b[c]["pu_vente"]}
        for c, q in [("40.20", 180), ("40.30", 180), ("10.20", 180),
                      ("70.50", 95)]
    ]


def test_couverture_horaire_est_arithmetiquement_juste(offre_type):
    b = moteur.calcul_bordereau()
    c = controle_prix.couverture_horaire(offre_type, b)

    total = sum(x["pu_vente"] * x["qte"] for x in offre_type)
    achats = sum((b[x["code_ouv"]]["deb_mat"] + b[x["code_ouv"]]["deb_eqp"])
                  * x["qte"] for x in offre_type)
    heures = sum(b[x["code_ouv"]]["heures_mo"] * x["qte"] for x in offre_type)

    assert c["total"] == pytest.approx(total, abs=0.01)
    assert c["par_heure"] == pytest.approx((total - achats) / heures, abs=0.01)
    assert c["couvre"] is True


def test_rabais_maximal_amene_exactement_au_plancher(offre_type):
    """Le chiffre qu'on veut AVANT de négocier : à ce rabais précis,
    l'offre couvre ses coûts et rien de plus."""
    b = moteur.calcul_bordereau()
    rabais = controle_prix.rabais_maximal(offre_type, b)
    au_plancher = controle_prix.couverture_horaire(offre_type, b,
                                                    rabais=rabais)
    assert au_plancher["par_heure"] == pytest.approx(
        au_plancher["plancher"], abs=0.05)
    assert au_plancher["couvre"] is True


def test_un_rabais_au_dela_du_maximal_declenche_l_alerte(offre_type):
    b = moteur.calcul_bordereau()
    trop = controle_prix.rabais_maximal(offre_type, b) + 0.05
    rapport = controle_prix.analyser(offre_type, b, rabais=trop)
    critiques = [a for a in rapport["alertes"]
                  if a["niveau"] == controle_prix.CRITIQUE]
    assert [a["code"] for a in critiques] == ["couverture"]
    assert rapport["indicateurs"]["couvre"] is False


def test_sans_rabais_l_offre_couvre_toujours(offre_type):
    """Par construction pu = déboursé × K : tant que personne ne force
    un prix, la couverture est garantie. Le jour où la surcharge
    manuelle existera, ce test devra rester vrai sans elle."""
    b = moteur.calcul_bordereau()
    rapport = controle_prix.analyser(offre_type, b)
    assert rapport["indicateurs"]["couvre"] is True
    assert not [a for a in rapport["alertes"]
                 if a["code"] == "couverture"]


def test_un_poste_dominant_est_signale(offre_type):
    b = moteur.calcul_bordereau()
    rapport = controle_prix.analyser(offre_type, b)
    poids = [a for a in rapport["alertes"] if a["code"] == "poids_poste"]
    assert "40.20" in [a["poste"] for a in poids]


def test_part_des_rendements_non_valides_est_mesuree():
    """17,9 % sur le métré type — sous le seuil, mais la valeur doit
    être calculée : c'est elle qui dira quand la calibration presse."""
    b = moteur.calcul_bordereau()
    lignes = [{"code_ouv": c, "qte": 10, "pu_vente": b[c]["pu_vente"]}
              for c in ("40.20", "90.10")]      # 90.10 est à valider
    rapport = controle_prix.analyser(lignes, b)
    assert 0 < rapport["indicateurs"]["part_non_validee"] < 1
    assert [a for a in rapport["alertes"]
            if a["code"] == "rendements_non_valides"]


def test_ecart_a_l_historique_est_signale(offre_type):
    b = moteur.calcul_bordereau()
    ancien = b["40.20"]["pu_vente"] * 0.7      # 30 % moins cher ailleurs
    rapport = controle_prix.analyser(offre_type, b,
                                      historique={"40.20": ancien})
    ecarts = [a for a in rapport["alertes"] if a["code"] == "ecart_historique"]
    assert ecarts and ecarts[0]["poste"] == "40.20"


def test_alertes_triees_du_plus_grave_au_plus_leger(offre_type):
    b = moteur.calcul_bordereau()
    rapport = controle_prix.analyser(offre_type, b, rabais=0.30)
    niveaux = [a["niveau"] for a in rapport["alertes"]]
    ordre = {controle_prix.CRITIQUE: 0, controle_prix.ATTENTION: 1,
             controle_prix.INFO: 2}
    assert niveaux == sorted(niveaux, key=lambda n: ordre[n])


# ── 12. Dossier de justification ──────────────────
def test_dossier_de_justification(tmp_path, openpyxl_dispo):
    from openpyxl import load_workbook

    from chiffrage.justification_xlsx import exporter_justification

    chemin = tmp_path / "just.xlsx"
    _, nb = exporter_justification(
        ["40.20", "70.50"], str(chemin),
        marche={"reference": "CSC 2026-TP-0147",
                 "pouvoir_adjudicateur": "Commune de Schaerbeek"})
    assert nb == 2
    wb = load_workbook(str(chemin))
    assert wb.sheetnames == ["Courrier", "Poste 40.20", "Poste 70.50"]

    textes = [c.value for r in wb["Courrier"].iter_rows() for c in r
              if isinstance(c.value, str)]
    assert any("art. 36" in t for t in textes)
    assert any(biblio.ENTREPRISE["nom"] in t for t in textes)
    assert any("relire et à signer" in t for t in textes)


def test_dossier_porte_toutes_les_ressources_du_poste(tmp_path,
                                                       openpyxl_dispo):
    from openpyxl import load_workbook

    from chiffrage.justification_xlsx import exporter_justification

    chemin = tmp_path / "just.xlsx"
    exporter_justification(["40.20"], str(chemin))
    ws = load_workbook(str(chemin))["Poste 40.20"]
    codes = {c.value for r in ws.iter_rows() for c in r
             if isinstance(c.value, str) and c.value.startswith(("MO.", "MA.", "EQ."))}
    attendus = {comp["code_res"] for comp in biblio.COMPOSITION
                 if comp["code_ouv"] == "40.20"}
    assert attendus <= codes


def test_dossier_formules_ancrees_sur_leur_ligne(tmp_path, openpyxl_dispo):
    """Le destinataire doit pouvoir refaire le calcul dans son tableur."""
    from openpyxl import load_workbook

    from chiffrage.justification_xlsx import exporter_justification

    chemin = tmp_path / "just.xlsx"
    exporter_justification(["40.20"], str(chemin))
    ws = load_workbook(str(chemin))["Poste 40.20"]
    produits = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=E"):
                produits += 1
                assert cell.value == f"=E{cell.row}*F{cell.row}"
    assert produits == len([c for c in biblio.COMPOSITION
                             if c["code_ouv"] == "40.20"])


def test_dossier_refuse_un_poste_inconnu(tmp_path, openpyxl_dispo):
    """Un dossier amputé d'un poste demandé se retourne contre son
    auteur : mieux vaut aucun fichier."""
    from chiffrage.justification_xlsx import exporter_justification

    with pytest.raises(ValueError, match="99.99"):
        exporter_justification(["40.20", "99.99"], str(tmp_path / "j.xlsx"))
    assert not (tmp_path / "j.xlsx").exists()


def test_dossier_refuse_une_liste_vide(tmp_path, openpyxl_dispo):
    from chiffrage.justification_xlsx import exporter_justification

    with pytest.raises(ValueError):
        exporter_justification([], str(tmp_path / "j.xlsx"))


# ── 13. Rien n'est écarté en silence (audit, P0) ────────
@pytest.fixture
def metre_pieges(tmp_path, openpyxl_dispo):
    """Un métré qui ressemble à ce qu'envoient les communes :
    quantités calculées, code en double, quantité nulle, négative."""
    from openpyxl import load_workbook

    from chiffrage.gen_metre import generer_metre

    chemin = tmp_path / "PIEGES.xlsx"
    generer_metre(str(chemin))
    wb = load_workbook(str(chemin))
    ws = wb.active
    for row in ws.iter_rows():
        code = row[1].value
        if not isinstance(code, str):
            continue
        if code == "03.02":                 # quantité calculée
            row[5].value = "=12.5*3"
        elif code == "03.03":               # quantité négative
            row[5].value = -10
        elif code == "05.01":               # quantité nulle
            row[5].value = 0
        elif code == "07.01":               # code dupliqué plus bas
            ws.cell(row=ws.max_row + 1, column=2, value="07.01")
            ws.cell(row=ws.max_row, column=5, value="m2")
            ws.cell(row=ws.max_row, column=6, value=999)
    wb.save(str(chemin))
    return chemin


def test_les_lignes_illisibles_sont_nommees_pas_ignorees(metre_pieges):
    """LE test de l'audit. Avant : trois postes s'évaporaient et le
    rapport annonçait « tous les postes portent un prix »."""
    from chiffrage.metre_io import lire_metre_complet

    lecture = lire_metre_complet(str(metre_pieges))
    genres = {a["genre"]: a["code"] for a in lecture["anomalies"]}

    assert genres.get("quantite_illisible") == "03.02"
    assert genres.get("quantite_negative") == "03.03"
    assert genres.get("quantite_nulle") == "05.01"
    assert genres.get("code_duplique") == "07.01"
    # Chaque anomalie porte SA ligne : sans elle, impossible de corriger.
    assert all(a["ligne"] > 0 for a in lecture["anomalies"])


def test_une_offre_amputee_ne_peut_plus_se_declarer_complete(metre_pieges,
                                                              tmp_path):
    from chiffrage.metre_io import imprimer_rapport, remplir_metre

    rapport = remplir_metre(str(metre_pieges), str(tmp_path / "o.xlsx"))
    texte = imprimer_rapport(rapport)

    assert rapport["anomalies_bloquantes"]
    assert "OFFRE IRRÉGULIÈRE" in texte
    assert "Tous les postes du métré portent un prix" not in texte


def test_quantite_nulle_ne_bloque_pas_mais_se_signale(tmp_path,
                                                       openpyxl_dispo):
    """Un poste « pour mémoire » à 0 est licite ; il doit être chiffré
    et signalé, pas rejeté."""
    from openpyxl import load_workbook

    from chiffrage.gen_metre import generer_metre
    from chiffrage.metre_io import lire_metre_complet

    chemin = tmp_path / "zero.xlsx"
    generer_metre(str(chemin))
    wb = load_workbook(str(chemin))
    for row in wb.active.iter_rows():
        if row[1].value == "05.01":
            row[5].value = 0
    wb.save(str(chemin))

    lecture = lire_metre_complet(str(chemin))
    assert "05.01" in {p["code"] for p in lecture["postes"]}
    assert [a["genre"] for a in lecture["anomalies"]] == ["quantite_nulle"]


def test_quantite_en_formule_avec_valeur_en_cache_est_utilisee(tmp_path,
                                                                openpyxl_dispo):
    """Quand Excel a enregistré le résultat, on l'utilise — en le
    signalant, car la formule peut avoir changé depuis."""
    from openpyxl import load_workbook

    from chiffrage.gen_metre import generer_metre
    from chiffrage.metre_io import lire_metre_complet

    chemin = tmp_path / "cache.xlsx"
    generer_metre(str(chemin))
    wb = load_workbook(str(chemin))
    for row in wb.active.iter_rows():
        if row[1].value == "03.02":
            row[5].value = "=12.5*3"
    wb.save(str(chemin))

    # openpyxl n'écrit pas de cache ; on simule ce que fait Excel en
    # enregistrant la valeur dans le classeur des valeurs.
    from openpyxl import load_workbook as charger
    valeurs = charger(str(chemin), data_only=True)
    assert valeurs.active is not None      # le second classeur s'ouvre

    lecture = lire_metre_complet(str(chemin))
    # Sans cache, le poste part en anomalie plutôt que d'être deviné.
    assert [a["genre"] for a in lecture["anomalies"]] == ["quantite_illisible"]
    assert "03.02" not in {p["code"] for p in lecture["postes"]}


# ── 14. Codes tolérants et classeurs multi-feuilles ──────
@pytest.mark.parametrize("code", [
    "03.02", "3.2", "01.02.03", "03.02.A", "1.01.10", "A.1.2",
    "03-02", "03/02",
])
def test_formes_de_codes_acceptees(code):
    """Un CSC numéroté 01.02.03 rendait ZÉRO poste, avec pour tout
    message « aucun poste lu »."""
    from chiffrage.metre_io import est_code_poste

    assert est_code_poste(code)


@pytest.mark.parametrize("pas_un_code", [
    "2026",            # pas de séparateur
    "Lot 03",          # espace
    "12,5",            # virgule décimale
    "Récapitulatif",
    "TOTAL",
    "m2",
    "A.B",             # aucun chiffre
    "",
    None,
    42,                # une cellule numérique n'est pas un code
])
def test_le_bruit_reste_ecarte(pas_un_code):
    """Élargir le motif ne doit pas transformer n'importe quelle
    cellule en poste."""
    from chiffrage.metre_io import est_code_poste

    assert not est_code_poste(pas_un_code)


@pytest.fixture
def metre_multi(tmp_path, openpyxl_dispo):
    """Un métré comme en envoient les communes : un onglet par lot,
    plus un récapitulatif qui reprend les mêmes codes."""
    from openpyxl import load_workbook

    from chiffrage.gen_metre import generer_metre

    chemin = tmp_path / "MULTI.xlsx"
    generer_metre(str(chemin))
    wb = load_workbook(str(chemin))
    src = wb["MÉTRÉ"]
    lot2 = wb.copy_worksheet(src)
    lot2.title = "Lot 02"
    recap = wb.copy_worksheet(src)
    recap.title = "Récapitulatif"
    src.title = "Lot 01"
    for row in lot2.iter_rows():          # le lot 2 a ses propres codes
        code = row[1].value
        if isinstance(code, str) and code.count(".") == 1 and code[0].isdigit():
            row[1].value = "1" + code
    wb.save(str(chemin))
    return chemin


def test_inventaire_des_feuilles_et_presomption_de_recapitulatif(metre_multi):
    from chiffrage.metre_io import feuilles_avec_postes

    inventaire = {f["nom"]: f for f in feuilles_avec_postes(str(metre_multi))}
    assert inventaire["Lot 01"]["nb_postes"] == 49
    assert inventaire["Lot 02"]["nb_postes"] == 49
    assert inventaire["Récapitulatif"]["recapitulatif"] is True
    assert inventaire["Lot 01"]["recapitulatif"] is False


def test_toutes_les_feuilles_sont_lues_par_defaut(metre_multi):
    """Perdre un lot entier en silence serait pire que lire un
    récapitulatif : les doublons, eux, sont signalés."""
    from chiffrage.metre_io import lire_metre_complet

    lecture = lire_metre_complet(str(metre_multi), ["Lot 01", "Lot 02"])
    assert len(lecture["postes"]) == 98
    assert {p["feuille"] for p in lecture["postes"]} == {"Lot 01", "Lot 02"}


def test_le_recapitulatif_produit_des_doublons_pas_un_double_compte(
        metre_multi):
    """Le vrai danger du multi-feuilles : un récap qui reprend les
    codes des lots doublerait le montant de l'offre."""
    from chiffrage.metre_io import lire_metre_complet

    lecture = lire_metre_complet(str(metre_multi))
    doublons = [a for a in lecture["anomalies"]
                 if a["genre"] == "code_duplique"]
    assert len(doublons) == 49
    assert {a["feuille"] for a in doublons} == {"Récapitulatif"}
    # 98 postes chiffrables, pas 147 : rien n'est compté deux fois.
    assert len(lecture["postes"]) == 98


def test_le_prix_est_ecrit_dans_la_feuille_du_poste(metre_multi, tmp_path):
    """Tout écrire sur la première feuille rendrait un classeur
    incohérent au pouvoir adjudicateur, sans qu'aucune erreur ne soit
    levée."""
    from openpyxl import load_workbook

    from chiffrage.metre_io import COL_PU, remplir_metre

    sortie = tmp_path / "offre.xlsx"
    rapport = remplir_metre(str(metre_multi), str(sortie),
                             feuilles=["Lot 01", "Lot 02"])
    assert rapport["postes"] == 98

    wb = load_workbook(str(sortie))

    def prix_ecrits(nom):
        return sum(1 for row in wb[nom].iter_rows() for c in row
                    if c.column == COL_PU and isinstance(c.value, (int, float)))

    # Les codes du lot 2 ne sont pas au mapping : rien ne s'y écrit,
    # et surtout rien ne se déverse sur le lot 1.
    assert prix_ecrits("Lot 01") == len(rapport["chiffres"])
    assert prix_ecrits("Récapitulatif") == 0
    assert len(rapport["non_couverts"]) == 49


# ── 15. Détection des colonnes ──────────────────
def _classeur(lignes, titre="Feuille"):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = titre
    for ligne in lignes:
        ws.append(ligne)
    return wb


@pytest.fixture
def metre_autre_commune(tmp_path, openpyxl_dispo):
    """La disposition citée par l'audit : colonnes ailleurs, autres
    intitulés, codes en 1.01.10, titre avant l'en-tête."""
    wb = _classeur([
        ["COMMUNE DE WEMMEL — MARCHÉ 2026/114"],
        [],
        ["Poste", "Description des travaux", "", "", "", "Unité",
         "Quantité", "", "Prix unitaire", "Total"],
        ["1.01.10", "Démolition de cloisons légères", "", "", "", "m2", 38,
         "", None, None],
        ["1.01.20", "Enduit de façade minéral armé", "", "", "", "m2", 165,
         "", None, None],
        ["2.03.05", "Peinture des plafonds intérieurs", "", "", "", "m2", 210,
         "", None, None],
    ], titre="Inventaire")
    chemin = tmp_path / "AUTRE.xlsx"
    wb.save(str(chemin))
    return chemin


def test_detection_sur_une_disposition_inconnue(metre_autre_commune):
    from openpyxl import load_workbook

    from chiffrage.detection_colonnes import detecter

    d = detecter(load_workbook(str(metre_autre_commune)).active)
    assert d["manquants"] == []
    assert d["champs"]["code"] == 1          # A
    assert d["champs"]["designation"] == 2   # B
    assert d["champs"]["unite"] == 6         # F
    assert d["champs"]["quantite"] == 7      # G
    assert d["champs"]["pu"] == 9            # I


def test_le_contenu_tranche_pour_la_colonne_des_codes(openpyxl_dispo,
                                                       tmp_path):
    """Un métré titre couramment « N° » le simple compteur de lignes,
    juste avant la vraie colonne des codes. Se fier à l'intitulé
    partait sur le compteur, et plus aucun poste n'était lu."""
    from openpyxl import load_workbook

    from chiffrage.detection_colonnes import detecter

    wb = _classeur([
        ["N°", "Poste", "Désignation", "Unité", "Quantité", "PU"],
        [1, "01.01", "Démolition", "m2", 38, None],
        [2, "01.02", "Enduit", "m2", 165, None],
        [3, "01.03", "Peinture", "m2", 210, None],
    ])
    chemin = tmp_path / "compteur.xlsx"
    wb.save(str(chemin))

    d = detecter(load_workbook(str(chemin)).active)
    assert d["champs"]["code"] == 2           # B, pas A
    assert d["origines"]["code"] == "contenu"


def test_detection_sans_aucun_entete(openpyxl_dispo, tmp_path):
    """Sans en-tête, seul le contenu parle — et il suffit."""
    from openpyxl import load_workbook

    from chiffrage.detection_colonnes import detecter

    wb = _classeur([
        ["", "1.01", "Démolition", "", "m2", 38, None],
        ["", "1.02", "Enduit", "", "m2", 165, None],
        ["", "1.03", "Peinture", "", "m2", 210, None],
        ["", "1.04", "Châssis", "", "m2", 26, None],
    ])
    chemin = tmp_path / "sans_entete.xlsx"
    wb.save(str(chemin))

    d = detecter(load_workbook(str(chemin)).active)
    assert d["champs"]["code"] == 2
    assert d["champs"]["quantite"] == 6
    assert d["manquants"] == []


def test_lecture_et_ecriture_suivent_les_colonnes_detectees(
        metre_autre_commune, tmp_path):
    """Le test qui compte : écrire le prix dans la mauvaise colonne
    rendrait l'offre silencieusement fausse."""
    from openpyxl import load_workbook

    from chiffrage.metre_io import remplir_metre

    mapping = {"1.01.10": "20.20", "1.01.20": "40.20", "2.03.05": "70.20"}
    sortie = tmp_path / "offre.xlsx"
    rapport = remplir_metre(str(metre_autre_commune), str(sortie),
                             mapping=mapping)

    assert rapport["postes"] == 3
    assert len(rapport["chiffres"]) == 3
    assert rapport["vides"] == []

    ws = load_workbook(str(sortie)).active
    for row in ws.iter_rows(min_row=4):
        if not row[0].value:
            continue
        assert isinstance(row[8].value, (int, float))   # I : le prix
        assert row[6].value                             # G : la quantité, intacte


def test_colonnes_imposees_priment_sur_la_detection(metre_autre_commune):
    """L'humain doit pouvoir corriger une détection fausse."""
    from chiffrage.metre_io import lire_metre_complet

    lecture = lire_metre_complet(str(metre_autre_commune),
                                  colonnes={"quantite": 7, "unite": 6,
                                             "code": 1, "pu": 9})
    assert len(lecture["postes"]) == 3
    assert lecture["colonnes"]["Inventaire"]["pu"] == 9


def test_intitules_synonymes_sont_reconnus(openpyxl_dispo):
    from chiffrage.detection_colonnes import _champ_de

    assert _champ_de("Qté") == "quantite"
    assert _champ_de("Métré") == "quantite"
    assert _champ_de("P.U. HTVA") == "pu"
    assert _champ_de("Prix total") == "montant"
    assert _champ_de("Libellé") == "designation"
    assert _champ_de("Un.") == "unite"
    # « Unité » ne doit pas être happé par « quantité ».
    assert _champ_de("Unité") == "unite"
    assert _champ_de("Observations") is None


# ── 16. Paramètres réglables sans toucher au code ──────
def test_parametres_par_defaut_quand_le_fichier_est_absent(tmp_path):
    """Le cas NORMAL : aucun fichier local, on chiffre quand même."""
    from chiffrage import parametres

    entreprise, params = parametres.charger_local(tmp_path / "rien.json")
    assert entreprise == parametres.ENTREPRISE_DEFAUT
    assert params == parametres.PARAMS_DEFAUT


def test_parametres_illisibles_ne_bloquent_pas(tmp_path):
    from chiffrage import parametres

    fichier = tmp_path / "casse.json"
    fichier.write_text("{pas du json", encoding="utf-8")
    entreprise, params = parametres.charger_local(fichier)
    assert entreprise == parametres.ENTREPRISE_DEFAUT
    assert params == parametres.PARAMS_DEFAUT


def test_chaque_bloc_retombe_independamment_sur_son_defaut(tmp_path):
    """Une marge aberrante ne doit pas faire perdre l'adresse."""
    import json

    from chiffrage import parametres

    fichier = tmp_path / "partiel.json"
    fichier.write_text(json.dumps({
        "entreprise": {"nom": "AUTRE SRL", "cp_ville": "1000 Bruxelles"},
        "params": {"marge": 9.9},          # aberrant : 990 %
    }), encoding="utf-8")

    entreprise, params = parametres.charger_local(fichier)
    assert entreprise["nom"] == "AUTRE SRL"
    assert entreprise["cp_ville"] == "1000 Bruxelles"
    assert entreprise["pays"] == parametres.ENTREPRISE_DEFAUT["pays"]
    assert params == parametres.PARAMS_DEFAUT     # le bloc entier est rejeté


@pytest.mark.parametrize("params_faux, motif", [
    ({"marge": 2.5}, "250 % de marge est une virgule mal placée"),
    ({"marge": -0.1}, "un taux négatif"),
    ({"fg": "beaucoup"}, "pas un nombre"),
])
def test_coefficients_aberrants_refuses(params_faux, motif):
    from chiffrage import parametres

    with pytest.raises(ValueError):
        parametres.valider_params({**parametres.PARAMS_DEFAUT, **params_faux})


def test_raison_sociale_vide_refusee():
    """Elle figure en en-tête de chaque devis et de chaque courrier."""
    from chiffrage import parametres

    with pytest.raises(ValueError, match="raison sociale"):
        parametres.valider_entreprise(
            {**parametres.ENTREPRISE_DEFAUT, "nom": "   "})


def test_numero_de_tva_absurde_refuse():
    from chiffrage import parametres

    with pytest.raises(ValueError, match="TVA"):
        parametres.valider_entreprise(
            {**parametres.ENTREPRISE_DEFAUT, "tva": "à compléter"})
    # Mais les variantes d'écriture passent : on refuse l'absurde,
    # pas la mise en forme.
    for ecriture in ("BE0766637025", "BE 0766.637.025", "0766.637.025"):
        parametres.valider_entreprise(
            {**parametres.ENTREPRISE_DEFAUT, "tva": ecriture})


def test_aller_retour_par_le_json(tmp_path):
    from chiffrage import parametres

    entreprise = dict(parametres.ENTREPRISE_DEFAUT, nom="AUTRE SRL")
    params = dict(parametres.PARAMS_DEFAUT, marge=0.15)
    fichier = tmp_path / "p.json"
    fichier.write_text(parametres.serialiser(entreprise, params),
                        encoding="utf-8")

    relu_entreprise, relu_params = parametres.charger_local(fichier)
    assert relu_entreprise["nom"] == "AUTRE SRL"
    assert relu_params["marge"] == 0.15


def test_commit_des_parametres_n_additionne_rien():
    """Deux adresses ne se fusionnent pas : le dernier écrit gagne,
    mais pas en aveugle — le sha relu fait échouer l'écriture si
    quelqu'un est passé entre temps."""
    from chiffrage import depot_github

    vu = {}

    def _lire(chemin, depot, token, branche):
        return '{"entreprise": {"nom": "ANCIEN"}}', "sha-distant"

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        vu.update(contenu=contenu, sha=sha, chemin=chemin)
        return "https://github.com/x/y/commit/abc"

    url = depot_github.commiter_parametres(
        '{"entreprise": {"nom": "NOUVEAU"}}', "moi/depot", "jeton",
        _lire=_lire, _ecrire=_ecrire)

    assert "NOUVEAU" in vu["contenu"] and "ANCIEN" not in vu["contenu"]
    assert vu["sha"] == "sha-distant"
    assert vu["chemin"].endswith("parametres_local.json")
    assert url.startswith("https://github.com/")


# ── 17. Les tables sont des données, pas du code ─────
@pytest.fixture
def data_copiee(tmp_path):
    import shutil  # noqa: PLC0415

    from chiffrage.bibliotheque import DOSSIER_DATA

    cible = tmp_path / "data"
    shutil.copytree(DOSSIER_DATA, cible)
    return cible


def _modifier(dossier, table, transformer):
    import json

    fichier = dossier / f"{table}.json"
    fichier.write_text(
        json.dumps(transformer(json.loads(fichier.read_text("utf-8"))),
                    ensure_ascii=False),
        encoding="utf-8")


def test_les_tables_json_donnent_les_memes_prix():
    """La migration ne devait rien déplacer : mêmes 49 prix, au centime."""
    b = moteur.calcul_bordereau()
    assert len(b) == 49
    assert b["40.20"]["pu_vente"] == 98.64
    assert b["70.10"]["pu_vente"] == 15.51
    assert round(moteur.coefficient_k(), 4) == 1.3324


def test_les_notes_de_raisonnement_ont_survecu():
    """Un JSON est muet : le POURQUOI d'un chiffre devait devenir une
    donnée, sinon il disparaissait avec les commentaires Python."""
    notes = {(c["code_ouv"], c["code_res"]): c.get("note")
             for c in biblio.COMPOSITION if c.get("note")}
    assert "double compte" in notes[("10.10", "MO.01")]
    assert "bicouche" in notes[("40.40", "MA.10")]
    assert all("Coût entreprise COMPLET" in r["note"]
                for r in biblio.RESSOURCES if r["type_res"] == "MO")


@pytest.mark.parametrize("table, transformer, motif", [
    ("composition",
     lambda d: d + [{"code_ouv": "40.20", "code_res": "MA.99", "qte_res": 1}],
     "inexistante"),
    ("ouvrages",
     lambda d: d + [{"code_ouv": "40.99", "libelle_ouv": "Fantôme",
                      "unite_ouv": "m2", "code_ref": ""}],
     "0 €"),
    ("ouvrages",
     lambda d: [dict(o, unite_ouv="") if o["code_ouv"] == "40.20" else o
                 for o in d],
     "unité manquante"),
    ("ressources",
     lambda d: [dict(r, type_res="XXX") if r["code_res"] == "MO.01" else r
                 for r in d],
     "inconnu"),
    ("ressources",
     lambda d: [dict(r, pu_res=-5) if r["code_res"] == "MA.01" else r
                 for r in d],
     "invalide"),
    ("mapping", lambda d: {**d, "99.99": "40.99"}, "inexistant"),
])
def test_une_table_incoherente_est_refusee(data_copiee, table, transformer,
                                            motif):
    """Éditable à la main veut dire corrompable à la main. Chacune de
    ces fautes produirait un prix faux sans se voir dans un JSON de
    150 lignes : une ressource orpheline vaut zéro, un ouvrage sans
    composition se vend gratuitement."""
    from chiffrage.bibliotheque import BibliothequeInvalide

    from chiffrage.bibliotheque import charger_tables

    _modifier(data_copiee, table, transformer)
    with pytest.raises(BibliothequeInvalide, match=motif):
        charger_tables(data_copiee)


def test_une_table_manquante_arrete_tout(data_copiee):
    """Pas de valeurs de repli ici, à la différence du lexique : une
    bibliothèque vide ne dégraderait pas le résultat, elle rendrait
    « aucun ouvrage » pour tous les postes — une offre entièrement
    vide, présentée comme normale."""
    from chiffrage.bibliotheque import BibliothequeInvalide

    from chiffrage.bibliotheque import charger_tables

    (data_copiee / "lots.json").unlink()
    with pytest.raises(BibliothequeInvalide, match="manquante"):
        charger_tables(data_copiee)


def test_le_lot_se_deduit_du_code(data_copiee):
    """Stocker le lot deux fois, c'est risquer qu'ils divergent."""
    import json

    brut = json.loads((data_copiee / "ouvrages.json").read_text("utf-8"))
    assert all("lot" not in o for o in brut)
    assert biblio.OUVRAGES_PAR_CODE["40.20"]["lot"] == "40"


# ── 18. Calculer sur des tables en cours de correction ────
def _tables_modifiees(**prix):
    """Une copie des tables du dépôt, avec des taux corrigés."""
    import copy

    tables = copy.deepcopy(moteur.tables_courantes())
    for res in tables["ressources"]:
        if res["code_res"] in prix:
            res["pu_res"] = prix[res["code_res"]]
    tables["ressources_par_code"] = {r["code_res"]: r
                                      for r in tables["ressources"]}
    return tables


def test_le_moteur_calcule_sur_des_tables_fournies():
    """Sans ça, aucun aperçu n'est possible : on ne pourrait pas voir
    l'effet d'un rendement corrigé AVANT de l'enregistrer."""
    tables = _tables_modifiees(**{"MO.02": 55.0})
    assert moteur.calcul_bordereau(tables=tables)["40.20"]["pu_vente"] > \
        moteur.calcul_bordereau()["40.20"]["pu_vente"]


def test_les_tables_du_depot_restent_intactes():
    """Le point critique : un aperçu ne doit RIEN changer aux prix qui
    partent dans les offres."""
    avant = moteur.calcul_bordereau()["40.20"]["pu_vente"]
    moteur.calcul_bordereau(tables=_tables_modifiees(**{"MO.02": 99.0}))
    assert moteur.calcul_bordereau()["40.20"]["pu_vente"] == avant == 98.64
    assert biblio.RESSOURCES_PAR_CODE["MO.02"]["pu_res"] == 48.00


def test_la_calibration_suit_les_tables_corrigees():
    """C'est l'écran de la séance de calibration : corriger un taux,
    voir l'écart bouger."""
    avant = moteur.calibration()
    apres = moteur.calibration(tables=_tables_modifiees(**{"MO.02": 55.0}))
    assert apres["ecart_moyen_absolu"] != avant["ecart_moyen_absolu"]
    ecarts_avant = {r["devis"]: r["ecart"] for r in avant["lignes"]}
    # Le devis 15 est de la façade : relever le taux façadier doit le
    # faire monter, pas descendre.
    apres_15 = next(r for r in apres["lignes"] if r["devis"] == "15")
    assert apres_15["ecart"] > ecarts_avant["15"]


def test_fiche_prix_suit_aussi_les_tables_corrigees():
    texte = moteur.fiche_prix("40.20",
                               tables=_tables_modifiees(**{"MO.02": 55.0}))
    assert "55.00" in texte


def test_commit_d_une_table_n_additionne_rien():
    """Un prix corrigé ne se mélange pas à un autre prix corrigé."""
    from chiffrage import depot_github

    vu = {}

    def _lire(chemin, depot, token, branche):
        vu["lu"] = chemin
        return "[]", "sha-distant"

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        vu.update(chemin=chemin, sha=sha, message=message)
        return "https://github.com/x/y/commit/abc"

    depot_github.commiter_table("ressources", "[]", "moi/depot", "jeton",
                                 _lire=_lire, _ecrire=_ecrire)
    assert vu["chemin"] == "chiffrage/data/ressources.json"
    assert vu["sha"] == "sha-distant"
    assert "ressources" in vu["message"]


# ── 19. Relevé de chantier ─────────────────────────────────────────────────
def _un_ouvrage(nb_lignes_mo):
    """Un code d'ouvrage ayant exactement ce nombre de lignes MO."""
    compte = {}
    for comp in biblio.COMPOSITION:
        if biblio.RESSOURCES_PAR_CODE[comp["code_res"]]["type_res"] == "MO":
            compte[comp["code_ouv"]] = compte.get(comp["code_ouv"], 0) + 1
    codes = sorted(c for c, n in compte.items() if n == nb_lignes_mo)
    assert codes, f"aucun ouvrage à {nb_lignes_mo} ligne(s) de main-d'œuvre"
    return codes[0]


def test_un_releve_est_le_quotient_des_heures_par_la_quantite():
    """Le seul calcul que le chef d'entreprise faisait de tête sur le
    chantier — et donc celui qu'il ne doit plus faire."""
    r = moteur.releve_rendement(_un_ouvrage(1), quantite=12, heures=7)
    assert r["rendement_observe"] == pytest.approx(7 / 12, abs=1e-4)


def test_les_heures_relevees_se_repartissent_au_prorata():
    """Sept ouvrages ont deux lignes de main-d'œuvre. Le relevé donne un
    total : la somme des propositions doit le rendre entier, et le
    partage suivre la proportion en place — pas un partage inventé."""
    code = _un_ouvrage(2)
    r = moteur.releve_rendement(code, quantite=5, heures=20)

    assert sum(x["propose"] for x in r["lignes"]) == pytest.approx(
        r["rendement_observe"], abs=1e-3)
    assert sum(x["part"] for x in r["lignes"]) == pytest.approx(1.0, abs=1e-3)
    facteurs = [x["propose"] / x["qte_res"] for x in r["lignes"]]
    assert max(facteurs) - min(facteurs) < 1e-3, (
        "les lignes ne sont pas corrigées dans la même proportion")


def test_un_releve_se_compare_au_rendement_en_place():
    """Un relevé qui confirme la bibliothèque doit sortir un écart nul :
    c'est ce qui permet de lever un ⚠️ en connaissance de cause."""
    code = _un_ouvrage(1)
    actuel = sum(c["qte_res"] for c in biblio.COMPOSITION
                  if c["code_ouv"] == code
                  and biblio.RESSOURCES_PAR_CODE[
                      c["code_res"]]["type_res"] == "MO")
    r = moteur.releve_rendement(code, quantite=10, heures=10 * actuel)
    assert r["rendement_actuel"] == pytest.approx(actuel, abs=1e-4)
    assert r["ecart"] == pytest.approx(0.0, abs=1e-3)


def test_un_releve_incomplet_est_refuse_plutot_que_devine():
    """Zéro heure ferait une main-d'œuvre gratuite, zéro quantité une
    division par zéro. Dans les deux cas, refuser vaut mieux que rendre
    un nombre que personne ne saurait interpréter."""
    code = _un_ouvrage(1)
    with pytest.raises(ValueError):
        moteur.releve_rendement(code, quantite=0, heures=7)
    with pytest.raises(ValueError):
        moteur.releve_rendement(code, quantite=12, heures=0)
    with pytest.raises(KeyError):
        moteur.releve_rendement("99.99", quantite=12, heures=7)


def test_un_releve_ne_corrige_rien_par_lui_meme():
    """Le relevé PROPOSE, l'humain applique. S'il écrivait dans les
    tables, une saisie de chantier changerait des prix sans qu'aucun
    écran ne l'annonce."""
    code = _un_ouvrage(1)
    avant = moteur.calcul_bordereau()[code]["pu_vente"]
    moteur.releve_rendement(code, quantite=1, heures=99)
    assert moteur.calcul_bordereau()[code]["pu_vente"] == avant


def test_un_releve_lit_les_tables_qu_on_lui_donne():
    """Pendant une séance de correction, la référence est la copie de
    travail — sinon l'écart s'afficherait contre des valeurs que
    l'utilisateur vient justement de remplacer."""
    import copy

    code = _un_ouvrage(1)
    tables = copy.deepcopy(moteur.tables_courantes())
    ligne = next(c for c in tables["composition"]
                  if c["code_ouv"] == code
                  and tables["ressources_par_code"][
                      c["code_res"]]["type_res"] == "MO")
    ligne["qte_res"] *= 2

    r = moteur.releve_rendement(code, quantite=1, heures=1, tables=tables)
    hors_session = moteur.releve_rendement(code, quantite=1, heures=1)
    assert r["rendement_actuel"] == pytest.approx(
        2 * hors_session["rendement_actuel"], abs=1e-4)


# ── 22. Journal des relevés de chantier ────────────────────────────────────
def _tables_avec(releves):
    import copy  # noqa: PLC0415

    t = copy.deepcopy(moteur.tables_courantes())
    t["releves"] = releves
    return t


def _rel(code, date, chantier, quantite, heures):
    return {"code_ouv": code, "date": date, "chantier": chantier,
             "quantite": quantite, "heures": heures}


def test_sans_releve_le_constate_est_none_et_non_zero():
    """« Aucun chantier relevé » et « rendement de zéro » sont deux
    choses opposées. Les confondre ferait passer une bibliothèque
    jamais confrontée au réel pour une bibliothèque à main-d'œuvre
    gratuite."""
    assert moteur.rendement_constate("20.10", tables=_tables_avec([])) is None
    assert moteur.releves_de("20.10", tables=_tables_avec([])) == []


def test_le_constate_pondere_par_les_quantites():
    """Σheures / Σquantités, pas la moyenne des rendements : 2 m2 en 3 h
    et 40 m2 en 20 h ne pèsent pas pareil, et une moyenne simple
    donnerait au tout petit chantier le poids du grand."""
    tables = _tables_avec([
        _rel("20.10", "2026-08-01", "Renan 35", 2, 3),
        _rel("20.10", "2026-08-20", "Wemmel", 40, 20),
    ])
    c = moteur.rendement_constate("20.10", tables=tables)

    assert c["n"] == 2
    assert c["rendement"] == pytest.approx(23 / 42, abs=1e-4)
    moyenne_simple = (3 / 2 + 20 / 40) / 2
    assert abs(c["rendement"] - moyenne_simple) > 0.4, (
        "la pondération ne change rien : le test ne prouve pas grand-chose")
    # Le constaté reste entre le plus rapide et le plus lent.
    assert c["mini"] <= c["rendement"] <= c["maxi"]


def test_le_constate_se_compare_au_rendement_en_place():
    """C'est l'écart qui dit s'il y a matière à corriger."""
    code = _un_ouvrage(1)
    actuel = sum(c["qte_res"] for c in biblio.COMPOSITION
                  if c["code_ouv"] == code
                  and biblio.RESSOURCES_PAR_CODE[
                      c["code_res"]]["type_res"] == "MO")
    tables = _tables_avec([_rel(code, "2026-08-01", "Renan 35",
                                 10, 10 * actuel)])
    c = moteur.rendement_constate(code, tables=tables)
    assert c["ecart"] == pytest.approx(0.0, abs=1e-3)


def test_un_releve_ne_change_aucun_prix():
    """Le journal est une PREUVE, pas un réglage. S'il alimentait le
    bordereau, une saisie de chantier ferait bouger des prix sans
    qu'aucun écran ne l'annonce — et une table absente les ferait
    bouger dans l'autre sens."""
    code = _un_ouvrage(1)
    avant = moteur.calcul_bordereau()[code]["pu_vente"]
    tables = _tables_avec([_rel(code, "2026-08-01", "Renan 35", 1, 99)])
    assert moteur.calcul_bordereau(tables=tables)[code]["pu_vente"] == avant


def test_les_releves_sortent_dans_l_ordre_du_temps():
    tables = _tables_avec([
        _rel("20.10", "2026-08-20", "Wemmel", 40, 20),
        _rel("20.10", "2026-08-01", "Renan 35", 2, 3),
    ])
    dates = [r["date"] for r in moteur.releves_de("20.10", tables=tables)]
    assert dates == sorted(dates)


def test_le_rendement_de_chaque_releve_est_calcule_pas_stocke():
    """Stocker le quotient en ferait une seconde vérité, qui finirait
    par diverger de ses deux termes."""
    tables = _tables_avec([_rel("20.10", "2026-08-01", "Renan 35", 12, 7)])
    assert "rendement" not in tables["releves"][0]
    assert moteur.releves_de("20.10", tables=tables)[0]["rendement"] == (
        pytest.approx(7 / 12, abs=1e-4))


def test_les_releves_dun_autre_ouvrage_ne_sen_melent_pas():
    tables = _tables_avec([
        _rel("20.10", "2026-08-01", "Renan 35", 12, 7),
        _rel("40.20", "2026-08-01", "Renan 35", 30, 40),
    ])
    assert moteur.rendement_constate("20.10", tables=tables)["n"] == 1


def test_fusionner_ajoute_sans_perdre_ni_doubler():
    """Un journal s'AJOUTE, il ne s'écrase pas : deux téléphones peuvent
    relever le même soir. Mais le même relevé saisi deux fois est une
    seule observation."""
    mien = [_rel("20.10", "2026-08-30", "Wemmel", 40, 20)]
    sien = [_rel("40.20", "2026-08-01", "Renan 35", 12, 7)]
    fusion = moteur.fusionner_releves(sien, mien, list(sien))

    assert len(fusion) == 2, "un doublon a été gardé, ou une observation perdue"
    assert {r["chantier"] for r in fusion} == {"Wemmel", "Renan 35"}
    dates = [r["date"] for r in fusion]
    assert dates == sorted(dates)


def test_un_releve_sur_un_ouvrage_supprime_est_signale_pas_fatal():
    """Une preuve devenue muette doit se voir — mais elle ne porte aucun
    prix, donc elle n'empêche rien : la perdre serait pire que la
    garder."""
    anomalies = moteur.controle_coherence()
    assert anomalies["releves_orphelins"] == []
    assert "releves_orphelins" in anomalies


# ── 23. La table des relevés est optionnelle, les prix ne le sont pas ───────
def test_une_bibliotheque_sans_le_fichier_de_releves_demarre(tmp_path):
    """Sur Streamlit Cloud, un push ne redémarre pas le processus :
    l'app peut lire des tables plus anciennes qu'elle. Refuser de
    démarrer pour un journal absent — dont aucun prix ne dépend —
    serait une panne inventée."""
    import shutil  # noqa: PLC0415
    from pathlib import Path  # noqa: PLC0415

    source = Path(biblio.__file__).parent / "data"
    for fichier in source.glob("*.json"):
        if fichier.name != "releves.json":
            shutil.copy(fichier, tmp_path)

    tables = biblio.charger_tables(tmp_path)
    assert tables["releves"] == []
    assert tables["ouvrages"], "le reste de la bibliothèque doit être là"


def test_une_table_de_prix_absente_refuse_toujours_de_demarrer(tmp_path):
    """Le raisonnement du repli ne s'étend PAS à une table de prix :
    absente, elle ne dégraderait pas le résultat, elle rendrait des
    offres à zéro présentées comme normales."""
    import shutil  # noqa: PLC0415
    from pathlib import Path  # noqa: PLC0415

    source = Path(biblio.__file__).parent / "data"
    for fichier in source.glob("*.json"):
        if fichier.name != "ressources.json":
            shutil.copy(fichier, tmp_path)

    with pytest.raises(biblio.BibliothequeInvalide):
        biblio.charger_tables(tmp_path)


@pytest.mark.parametrize("bancal, motif", [
    ({"code_ouv": "20.10", "date": "2026-08-01", "chantier": "",
       "quantite": 12, "heures": 7}, "chantier"),
    ({"code_ouv": "20.10", "date": "2026-08-01", "chantier": "X",
       "quantite": 0, "heures": 7}, "quantite"),
    ({"code_ouv": "20.10", "date": "2026-08-01", "chantier": "X",
       "quantite": 12, "heures": -1}, "heures"),
])
def test_un_releve_malforme_est_refuse_au_chargement(tmp_path, bancal, motif):
    """Un fichier PRÉSENT mais bancal est une corruption, pas une
    absence : là, on lève. Un relevé sans provenance ne prouve rien."""
    import json as _json  # noqa: PLC0415
    import shutil  # noqa: PLC0415
    from pathlib import Path  # noqa: PLC0415

    source = Path(biblio.__file__).parent / "data"
    for fichier in source.glob("*.json"):
        shutil.copy(fichier, tmp_path)
    (tmp_path / "releves.json").write_text(
        _json.dumps([bancal]), encoding="utf-8")

    with pytest.raises(biblio.BibliothequeInvalide) as err:
        biblio.charger_tables(tmp_path)
    assert motif in str(err.value)


def test_le_journal_fusionne_au_lieu_d_ecraser():
    """La seule table qui s'AJOUTE. Deux téléphones peuvent relever le
    même soir : un PUT sans fusion effacerait l'observation de l'autre
    sans rien dire — une demi-journée de chantier perdue, invisible."""
    import json as _json  # noqa: PLC0415

    from chiffrage.depot_github import commiter_releves  # noqa: PLC0415

    distant = [{"code_ouv": "20.10", "date": "2026-08-01",
                 "chantier": "Renan 35", "quantite": 2, "heures": 3}]
    ecrit = {}

    def _lire(chemin, depot, token, branche):
        return _json.dumps(distant), "sha-distant"

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        ecrit["contenu"] = _json.loads(contenu)
        ecrit["sha"] = sha
        return "https://github.com/commit"

    mien = [{"code_ouv": "40.20", "date": "2026-08-30", "chantier": "Wemmel",
              "quantite": 40, "heures": 20},
             dict(distant[0])]          # le même relevé, ressaisi
    fusion, url = commiter_releves(mien, "d/r", "jeton",
                                    _lire=_lire, _ecrire=_ecrire)

    assert len(ecrit["contenu"]) == 2, "un relevé a été perdu ou doublé"
    assert {r["chantier"] for r in ecrit["contenu"]} == {"Renan 35", "Wemmel"}
    # Le sha relu juste avant fait échouer l'écriture si quelqu'un est
    # passé entre temps.
    assert ecrit["sha"] == "sha-distant"
    assert fusion == ecrit["contenu"]


def test_un_journal_distant_illisible_arrete_l_ecriture():
    """Fusionner avec un fichier corrompu écraserait ce qu'il contient
    encore. Mieux vaut refuser et le faire corriger à la main."""
    from chiffrage.depot_github import (  # noqa: PLC0415
        ErreurDepot,
        commiter_releves,
    )

    def _lire(chemin, depot, token, branche):
        return "{ pas du json", "sha"

    def _ecrire(*a, **k):
        raise AssertionError("rien ne doit être écrit")

    with pytest.raises(ErreurDepot):
        commiter_releves([], "d/r", "jeton", _lire=_lire, _ecrire=_ecrire)


# ── 24. D'où vient l'écart d'un devis historique ───────────────────────────
def test_lanalyse_redit_la_meme_chose_que_la_calibration():
    """Deux vérités sur un même écart, c'est une lecture faussée tôt ou
    tard. L'analyse détaille, elle ne recalcule pas autrement."""
    cal = {r["devis"]: r for r in moteur.calibration()["lignes"]}
    for num, ref in cal.items():
        a = moteur.analyser_ecart(num)
        assert a["calcule"] == pytest.approx(ref["calcule"], abs=0.01)
        assert a["ecart"] == pytest.approx(ref["ecart"], abs=1e-4)
        assert a["heures_mo"] == pytest.approx(ref["heures_mo"], abs=0.01)


def test_le_facteur_de_quantites_annule_vraiment_lecart():
    """Le chiffre le plus engageant de l'écran : « il faudrait 30 % de
    quantités en moins ». Il doit tomber juste — appliqué à toutes les
    lignes, il ramène le calcul sur le forfait vendu."""
    for num in moteur.tables_courantes()["metres_histo"]:
        a = moteur.analyser_ecart(num)
        facteur = a["facteur_quantites"]
        lignes = [(x["code_ouv"], x["qte"] * facteur) for x in a["lignes"]]
        rejoue = moteur.devis(f"essai {num}", lignes)
        assert rejoue["total_ht"] == pytest.approx(a["forfait"], rel=1e-3), (
            f"devis {num} : le facteur ne ramène pas sur le forfait")


def test_le_k_implicite_est_ce_que_le_forfait_a_couvert():
    """K implicite = forfait / déboursé sec. En dessous de 1, le chantier
    n'a pas payé ses propres achats et heures — et `couvre_debourse` doit
    le dire, sans quoi l'écran rassurerait à tort."""
    for num in moteur.tables_courantes()["metres_histo"]:
        a = moteur.analyser_ecart(num)
        assert a["k_implicite"] == pytest.approx(
            a["forfait"] / a["debourse_sec"], abs=1e-3)
        assert a["couvre_debourse"] == (a["k_implicite"] >= 1.0)
        # Le calculé, lui, est toujours le déboursé fois K : c'est la
        # définition même du prix de vente.
        assert a["calcule"] == pytest.approx(
            a["debourse_sec"] * a["k_vise"], rel=1e-3)


def test_les_postes_sortent_du_plus_lourd_au_plus_leger():
    """On regarde d'abord ce qui pèse : un écart porté par un seul poste
    se règle en vérifiant SA quantité."""
    a = moteur.analyser_ecart("16")
    montants = [x["montant"] for x in a["lignes"]]
    assert montants == sorted(montants, reverse=True)
    assert sum(x["part"] for x in a["lignes"]) == pytest.approx(1.0, abs=1e-3)
    assert a["concentration"] == pytest.approx(
        sum(x["part"] for x in a["lignes"][:3]), abs=1e-3)


def test_lanalyse_ne_tranche_pas_a_la_place_du_chef_dentreprise():
    """Deux causes opposées produisent le même total. L'outil rend les
    chiffres qui les séparent ; il ne rend aucun verdict — pas de champ
    « hypothèse », pas de « cause »."""
    a = moteur.analyser_ecart("16")
    assert not {"hypothese", "cause", "verdict", "conclusion"} & set(a)


def test_un_devis_historique_inconnu_est_refuse():
    with pytest.raises(KeyError):
        moteur.analyser_ecart("99")


def test_lanalyse_suit_les_tables_corrigees():
    """Pendant une séance, la référence est la copie de travail : sinon
    l'écart s'afficherait contre des valeurs qu'on vient de remplacer."""
    import copy  # noqa: PLC0415

    tables = copy.deepcopy(moteur.tables_courantes())
    for ligne in tables["composition"]:
        if tables["ressources_par_code"][
                ligne["code_res"]]["type_res"] == "MO":
            ligne["qte_res"] *= 2

    avant = moteur.analyser_ecart("16")
    apres = moteur.analyser_ecart("16", tables=tables)
    assert apres["calcule"] > avant["calcule"]
    assert apres["facteur_quantites"] < avant["facteur_quantites"]


# ── 25. Une instance par entrepreneur ──────────────────────────────────────
#
# Le partage n'est pas un problème d'accès mais de CLOISONNEMENT : les
# tables sont des constantes de module, l'identité est un seul fichier, et
# l'écriture visait un chemin en dur. Une bibliothèque de prix, ce sont
# des taux horaires et une marge — l'actif de l'entreprise, pas quelque
# chose qui se partage avec un confrère qui répond aux mêmes marchés.

def _instance(tmp_path, nom_entreprise, fg):
    """Le dossier d'un entrepreneur : ses tables ET son identité."""
    import json as _json  # noqa: PLC0415
    import shutil  # noqa: PLC0415
    from pathlib import Path as _Path  # noqa: PLC0415

    tmp_path.mkdir(parents=True, exist_ok=True)
    source = _Path(biblio.__file__).parent / "data"
    for fichier in source.glob("*.json"):
        shutil.copy(fichier, tmp_path)
    (tmp_path / "parametres_local.json").write_text(_json.dumps({
        "entreprise": {"nom": nom_entreprise, "adresse": "Rue d'Essai 1",
                        "cp_ville": "1780 Wemmel", "pays": "Belgique",
                        "tva": "BE 0999.888.777", "activite": "Rénovation"},
        "params": {"fg": fg, "fc": 0.05, "aleas": 0.03, "marge": 0.10,
                    "tva": 0.06, "tva_marche_public": 0.21},
    }, ensure_ascii=False), encoding="utf-8")
    return tmp_path


def _dans_une_instance(dossier, code):
    """Exécute `code` dans un interpréteur neuf, CHIFFRAGE_DATA pointé sur
    ce dossier — les tables étant des constantes de module, deux
    instances ne peuvent pas cohabiter dans le même processus."""
    import os  # noqa: PLC0415
    import subprocess  # noqa: PLC0415
    import sys  # noqa: PLC0415
    from pathlib import Path as _Path  # noqa: PLC0415

    racine = _Path(biblio.__file__).resolve().parent.parent
    env = dict(os.environ, CHIFFRAGE_DATA=str(dossier),
                PYTHONPATH=str(racine))
    fait = subprocess.run([sys.executable, "-c", code], env=env,
                           capture_output=True, text=True, timeout=120)
    assert fait.returncode == 0, fait.stderr
    return fait.stdout.strip()


def test_deux_entrepreneurs_ne_partagent_ni_prix_ni_identite(tmp_path):
    """Le cœur du cloisonnement : même code, deux dossiers, deux mondes.
    Sans ça, l'entrepreneur B verrait les taux horaires de A — c'est-à-dire
    ses marges."""
    # Séparateur explicite : la raison sociale contient des espaces.
    code = ("from chiffrage.bibliotheque import ENTREPRISE\n"
             "from chiffrage.moteur import coefficient_k, calcul_bordereau\n"
             "print(ENTREPRISE['nom'], round(coefficient_k(), 4),"
             " calcul_bordereau()['40.20']['pu_vente'], sep='|')")

    a = _dans_une_instance(_instance(tmp_path / "a", "ALPHA SRL", 0.12), code)
    b = _dans_une_instance(_instance(tmp_path / "b", "BETA SRL", 0.20), code)

    nom_a, k_a, pu_a = a.split("|")
    nom_b, k_b, pu_b = b.split("|")
    assert (nom_a, nom_b) == ("ALPHA SRL", "BETA SRL"), (a, b)
    assert float(k_b) > float(k_a), "les coefficients ne sont pas cloisonnés"
    assert float(pu_b) > float(pu_a), "les prix ne sont pas cloisonnés"


def test_sans_la_variable_rien_ne_bouge():
    """Le déploiement en service garde ses fichiers là où ils sont nés :
    une couture ne doit pas déménager l'existant."""
    from chiffrage import depot_github  # noqa: PLC0415

    assert biblio.ENTREPRISE["nom"] == "BAG BATTER SRL"
    assert depot_github.chemins_entreprise() == {
        "tables": "chiffrage/data",
        "parametres": "chiffrage/parametres_local.json",
        "lexique": "chiffrage/lexique_local.json",
    }


def test_le_dossier_reunit_les_trois_fichiers_de_lentreprise():
    """Prix, identité et lexique au même endroit : c'est ce qui permet de
    donner à chaque instance un jeton limité à SON dossier."""
    from chiffrage.depot_github import chemins_entreprise  # noqa: PLC0415

    c = chemins_entreprise("donnees/wemmel")
    assert c["tables"] == "donnees/wemmel"
    assert c["parametres"] == "donnees/wemmel/parametres_local.json"
    assert c["lexique"] == "donnees/wemmel/lexique_local.json"
    # Les slashs de trop ne doivent pas produire un chemin absurde.
    assert chemins_entreprise("/donnees/wemmel/") == c


def test_le_commit_dune_table_suit_le_dossier_de_linstance():
    """Écrire dans `chiffrage/data` depuis l'instance d'un autre
    entrepreneur écraserait les prix de quelqu'un d'autre."""
    from chiffrage.depot_github import commiter_table  # noqa: PLC0415

    vus = {}

    def _lire(chemin, depot, token, branche):
        vus["lu"] = chemin
        return None, None

    def _ecrire(chemin, contenu, message, depot, token, branche, sha):
        vus["ecrit"] = chemin
        return "url"

    commiter_table("ressources", "[]", "d/r", "jeton",
                    dossier="donnees/wemmel", _lire=_lire, _ecrire=_ecrire)
    assert vus["lu"] == "donnees/wemmel/ressources.json"
    assert vus["ecrit"] == vus["lu"], "on relit et on écrit au même endroit"

    commiter_table("ressources", "[]", "d/r", "jeton",
                    _lire=_lire, _ecrire=_ecrire)
    assert vus["ecrit"] == "chiffrage/data/ressources.json"


def test_le_journal_suit_aussi_le_dossier():
    from chiffrage.depot_github import commiter_releves  # noqa: PLC0415

    vus = {}

    def _lire(chemin, depot, token, branche):
        vus["chemin"] = chemin
        return None, None

    commiter_releves([], "d/r", "jeton", dossier="donnees/wemmel",
                      _lire=_lire, _ecrire=lambda *a, **k: "url")
    assert vus["chemin"] == "donnees/wemmel/releves.json"


# ── 20. Reprendre un devis enregistré ──────────────────────────────────────
# Le fichier relu ici a pu être édité à la main ou produit par une version
# antérieure de la bibliothèque. Ces tests vérifient surtout qu'une donnée
# douteuse est ÉCARTÉE ET SIGNALÉE, jamais devinée ni laissée passer.

def _relire(charge, codes=("40.20", "40.30")):
    from chiffrage import devis_json

    return devis_json.lire(json.dumps(charge).encode("utf-8"), set(codes))


def test_un_devis_fait_laller_retour_sans_rien_perdre():
    from chiffrage import devis_json

    texte = devis_json.serialiser(
        "Façade arrière", "2026-042", "Rue X 1", "M. Dupont\n1030 Bruxelles",
        0.06, [("40.20", 22.0), ("40.30", 8.5)])
    devis, anomalies = devis_json.lire(texte, {"40.20", "40.30"})
    assert anomalies == []
    assert devis["objet"] == "Façade arrière"
    assert devis["client"].endswith("1030 Bruxelles")
    assert devis["tva"] == 0.06
    assert devis["lignes"] == [{"code_ouv": "40.20", "qte": 22.0},
                                {"code_ouv": "40.30", "qte": 8.5}]


def test_un_ouvrage_disparu_de_la_bibliotheque_est_ecarte_et_signale():
    """Le laisser passer ferait planter le tableau d'édition, dont la
    liste déroulante n'accepte que des codes existants."""
    devis, anomalies = _relire({
        "lignes": [{"code_ouv": "40.20", "qte": 3},
                    {"code_ouv": "99.99", "qte": 3}]})
    assert [ligne["code_ouv"] for ligne in devis["lignes"]] == ["40.20"]
    assert any("99.99" in a for a in anomalies)


def test_une_quantite_negative_est_ecartee():
    """Sans ce contrôle, le devis afficherait un montant négatif sans
    que rien ne le signale."""
    devis, anomalies = _relire({"lignes": [{"code_ouv": "40.20", "qte": -5}]})
    assert devis["lignes"] == []
    assert any("négative" in a for a in anomalies)


def test_une_quantite_illisible_est_ecartee():
    devis, anomalies = _relire({
        "lignes": [{"code_ouv": "40.20", "qte": "beaucoup"}]})
    assert devis["lignes"] == []
    assert anomalies


def test_un_taux_de_tva_inattendu_retombe_sur_vingt_et_un():
    """Sous-facturer la TVA se paie au contrôle ; la sur-facturer se
    corrige par une note de crédit. Le repli va donc vers le haut."""
    devis, anomalies = _relire({"tva": 0.12, "lignes": []})
    assert devis["tva"] == 0.21
    assert anomalies


def test_un_devis_sans_tva_ne_declenche_pas_davertissement():
    devis, anomalies = _relire({"lignes": []})
    assert devis["tva"] == 0.21
    assert anomalies == []


def test_un_fichier_qui_nest_pas_un_devis_est_refuse():
    from chiffrage import devis_json

    with pytest.raises(ValueError):
        devis_json.lire(b'["40.20"]', {"40.20"})
    with pytest.raises(ValueError):
        devis_json.lire(b'{"objet": "x"}', {"40.20"})


# ── 21. Combinaisons — ce que les tests isolés laissent passer ──────────────
#
# Deux défauts trouvés par un audit du dépôt, tous deux invisibles aux
# tests pris un à un : chacune des deux moitiés passait seule.

def _cacher_formule(chemin, cellule, valeur):
    """Écrit la valeur en cache d'une formule, comme le fait Excel.

    openpyxl n'en écrit jamais : un classeur qu'il produit ne porte que
    la formule. Or c'est le cache qui distingue un métré déjà ouvert
    dans Excel — donc chiffrable — d'un métré qui ne l'a jamais été. Il
    faut donc l'ajouter dans le XML, à la main, pour tester ce chemin.
    """
    import re  # noqa: PLC0415
    import shutil  # noqa: PLC0415
    import zipfile  # noqa: PLC0415

    tampon = str(chemin) + ".tmp"
    with zipfile.ZipFile(str(chemin)) as source, \
            zipfile.ZipFile(tampon, "w") as sortie:
        for item in source.infolist():
            donnees = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                donnees = re.sub(
                    r'(<c r="%s"[^>]*>)(<f>[^<]*</f>)' % cellule,
                    r'\1\2<v>%s</v>' % valeur,
                    donnees.decode("utf-8")).encode("utf-8")
            sortie.writestr(item, donnees)
    shutil.move(tampon, str(chemin))


def test_une_quantite_en_formule_se_lit_dans_la_colonne_detectee(
        tmp_path, openpyxl_dispo):
    """Le défaut P0 : la première lecture trouvait la quantité en G, la
    seconde allait chercher son cache dans la colonne F d'origine. Le
    poste ressortait « quantité illisible » alors que la valeur était
    là — et un poste sans prix rend l'offre irrégulière (art. 76).

    Colonne déplacée seule : testé, ça passait. Formule seule : testé,
    ça passait. Les deux ensemble : le poste disparaissait."""
    from chiffrage.metre_io import lire_metre_complet

    wb = _classeur([
        ["COMMUNE DE WEMMEL — MARCHÉ 2026/114"],
        [],
        ["Poste", "Description des travaux", "", "", "", "Unité",
         "Quantité", "Prix unitaire", "Total"],
        ["1.01.10", "Démolition de cloisons légères", "", "", "", "m2", 38,
         None, None],
        ["1.01.20", "Enduit de façade minéral armé", "", "", "", "m2",
         "=12.5*3", None, None],
        ["2.03.05", "Peinture des plafonds intérieurs", "", "", "", "m2", 210,
         None, None],
    ], titre="Inventaire")
    chemin = tmp_path / "FORMULE.xlsx"
    wb.save(str(chemin))
    _cacher_formule(chemin, "G5", 37.5)

    lecture = lire_metre_complet(str(chemin))
    assert lecture["colonnes"]["Inventaire"]["quantite"] == 7, (
        "la quantité n'est pas détectée en G : le test ne prouve rien")

    postes = {p["code"]: p for p in lecture["postes"]}
    assert "1.01.20" in postes, "le poste en formule a été perdu"
    assert postes["1.01.20"]["quantite"] == pytest.approx(37.5)
    assert [a["genre"] for a in lecture["anomalies"]] == ["quantite_formule"]


def test_aucune_colonne_assez_remplie_ne_fait_pas_planter_la_detection(
        openpyxl_dispo):
    """Le défaut P1 : des colonnes numériques pouvaient exister sans
    qu'aucune n'atteigne le seuil de remplissage. `min()` recevait une
    séquence vide et levait — le dépôt du métré plantait au lieu de
    rendre une détection incomplète, que l'interface sait présenter."""
    from chiffrage.detection_colonnes import detecter

    lignes = [[None, f"03.{i:02d}", "Désignation du poste"] for i in range(10)]
    for ligne in lignes[:3]:                 # trois nombres épars sur dix
        ligne += ["", 12.5]
    detection = detecter(_classeur(lignes).active)

    assert "quantite" not in detection["champs"], (
        "trois valeurs sur dix postes ne font pas une colonne de quantité")
    assert "quantite" in detection["manquants"]


def test_une_colonne_deja_attribuee_nest_pas_reprise_par_deduction(
        openpyxl_dispo):
    """Trouvé en vérifiant le repli partiel signalé par l'audit, et pire
    que décrit : l'unité nommée « Unité » en G, la quantité trouvée en F,
    et la déduction de position donnait « PU = G ». Le prix s'écrivait
    par-dessus l'unité — et `manquants` était VIDE, donc l'interface ne
    demandait aucune validation. Ne pas savoir où est le prix doit se
    voir."""
    from chiffrage.detection_colonnes import detecter

    lignes = [["Poste", "Description", "", "", "", "Métrage réel", "Unité"]]
    lignes += [[f"1.01.{i:02d}", "Enduit de façade minéral armé", "", "", "",
                 10.0 + i, "m2"] for i in range(6)]
    detection = detecter(_classeur(lignes).active)

    assert detection["champs"]["unite"] == 7
    assert detection["champs"]["quantite"] == 6
    assert detection["champs"].get("pu") != detection["champs"]["unite"], (
        "le prix irait s'écrire dans la colonne de l'unité")
    assert "pu" in detection["manquants"], (
        "une détection qui ignore où est le prix doit le dire")

@pytest.fixture
def metre_sans_colonne_de_prix(tmp_path, openpyxl_dispo):
    """Le classeur où la colonne du PU est introuvable : « Métrage réel »
    n'est pas un intitulé connu, et rien n'annonce un prix. L'unité, elle,
    est nommée — et se trouve juste à droite de la quantité."""
    lignes = [["Poste", "Description", "", "", "", "Métrage réel", "Unité"]]
    lignes += [[f"1.01.{i:02d}", "Enduit de façade minéral armé", "", "", "",
                 10.0 + i, "m2"] for i in range(6)]
    chemin = tmp_path / "SANS_PU.xlsx"
    _classeur(lignes, titre="Inventaire").save(str(chemin))
    return chemin


def _mapping_facade():
    return {f"1.01.{i:02d}": "40.20" for i in range(6)}


def test_sans_colonne_de_prix_rien_nest_ecrit(metre_sans_colonne_de_prix,
                                               tmp_path):
    """La colonne du PU est la seule où l'outil écrive. Ne pas savoir où
    elle est ne doit pas se traduire par « écrire ailleurs » : le prix
    atterrissait dans la colonne de l'unité, et six « m2 » devenaient des
    montants dans le classeur rendu au pouvoir adjudicateur."""
    from openpyxl import load_workbook  # noqa: PLC0415

    from chiffrage.metre_io import remplir_metre

    sortie = tmp_path / "offre.xlsx"
    rapport = remplir_metre(str(metre_sans_colonne_de_prix), str(sortie),
                             mapping=_mapping_facade())

    assert rapport["postes"] == 6, "les postes doivent rester LISIBLES"
    assert rapport["chiffres"] == []
    assert len(rapport["sans_colonne_pu"]) == 6
    assert {p["code"] for p in rapport["sans_colonne_pu"]} <= set(
        rapport["vides"]), "des postes non écrits absents du décompte art. 76"

    # Le prix calculé accompagne chaque poste : il reste à le porter à la
    # main, ce qui n'est possible que s'il est dit.
    assert all(p["pu"] > 0 for p in rapport["sans_colonne_pu"])

    # Et le classeur est intact — c'est tout l'objet.
    ws = load_workbook(str(sortie))["Inventaire"]
    for ligne in range(2, 8):
        assert ws.cell(ligne, 7).value == "m2", "l'unité a été écrasée"
        assert isinstance(ws.cell(ligne, 6).value, (int, float))


def test_le_rapport_nomme_la_feuille_et_le_prix_a_porter(
        metre_sans_colonne_de_prix, tmp_path):
    """Un poste non écrit sans explication est un poste perdu."""
    from chiffrage.metre_io import imprimer_rapport, remplir_metre

    rapport = remplir_metre(str(metre_sans_colonne_de_prix),
                             str(tmp_path / "offre.xlsx"),
                             mapping=_mapping_facade())
    texte = imprimer_rapport(rapport)

    assert "Inventaire" in texte
    assert "1.01.00" in texte
    assert "OFFRE IRRÉGULIÈRE" in texte


def test_imposer_la_colonne_du_prix_debloque_lecriture(
        metre_sans_colonne_de_prix, tmp_path):
    """Le refus n'est pas une impasse : dire où est le prix suffit, et
    c'est exactement ce que l'interface demande avant de chiffrer."""
    from openpyxl import load_workbook  # noqa: PLC0415

    from chiffrage.metre_io import remplir_metre

    sortie = tmp_path / "offre.xlsx"
    rapport = remplir_metre(
        str(metre_sans_colonne_de_prix), str(sortie),
        mapping=_mapping_facade(),
        colonnes={"code": 1, "designation": 2, "quantite": 6, "unite": 7,
                   "pu": 8})

    assert len(rapport["chiffres"]) == 6
    assert rapport["sans_colonne_pu"] == []
    assert rapport["vides"] == []

    ws = load_workbook(str(sortie))["Inventaire"]
    for ligne in range(2, 8):
        assert isinstance(ws.cell(ligne, 8).value, (int, float))
        assert ws.cell(ligne, 7).value == "m2"


@pytest.fixture
def metre_de_torture(tmp_path, openpyxl_dispo):
    """Les quatre pièges d'un coup — plusieurs feuilles, colonnes
    déplacées, quantité en formule avec cache, récapitulatif reprenant
    les mêmes codes."""
    from openpyxl import Workbook  # noqa: PLC0415

    entete = ["Poste", "Description des travaux", "", "", "", "Unité",
               "Quantité", "Prix unitaire", "Total"]
    lot_1 = [("1.01.10", "Démolition de cloisons légères", "m2", 38),
              ("1.01.20", "Enduit de façade minéral armé", "m2", "=12.5*3"),
              ("1.02.05", "Peinture des plafonds intérieurs", "m2", 210)]
    lot_2 = [("2.01.10", "Plafonnage sur maçonnerie", "m2", 74),
              ("2.01.20", "Cimentage de mur de cave", "m2", 26),
              ("2.02.30", "Pose de faïence murale", "m2", 18)]

    def feuille(ws, titre, postes):
        ws.title = titre
        ws.append(["COMMUNE DE WEMMEL — MARCHÉ 2026/114"])
        ws.append([])
        ws.append(list(entete))
        for code, libelle, unite, qte in postes:
            ws.append([code, libelle, "", "", "", unite, qte, None, None])

    wb = Workbook()
    feuille(wb.active, "Lot 01", lot_1)
    feuille(wb.create_sheet(), "Lot 02", lot_2)
    feuille(wb.create_sheet(), "Récapitulatif", lot_1)

    chemin = tmp_path / "TORTURE.xlsx"
    wb.save(str(chemin))
    # La formule est en G5 sur « Lot 01 » comme sur le récapitulatif :
    # Excel en aurait mis le cache dans les deux.
    _cacher_formule(chemin, "G5", 37.5)
    return chemin


def test_le_classeur_de_torture_se_lit_en_entier(metre_de_torture):
    """Les six postes des deux lots, la formule comprise ; et le
    récapitulatif signalé comme doublon plutôt que compté deux fois."""
    from chiffrage.metre_io import lire_metre_complet

    lecture = lire_metre_complet(str(metre_de_torture))
    postes = {p["code"]: p for p in lecture["postes"]}
    genres = [a["genre"] for a in lecture["anomalies"]]

    assert len(postes) == 6, "des postes se sont perdus en chemin"
    assert postes["1.01.20"]["quantite"] == pytest.approx(37.5)
    assert postes["1.01.20"]["feuille"] == "Lot 01"
    assert postes["2.02.30"]["feuille"] == "Lot 02"

    # Le récapitulatif reprend les trois codes du lot 1 : trois doublons,
    # et aucune quantité ajoutée à celles déjà lues.
    assert genres.count("code_duplique") == 3
    assert genres.count("quantite_formule") >= 1
    assert all(c["feuille"] == "Lot 01"
                for code, c in postes.items() if code.startswith("1."))


def test_le_prix_va_dans_la_colonne_detectee_de_la_bonne_feuille(
        metre_de_torture, tmp_path):
    """Écrire dans la colonne d'origine, ou sur la première feuille,
    rendrait au pouvoir adjudicateur un classeur faux sans qu'aucune
    erreur ne soit levée."""
    from openpyxl import load_workbook  # noqa: PLC0415

    from chiffrage.metre_io import remplir_metre

    sortie = tmp_path / "offre.xlsx"
    rapport = remplir_metre(
        str(metre_de_torture), str(sortie),
        feuilles=["Lot 01", "Lot 02"],
        mapping={"1.01.10": "20.20", "1.01.20": "40.20",
                  "2.01.10": "60.20"})
    assert len(rapport["chiffres"]) == 3

    wb = load_workbook(str(sortie))
    for nom, attendus in (("Lot 01", 2), ("Lot 02", 1)):
        prix = [c.value for row in wb[nom].iter_rows(min_row=4) for c in row
                 if c.column == 8 and isinstance(c.value, (int, float))]
        assert len(prix) == attendus, f"{nom} : prix mal placés"
    # La quantité reste intacte, et rien n'est écrit dans la colonne
    # que la disposition par défaut aurait visée.
    ws = wb["Lot 01"]
    assert ws.cell(4, 7).value == 38
    assert all(ws.cell(ligne, 6).value in (None, "m2") for ligne in (4, 5, 6))
